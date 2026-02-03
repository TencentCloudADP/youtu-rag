"""Knowledge config routes, including tool configuration, building, and validation"""
import logging

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session

from ..database import get_db
from ..models.kb_config import (
    KBConfigurationUpdate,
    KBBuildRequest,
    KBBuildResponse,
    QAValidationResult,
    DBConnectionTestRequest,
    DBConnectionTestResponse,
)
from ..services.kb_config_service import KBConfigService
from ..config import settings

from ....config import ConfigLoader
from ....tools.memory_toolkit import VectorMemoryToolkit


logger = logging.getLogger(__name__)
router = APIRouter()


@router.put("/{kb_id}/configuration")
async def update_kb_configuration(
    kb_id: int,
    config_update: KBConfigurationUpdate,
    db: Session = Depends(get_db)
):
    """Update knowledge base config (tools, files, and connections)
    
    Args:
        kb_id: Knowledge base ID.
        config_update: Update configuration.
        
    Returns:
        Result of the update.
        
    Example:
        ```
        PUT /api/knowledge/{kb_id}/configuration
        {
            "configuration": {
                "tools": {...},
                "selectedFiles": ["doc1.pdf"],
                "selectedQAFiles": ["qa.xlsx"],
                "dbConnections": [...]
            }
        }
        ```
    """
    try:
        tools_config_dict = {}
        for tool_name, tool_config in config_update.configuration.tools.items():
            tools_config_dict[tool_name] = {
                "enabled": tool_config.enabled,
                "settings": tool_config.settings
            }
        
        result = await KBConfigService.update_configuration(
            kb_id=kb_id,
            tools_config=tools_config_dict,
            selected_files=config_update.configuration.selectedFiles,
            selected_qa_files=config_update.configuration.selectedQAFiles,
            db_connections=config_update.configuration.dbConnections,
            db=db
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        db.rollback()
        logger.error(f"Update configuration error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{kb_id}/build", response_model=KBBuildResponse)
async def build_knowledge_base(
    kb_id: int,
    build_request: KBBuildRequest,
    db: Session = Depends(get_db)
):
    """Build/rebuild knowledge base using configured tools and sources.

    The build process:
    1. Load sources from config (MinIO files, database, QA files);
    2. Process sources in parallel using KnowledgeBuilderAgent;
    3. Store vectors in ChromaDB, structured data in SQLite;
    4. Update build status and logs.
    
    Args:
        kb_id: Knowledge base ID.
        build_request: Build options.
        
    Returns:
        Build results.
        
    Example:
        ```
        POST /api/knowledge/{kb_id}/build
        {
            "force_rebuild": false,
            "file_filter": ["doc1.pdf"]
        }
        ```
    """
    try:
        from ..kb_config_routes import build_knowledge_base as _build_impl
        return await _build_impl(kb_id, build_request, db)
    except Exception as e:
        logger.error(f"Build error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/files/validate-qa/{filename}", response_model=QAValidationResult)
async def validate_qa_file(filename: str):
    """Validate the format of a QA Excel file.
    
    Expected format:
    - Sheet name: "example"
    - Headers: "question", "answer", "howtofind"
    
    Args:
        filename: QA Excel name in MinIO
        
    Returns:
        Validation results.
        
    Example:
        ```
        GET /api/knowledge/files/validate-qa/qa_examples.xlsx
        ```
    """
    try:
        result = await KBConfigService.validate_qa_file(filename)
        # ----------------- 新增逻辑 Start: 存储到 Working Memory -----------------
        # 只有当校验成功 (valid=True) 时，才尝试读取数据并存储
        if result.get("valid"):
            try:
                # 注意：Service 没有直接返回所有行的数据，只返回了 sample_data。
                # 所以要实现存储，我们必须在这里再次读取 Excel 文件，或者修改 Service 让它返回所有数据。
                # 这种重复读取在大文件时效率较低，但为了不修改 Service 接口，我们先这样做。
                
                # --- 为了读取数据，我们需要重新下载一次文件 (或者依赖 Service 的修改) ---
                # 但更安全的做法是：既然 Service 返回了 valid=True，我们就利用它。
                # 这里为了简单起见，我们只能复用 Service 里的部分逻辑或者重新读取。
                
                # 最佳实践：这里应该调用另一个 Service 方法 `import_qa_to_memory`，
                # 但既然要在路由层做，我们只能在这里写。
                
                import os
                import io
                import openpyxl
                from ..minio_client import MinIOClient 
                
                # 初始化 MinIO (复用环境变量)
                minio_client = MinIOClient(
                    endpoint=os.getenv("MINIO_ENDPOINT", "localhost:9000"),
                    access_key=os.getenv("MINIO_ACCESS_KEY", "minioadmin"),
                    secret_key=os.getenv("MINIO_SECRET_KEY", "minioadmin"),
                    bucket_name=os.getenv("MINIO_BUCKET", "rag-documents"),
                    secure=os.getenv("MINIO_SECURE", "false").lower() == "true"
                )
                
                file_data = minio_client.download_file(filename)
                if file_data:
                    wb = openpyxl.load_workbook(io.BytesIO(file_data.read()))
                    if "example" in wb.sheetnames:
                        sheet = wb["example"]
                        # 简单的标头定位 (假设已经 validate 过了)
                        headers = [str(c.value).lower().strip() for c in sheet[1]]
                        q_idx = headers.index("question")
                        a_idx = headers.index("answer")
                        h_idx = headers.index("howtofind") if "howtofind" in headers else -1
                        
                        # # 初始化 Memory
                        # agent_config = ConfigLoader.load_agent_config("simple/chat")
                        # memory_toolkit = VectorMemoryToolkit(config=agent_config)
                        memory_toolkit = VectorMemoryToolkit(
                            persist_directory=settings.memory_store_path,
                            collection_prefix="rag_chat",
                            default_user_id="default_user",
                            max_working_memory_turns=10000,
                        )

                        logger.info(f"💾 [Auto-Import] Starting to import {sheet.max_row - 1} rows to memory...")
                        
                        count = 0
                        for row_idx in range(2, sheet.max_row + 1):
                            row = sheet[row_idx]
                            q = row[q_idx].value
                            a = row[a_idx].value
                            h = row[h_idx].value if h_idx >= 0 else None
                            
                            if q and a:
                                lines = [f"Question: {str(q).strip()}", f"Answer: {str(a).strip()}"]
                                if h:
                                    lines.append(f"HowToFind: {str(h).strip()}")
                                    
                                # await memory_toolkit.store_working_memory("\n".join(lines), role="assistant")
                                await memory_toolkit.save_conversation_to_episodic(
                                    question=str(q).strip(),
                                    answer=str(a).strip(),
                                    importance_score=0.5, 
                                )
                                count += 1
                        
                        logger.info(f"✅ [Auto-Import] Successfully imported {count} items.")

            except Exception as e:
                # 仅记录日志，不影响校验结果返回给前端
                logger.error(f"❌ Failed to auto-import QA to memory: {str(e)}")

        
        return QAValidationResult(**result)
    except Exception as e:
        logger.error(f"Validation error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/database/test-connection", response_model=DBConnectionTestResponse)
async def test_database_connection(request: DBConnectionTestRequest):
    """Test database connection and retrieve table list.

    Supporting MySQL and SQLite connections.
    - MySQL: requires host, port, database, username, password
    - SQLite: requires file_path
    
    Returns available table list on success.
    
    Args:
        request: Request for database connection test.
        
    Returns:
        Connection test results.
        
    Example:
        ```
        POST /api/knowledge/database/test-connection
        {
            "db_type": "mysql",
            "host": "localhost",
            "port": 3306,
            "database": "mydb",
            "username": "user",
            "password": "pass"
        }
        ```
    """
    try:
        result = await KBConfigService.test_database_connection(
            db_type=request.db_type,
            host=request.host,
            port=request.port,
            database=request.database,
            username=request.username,
            password=request.password,
            file_path=request.file_path
        )
        return DBConnectionTestResponse(**result)
    except Exception as e:
        logger.error(f"Connection test error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

