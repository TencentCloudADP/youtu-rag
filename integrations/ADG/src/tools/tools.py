"""
Tool Definition Module
Defines various tools used in table processing and Q&A workflows
Contains complete tool chain for construction and retrieval phases
"""

from typing import Any, Dict, List, Optional, Tuple
from abc import ABC, abstractmethod
import pandas as pd
import openpyxl
import networkx as nx
from pathlib import Path
import re


class BaseTool(ABC):
    """Base tool class"""
    
    def __init__(self, name: str, description: str, tool_type: str):
        self.name = name
        self.description = description
        self.tool_type = tool_type
    
    @abstractmethod
    def execute(self, *args, **kwargs) -> Any:
        """Execute tool functionality"""
        pass
    
    def __repr__(self) -> str:
        return f"{self.__class__.__name__}(name='{self.name}', type='{self.tool_type}')"


# ============================================================
# Construction Phase Tools - Phase 1: Construction
# ============================================================

# ============ 1. Excel Loading Tools ============

class ExcelLoader(BaseTool):
    """Excel file loader"""
    
    def __init__(self):
        super().__init__(
            name="ExcelLoader",
            description="Load Excel files (.xlsx format)",
            tool_type="loader"
        )
    
    def execute(self, file_path: str) -> openpyxl.Workbook:
        """
        Load Excel file
        
        Args:
            file_path: Excel file path
            
        Returns:
            openpyxl workbook object
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # 不使用data_only=True以保留合并单元格等格式信息
        workbook = openpyxl.load_workbook(file_path)
        return workbook


class WorkbookReader(BaseTool):
    """Workbook reader"""
    
    def __init__(self):
        super().__init__(
            name="WorkbookReader",
            description="Read all worksheets in workbook",
            tool_type="reader"
        )
    
    def execute(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """
        Read workbook information
        
        Args:
            workbook: openpyxl workbook object
            
        Returns:
            Workbook information dictionary
        """
        return {
            "sheet_names": workbook.sheetnames,
            "sheet_count": len(workbook.sheetnames),
            "active_sheet": workbook.active.title,
            "workbook": workbook
        }


# ============ 2. Table Region Recognition Tools ============

class NestedTableDetector(BaseTool):
    """Nested table detector"""
    
    def __init__(self):
        super().__init__(
            name="NestedTableDetector",
            description="Detect nested table structures in Excel",
            tool_type="detector"
        )
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> List[Dict[str, Any]]:
        """
        Detect nested tables
        
        Args:
            worksheet: Worksheet object
            
        Returns:
            List of nested tables
        """
        nested_tables = []
        # Simple implementation: detect multiple tables by empty rows/columns
        # More complex algorithms needed in actual applications
        
        return nested_tables


class RegionSegmenter(BaseTool):
    """Region segmenter (enhanced version: intelligent empty row handling)"""
    
    def __init__(self):
        super().__init__(
            name="RegionSegmenter",
            description="Segment worksheet into different table regions with intelligent empty row handling",
            tool_type="segmenter"
        )
        # Empty row processing parameters
        self.min_data_density = 0.3  # Minimum data density threshold
        self.max_empty_gap = 3       # Maximum empty row gap
        self.similarity_threshold = 0.7  # Header similarity threshold
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> List[Dict[str, Any]]:
        """
        Segment table regions with intelligent empty row handling
        
        Args:
            worksheet: Worksheet object
            
        Returns:
            List of regions, each containing boundary info and empty row processing results
        """
        # 1. First get original regions
        original_regions = self._get_original_regions(worksheet)
        
        # 2. Apply intelligent empty row processing to each region
        processed_regions = []
        for region in original_regions:
            processed_region = self._process_region_with_smart_empty_rows(worksheet, region)
            processed_regions.append(processed_region)
        
        return processed_regions
    
    def _get_original_regions(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> List[Dict[str, Any]]:
        """获取有效数据区域（智能检测）"""
        regions = []
        
        # 智能检测有效数据区域
        effective_region = self._detect_effective_data_region(worksheet)
        
        if effective_region:
            regions.append(effective_region)
        else:
            # 如果检测失败，使用默认区域
            min_row = worksheet.min_row
            max_row = worksheet.max_row
            min_col = worksheet.min_column
            max_col = worksheet.max_column
            
            region = {
                "min_row": min_row,
                "max_row": max_row,
                "min_col": min_col,
                "max_col": max_col,
                "area": (max_row - min_row + 1) * (max_col - min_col + 1)
            }
            regions.append(region)
        
        return regions
    
    def _detect_effective_data_region(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Any]:
        """检测有效数据区域，跳过大量空行"""
        # 1. 快速扫描找到数据边界
        data_rows = []
        data_cols = []
        
        # 限制扫描范围，避免处理超大文件
        max_scan_rows = min(worksheet.max_row, 5000)  # 最多扫描5000行
        max_scan_cols = min(worksheet.max_column, 50)  # 最多扫描50列
        
        # 扫描行
        for row in range(1, max_scan_rows + 1):
            has_data = False
            for col in range(1, max_scan_cols + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                    break
            if has_data:
                data_rows.append(row)
        
        # 扫描列
        for col in range(1, max_scan_cols + 1):
            has_data = False
            for row in range(1, max_scan_rows + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value is not None and str(cell_value).strip():
                    has_data = True
                    break
            if has_data:
                data_cols.append(col)
        
        if not data_rows or not data_cols:
            return None
        
        # 2. 计算有效区域
        min_row = min(data_rows)
        max_row = max(data_rows)
        min_col = min(data_cols)
        max_col = max(data_cols)
        
        # 3. 添加一些缓冲区域
        buffer_rows = 2
        buffer_cols = 1
        
        effective_region = {
            "min_row": max(1, min_row - buffer_rows),
            "max_row": min(worksheet.max_row, max_row + buffer_rows),
            "min_col": max(1, min_col - buffer_cols),
            "max_col": min(worksheet.max_column, max_col + buffer_cols),
            "area": (max_row - min_row + 1) * (max_col - min_col + 1),
            "data_density": self._calculate_region_density(worksheet, min_row, max_row, min_col, max_col),
            "original_size": worksheet.max_row * worksheet.max_column,
            "compression_ratio": (worksheet.max_row * worksheet.max_column) / ((max_row - min_row + 1) * (max_col - min_col + 1))
        }
        
        return effective_region
    
    def _calculate_region_density(self, worksheet, min_row, max_row, min_col, max_col):
        """计算区域数据密度"""
        total_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
        non_empty_cells = 0
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value is not None and str(cell_value).strip():
                    non_empty_cells += 1
        
        return non_empty_cells / total_cells if total_cells > 0 else 0.0
    
    def _process_region_with_smart_empty_rows(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                                            region: Dict[str, Any]) -> Dict[str, Any]:
        """对区域进行智能空行处理（简化版：直接返回有效区域）"""
        # 直接返回检测到的有效区域，不进行复杂的空行分析
        return {
            "original_region": region,
            "processed_regions": [region]  # 直接返回检测到的有效区域
        }
    
    def _analyze_empty_rows(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                           region: Dict[str, Any]) -> Dict[str, Any]:
        """分析空行类型"""
        empty_rows = self._find_empty_rows(worksheet, region)
        analysis = {
            "empty_rows": empty_rows,
            "nested_table_separators": [],
            "meaningless_empty_rows": [],
            "uncertain_rows": []
        }
        
        for row_idx in empty_rows:
            row_analysis = self._analyze_single_empty_row(worksheet, row_idx, region, empty_rows)
            
            if row_analysis["type"] == "nested_table_separator":
                analysis["nested_table_separators"].append(row_analysis)
            elif row_analysis["type"] == "meaningless_empty":
                analysis["meaningless_empty_rows"].append(row_analysis)
            else:
                analysis["uncertain_rows"].append(row_analysis)
        
        return analysis
    
    def _find_empty_rows(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                        region: Dict[str, Any]) -> List[int]:
        """查找空行"""
        empty_rows = []
        
        for row in range(region['min_row'], region['max_row'] + 1):
            if self._is_row_empty(worksheet, row, region):
                empty_rows.append(row)
        
        return empty_rows
    
    def _is_row_empty(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, row: int, 
                     region: Dict[str, Any]) -> bool:
        """判断行是否为空"""
        for col in range(region['min_col'], region['max_col'] + 1):
            cell_value = worksheet.cell(row, col).value
            if cell_value is not None and str(cell_value).strip():
                return False
        return True
    
    def _analyze_single_empty_row(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                                 row_idx: int, region: Dict[str, Any], 
                                 all_empty_rows: List[int]) -> Dict[str, Any]:
        """分析单个空行"""
        # 获取前后数据块信息
        prev_block = self._get_data_block_before(worksheet, row_idx, region)
        next_block = self._get_data_block_after(worksheet, row_idx, region)
        
        # 计算各种指标
        metrics = self._calculate_metrics(worksheet, row_idx, region, 
                                        prev_block, next_block, all_empty_rows)
        
        # 判断空行类型
        empty_row_type, confidence, reason = self._classify_empty_row(metrics)
        
        return {
            "row_index": row_idx,
            "type": empty_row_type,
            "confidence": confidence,
            "reason": reason,
            "metrics": metrics,
            "prev_block": prev_block,
            "next_block": next_block
        }
    
    def _get_data_block_before(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                              row_idx: int, region: Dict[str, Any]) -> Dict[str, Any]:
        """获取空行前的数据块信息"""
        block_info = {
            'start_row': None,
            'end_row': None,
            'data_density': 0.0,
            'has_headers': False,
            'header_pattern': None,
            'data_types': set()
        }
        
        # 向前查找数据块
        data_start = None
        for row in range(row_idx - 1, region['min_row'] - 1, -1):
            if not self._is_row_empty(worksheet, row, region):
                if data_start is None:
                    data_start = row
                block_info['end_row'] = row
            elif data_start is not None:
                break
        
        if data_start is not None:
            block_info['start_row'] = data_start
            block_info['data_density'] = self._calculate_data_density(
                worksheet, data_start, block_info['end_row'], region
            )
            block_info['has_headers'] = self._has_header_pattern(
                worksheet, data_start, block_info['end_row'], region
            )
            block_info['header_pattern'] = self._extract_header_pattern(
                worksheet, data_start, block_info['end_row'], region
            )
            block_info['data_types'] = self._extract_data_types(
                worksheet, data_start, block_info['end_row'], region
            )
        
        return block_info
    
    def _get_data_block_after(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                             row_idx: int, region: Dict[str, Any]) -> Dict[str, Any]:
        """获取空行后的数据块信息"""
        block_info = {
            'start_row': None,
            'end_row': None,
            'data_density': 0.0,
            'has_headers': False,
            'header_pattern': None,
            'data_types': set()
        }
        
        # 向后查找数据块
        data_start = None
        for row in range(row_idx + 1, region['max_row'] + 1):
            if not self._is_row_empty(worksheet, row, region):
                if data_start is None:
                    data_start = row
                block_info['end_row'] = row
            elif data_start is not None:
                break
        
        if data_start is not None:
            block_info['start_row'] = data_start
            block_info['data_density'] = self._calculate_data_density(
                worksheet, data_start, block_info['end_row'], region
            )
            block_info['has_headers'] = self._has_header_pattern(
                worksheet, data_start, block_info['end_row'], region
            )
            block_info['header_pattern'] = self._extract_header_pattern(
                worksheet, data_start, block_info['end_row'], region
            )
            block_info['data_types'] = self._extract_data_types(
                worksheet, data_start, block_info['end_row'], region
            )
        
        return block_info
    
    def _calculate_data_density(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                               start_row: int, end_row: int, 
                               region: Dict[str, Any]) -> float:
        """计算数据密度"""
        if start_row is None or end_row is None:
            return 0.0
        
        total_cells = (end_row - start_row + 1) * (region['max_col'] - region['min_col'] + 1)
        non_empty_cells = 0
        
        for row in range(start_row, end_row + 1):
            for col in range(region['min_col'], region['max_col'] + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value is not None and str(cell_value).strip():
                    non_empty_cells += 1
        
        return non_empty_cells / total_cells if total_cells > 0 else 0.0
    
    def _has_header_pattern(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                           start_row: int, end_row: int, 
                           region: Dict[str, Any]) -> bool:
        """判断是否有表头模式"""
        if start_row is None or end_row is None:
            return False
        
        # 检查前几行是否有表头特征
        header_rows = min(3, end_row - start_row + 1)
        for row in range(start_row, start_row + header_rows):
            if self._is_header_row(worksheet, row, region):
                return True
        return False
    
    def _is_header_row(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, row: int, 
                      region: Dict[str, Any]) -> bool:
        """判断是否为表头行"""
        # 检查是否包含文本且不包含数值
        has_text = False
        has_numeric = False
        
        for col in range(region['min_col'], region['max_col'] + 1):
            cell_value = worksheet.cell(row, col).value
            if cell_value is not None:
                if isinstance(cell_value, (int, float)):
                    has_numeric = True
                elif isinstance(cell_value, str) and cell_value.strip():
                    has_text = True
        
        return has_text and not has_numeric
    
    def _extract_header_pattern(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                               start_row: int, end_row: int, 
                               region: Dict[str, Any]) -> List[str]:
        """提取表头模式"""
        if start_row is None or end_row is None:
            return []
        
        header_pattern = []
        for col in range(region['min_col'], region['max_col'] + 1):
            cell_value = worksheet.cell(start_row, col).value
            header_pattern.append(str(cell_value) if cell_value else "")
        
        return header_pattern
    
    def _extract_data_types(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                           start_row: int, end_row: int, 
                           region: Dict[str, Any]) -> set:
        """提取数据类型"""
        if start_row is None or end_row is None:
            return set()
        
        data_types = set()
        for row in range(start_row, end_row + 1):
            for col in range(region['min_col'], region['max_col'] + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value is not None:
                    data_types.add(type(cell_value).__name__)
        
        return data_types
    
    def _calculate_metrics(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, row_idx: int,
                          region: Dict[str, Any], prev_block: Dict[str, Any],
                          next_block: Dict[str, Any], all_empty_rows: List[int]) -> Dict[str, Any]:
        """计算各种指标"""
        # 确保所有值都是数值类型
        prev_density = prev_block.get('data_density', 0.0)
        next_density = next_block.get('data_density', 0.0)
        
        # 如果返回的是字典，取默认值
        if isinstance(prev_density, dict):
            prev_density = 0.0
        if isinstance(next_density, dict):
            next_density = 0.0
            
        metrics = {
            'prev_block_density': float(prev_density),
            'next_block_density': float(next_density),
            'prev_has_headers': bool(prev_block.get('has_headers', False)),
            'next_has_headers': bool(next_block.get('has_headers', False)),
            'header_similarity': 0.0,
            'data_type_similarity': 0.0,
            'empty_gap_size': 0,
            'position_ratio': 0.0,
            'surrounding_empty_count': 0
        }
        
        # 计算表头相似度
        if prev_block.get('header_pattern') and next_block.get('header_pattern'):
            metrics['header_similarity'] = self._calculate_header_similarity(
                prev_block['header_pattern'], next_block['header_pattern']
            )
        
        # 计算数据类型相似度
        if prev_block.get('data_types') and next_block.get('data_types'):
            metrics['data_type_similarity'] = self._calculate_data_type_similarity(
                prev_block['data_types'], next_block['data_types']
            )
        
        # 计算空行间隔大小
        gap_size = self._calculate_empty_gap_size(row_idx, all_empty_rows)
        metrics['empty_gap_size'] = int(gap_size) if isinstance(gap_size, (int, float)) else 0
        
        # 计算位置比例
        try:
            position_ratio = (row_idx - region['min_row']) / (
                region['max_row'] - region['min_row']
            )
            metrics['position_ratio'] = float(position_ratio)
        except (ZeroDivisionError, TypeError):
            metrics['position_ratio'] = 0.0
        
        # 计算周围空行数量
        empty_count = self._count_surrounding_empty_rows(row_idx, all_empty_rows)
        metrics['surrounding_empty_count'] = int(empty_count) if isinstance(empty_count, (int, float)) else 0
        
        return metrics
    
    def _calculate_header_similarity(self, pattern1: List[str], pattern2: List[str]) -> float:
        """计算表头相似度"""
        if not pattern1 or not pattern2:
            return 0.0
        
        min_len = min(len(pattern1), len(pattern2))
        matches = 0
        
        for i in range(min_len):
            if pattern1[i] == pattern2[i]:
                matches += 1
        
        return matches / min_len
    
    def _calculate_data_type_similarity(self, types1: set, types2: set) -> float:
        """计算数据类型相似度"""
        if not types1 or not types2:
            return 0.0
        
        intersection = types1.intersection(types2)
        union = types1.union(types2)
        
        return len(intersection) / len(union) if union else 0.0
    
    def _calculate_empty_gap_size(self, row_idx: int, all_empty_rows: List[int]) -> int:
        """计算空行间隔大小"""
        if row_idx not in all_empty_rows:
            return 0
        
        # 计算连续空行的数量
        gap_size = 1
        current_idx = all_empty_rows.index(row_idx)
        
        # 向前计算
        for i in range(current_idx - 1, -1, -1):
            if all_empty_rows[i] == row_idx - gap_size:
                gap_size += 1
            else:
                break
        
        # 向后计算
        for i in range(current_idx + 1, len(all_empty_rows)):
            if all_empty_rows[i] == row_idx + gap_size:
                gap_size += 1
            else:
                break
        
        return gap_size
    
    def _count_surrounding_empty_rows(self, row_idx: int, all_empty_rows: List[int]) -> int:
        """计算周围空行数量"""
        if row_idx not in all_empty_rows:
            return 0
        
        current_idx = all_empty_rows.index(row_idx)
        count = 1
        
        # 向前计算
        for i in range(current_idx - 1, -1, -1):
            if all_empty_rows[i] == row_idx - count:
                count += 1
            else:
                break
        
        # 向后计算
        for i in range(current_idx + 1, len(all_empty_rows)):
            if all_empty_rows[i] == row_idx + count:
                count += 1
            else:
                break
        
        return count
    
    def _classify_empty_row(self, metrics: Dict[str, Any]) -> tuple:
        """分类空行类型"""
        # 嵌套表分隔的特征
        nested_table_score = 0.0
        nested_table_reasons = []
        
        # 1. 前后都有数据块
        if metrics['prev_block_density'] > 0 and metrics['next_block_density'] > 0:
            nested_table_score += 0.3
            nested_table_reasons.append("前后都有数据块")
        
        # 2. 前后都有表头
        if metrics['prev_has_headers'] and metrics['next_has_headers']:
            nested_table_score += 0.2
            nested_table_reasons.append("前后都有表头")
        
        # 3. 表头相似度低（可能是不同的表）
        if metrics['header_similarity'] < 0.5:
            nested_table_score += 0.2
            nested_table_reasons.append("表头相似度低")
        
        # 4. 数据类型相似度低
        if metrics['data_type_similarity'] < 0.5:
            nested_table_score += 0.1
            nested_table_reasons.append("数据类型相似度低")
        
        # 5. 空行间隔适中
        if 1 <= metrics['empty_gap_size'] <= 3:
            nested_table_score += 0.1
            nested_table_reasons.append("空行间隔适中")
        
        # 6. 位置不在边缘
        if 0.1 < metrics['position_ratio'] < 0.9:
            nested_table_score += 0.1
            nested_table_reasons.append("位置不在边缘")
        
        # 无意义空行的特征
        meaningless_score = 0.0
        meaningless_reasons = []
        
        # 1. 前后数据密度都高
        if metrics['prev_block_density'] > 0.7 and metrics['next_block_density'] > 0.7:
            meaningless_score += 0.3
            meaningless_reasons.append("前后数据密度都高")
        
        # 2. 表头相似度高
        if metrics['header_similarity'] > 0.8:
            meaningless_score += 0.2
            meaningless_reasons.append("表头相似度高")
        
        # 3. 数据类型相似度高
        if metrics['data_type_similarity'] > 0.8:
            meaningless_score += 0.2
            meaningless_reasons.append("数据类型相似度高")
        
        # 4. 空行间隔很小
        if metrics['empty_gap_size'] == 1:
            meaningless_score += 0.2
            meaningless_reasons.append("空行间隔很小")
        
        # 5. 周围空行很多
        if metrics['surrounding_empty_count'] > 3:
            meaningless_score += 0.1
            meaningless_reasons.append("周围空行很多")
        
        # 判断结果
        if nested_table_score > meaningless_score and nested_table_score > 0.5:
            return "nested_table_separator", nested_table_score, "; ".join(nested_table_reasons)
        elif meaningless_score > nested_table_score and meaningless_score > 0.5:
            return "meaningless_empty", meaningless_score, "; ".join(meaningless_reasons)
        else:
            return "uncertain", max(nested_table_score, meaningless_score), "特征不明显"
    
    def _split_regions_by_empty_rows(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                                    region: Dict[str, Any], 
                                    empty_row_analysis: Dict[str, Any]) -> List[Dict[str, Any]]:
        """根据空行分析结果分割区域"""
        sub_regions = []
        
        # 获取嵌套表分隔空行
        separator_rows = [analysis["row_index"] for analysis in empty_row_analysis["nested_table_separators"]]
        
        if not separator_rows:
            # 没有嵌套表分隔，返回原区域
            sub_regions.append(region)
        else:
            # 根据分隔空行分割区域
            separator_rows.sort()
            current_start = region['min_row']
            
            for separator_row in separator_rows:
                # 创建子区域（不包含分隔空行）
                if current_start < separator_row:
                    sub_region = {
                        "min_row": current_start,
                        "max_row": separator_row - 1,
                        "min_col": region['min_col'],
                        "max_col": region['max_col'],
                        "area": (separator_row - current_start) * (region['max_col'] - region['min_col'] + 1),
                        "type": "data_region"
                    }
                    sub_regions.append(sub_region)
                
                current_start = separator_row + 1
            
            # 添加最后一个区域
            if current_start <= region['max_row']:
                sub_region = {
                    "min_row": current_start,
                    "max_row": region['max_row'],
                    "min_col": region['min_col'],
                    "max_col": region['max_col'],
                    "area": (region['max_row'] - current_start + 1) * (region['max_col'] - region['min_col'] + 1),
                    "type": "data_region"
                }
                sub_regions.append(sub_region)
        
        return sub_regions
    
    def _split_regions_by_analysis(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                                  region: Dict[str, Any], 
                                  empty_row_analysis: Dict[str, Any],
                                  nested_patterns: Dict[str, Any]) -> List[Dict[str, Any]]:
        """基于空行分析和嵌套模式分析分割区域"""
        sub_regions = []
        
        # 1. 基于空行分割
        if empty_row_analysis["nested_table_separators"]:
            sub_regions.extend(self._split_by_empty_rows(worksheet, region, empty_row_analysis))
        
        # 2. 基于嵌套模式分割
        if any(patterns for patterns in nested_patterns.values()):
            pattern_regions = self._split_by_nested_patterns(worksheet, region, nested_patterns)
            sub_regions.extend(pattern_regions)
        
        # 3. 如果没有检测到分割模式，返回原区域
        if not sub_regions:
            sub_regions = [region]
        
        return sub_regions
    
    def _split_by_empty_rows(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                           region: Dict[str, Any], 
                           empty_row_analysis: Dict[str, Any]) -> List[Dict[str, Any]]:
        """基于空行分割区域"""
        sub_regions = []
        separators = empty_row_analysis["nested_table_separators"]
        
        if not separators:
            return [region]
        
        # 按分隔符分割区域
        start_row = region['min_row']
        end_row = region['max_row']
        
        current_start = start_row
        for separator_row in sorted(separators):
            if current_start < separator_row:
                sub_regions.append({
                    "min_row": current_start,
                    "max_row": separator_row - 1,
                    "min_col": region['min_col'],
                    "max_col": region['max_col'],
                    "split_type": "empty_row"
                })
            current_start = separator_row + 1
        
        # 添加最后一个区域
        if current_start <= end_row:
            sub_regions.append({
                "min_row": current_start,
                "max_row": end_row,
                "min_col": region['min_col'],
                "max_col": region['max_col'],
                "split_type": "empty_row"
            })
        
        return sub_regions
    
    def _split_by_nested_patterns(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                                region: Dict[str, Any], 
                                nested_patterns: Dict[str, Any]) -> List[Dict[str, Any]]:
        """基于嵌套模式分割区域"""
        sub_regions = []
        
        # 1. 基于重复模式分割
        repetition_patterns = nested_patterns.get("repetition_patterns", [])
        for pattern in repetition_patterns:
            if pattern["type"] == "row_repetition":
                # 按重复模式分割行
                pattern_length = pattern.get("pattern_length", pattern.get("period", 1))
                start_row = pattern["start_row"]
                end_row = pattern["end_row"]
                
                current_start = start_row
                while current_start <= end_row:
                    current_end = min(current_start + pattern_length - 1, end_row)
                    sub_regions.append({
                        "min_row": current_start,
                        "max_row": current_end,
                        "min_col": region['min_col'],
                        "max_col": region['max_col'],
                        "split_type": "repetition_pattern",
                        "pattern_info": pattern
                    })
                    current_start = current_end + 1
        
        # 2. 基于周期性模式分割
        periodic_patterns = [p for p in repetition_patterns if p["type"] == "periodic"]
        for pattern in periodic_patterns:
            period = pattern["period"]
            start_row = pattern["start_row"]
            end_row = pattern["end_row"]
            
            current_start = start_row
            while current_start <= end_row:
                current_end = min(current_start + period - 1, end_row)
                sub_regions.append({
                    "min_row": current_start,
                    "max_row": current_end,
                    "min_col": region['min_col'],
                    "max_col": region['max_col'],
                    "split_type": "periodic_pattern",
                    "pattern_info": pattern
                })
                current_start = current_end + 1
        
        # 3. 基于缩进模式分割
        indentation_patterns = nested_patterns.get("indentation_patterns", [])
        if indentation_patterns:
            # 按缩进级别分组
            indent_groups = self._group_by_indentation(indentation_patterns)
            for group in indent_groups:
                if len(group) > 1:  # 有多个缩进级别
                    sub_regions.append({
                        "min_row": min(p["row"] for p in group),
                        "max_row": max(p["row"] for p in group),
                        "min_col": region['min_col'],
                        "max_col": region['max_col'],
                        "split_type": "indentation_pattern",
                        "pattern_info": group
                    })
        
        # 4. 基于字体模式分割
        font_patterns = nested_patterns.get("font_patterns", [])
        if font_patterns:
            # 按字体变化分组
            font_groups = self._group_by_font_changes(font_patterns)
            for group in font_groups:
                if len(group) > 1:  # 有多个字体变化
                    sub_regions.append({
                        "min_row": min(p["row"] for p in group),
                        "max_row": max(p["row"] for p in group),
                        "min_col": region['min_col'],
                        "max_col": region['max_col'],
                        "split_type": "font_pattern",
                        "pattern_info": group
                    })
        
        return sub_regions
    
    def _group_by_indentation(self, indentation_patterns: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
        """按缩进级别分组"""
        groups = []
        current_group = []
        current_indent_level = None
        
        # 过滤出有 "row" 键的模式
        valid_patterns = [p for p in indentation_patterns if "row" in p]
        
        for pattern in sorted(valid_patterns, key=lambda x: x["row"]):
            analysis = pattern["analysis"]
            indent_levels = analysis.get("indent_levels", [])
            
            if indent_levels:
                avg_indent = sum(indent_levels) / len(indent_levels)
                
                if current_indent_level is None:
                    current_indent_level = avg_indent
                    current_group = [pattern]
                elif abs(avg_indent - current_indent_level) <= 1:  # 缩进级别相近
                    current_group.append(pattern)
                else:
                    if current_group:
                        groups.append(current_group)
                    current_group = [pattern]
                    current_indent_level = avg_indent
            else:
                if current_group:
                    groups.append(current_group)
                    current_group = []
                    current_indent_level = None
        
        if current_group:
            groups.append(current_group)
        
        return groups
    
    def _group_by_font_changes(self, font_patterns: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
        """按字体变化分组"""
        groups = []
        current_group = []
        current_font_style = None
        
        # 过滤出有 "row" 键的模式
        valid_patterns = [p for p in font_patterns if "row" in p]
        
        for pattern in sorted(valid_patterns, key=lambda x: x["row"]):
            analysis = pattern["analysis"]
            is_header = analysis.get("is_header", False)
            is_subheader = analysis.get("is_subheader", False)
            
            font_style = (is_header, is_subheader)
            
            if current_font_style is None:
                current_font_style = font_style
                current_group = [pattern]
            elif font_style == current_font_style:
                current_group.append(pattern)
            else:
                if current_group:
                    groups.append(current_group)
                current_group = [pattern]
                current_font_style = font_style
        
        if current_group:
            groups.append(current_group)
        
        return groups
    
    def _get_final_regions(self, sub_regions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """获取最终的区域列表"""
        # 过滤掉面积太小的区域
        final_regions = []
        for region in sub_regions:
            if region.get('area', 0) > 10:  # 至少10个单元格
                final_regions.append(region)
        
        return final_regions
    
    def _detect_nested_tables_by_pattern(self, worksheet, start_row, end_row, start_col, end_col):
        """检测无空行嵌套表（基于缩进、字体差异、重复模式）"""
        patterns = {
            "indentation_patterns": [],
            "font_patterns": [],
            "repetition_patterns": [],
            "separator_patterns": []
        }
        
        # 1. 检测缩进模式
        patterns["indentation_patterns"] = self._detect_indentation_patterns(worksheet, start_row, end_row, start_col, end_col)
        
        # 2. 检测字体差异模式
        patterns["font_patterns"] = self._detect_font_patterns(worksheet, start_row, end_row, start_col, end_col)
        
        # 3. 检测重复模式（如每N行一个子表）
        patterns["repetition_patterns"] = self._detect_repetition_patterns_enhanced(worksheet, start_row, end_row, start_col, end_col)
        
        # 4. 检测分隔符模式
        patterns["separator_patterns"] = self._detect_separator_patterns(worksheet, start_row, end_row, start_col, end_col)
        
        return patterns
    
    def _detect_indentation_patterns(self, worksheet, start_row, end_row, start_col, end_col):
        """检测缩进模式"""
        patterns = []
        
        # 分析每行的缩进情况
        for row_idx in range(start_row, end_row + 1):
            row_indentation = []
            
            for col_idx in range(start_col, end_col + 1):
                cell = worksheet.cell(row_idx, col_idx)
                cell_value = cell.value
                
                if cell_value and isinstance(cell_value, str):
                    # 检测缩进（空格、制表符等）
                    indent_level = self._calculate_indent_level(cell_value)
                    
                    # 检测对齐方式
                    alignment = self._get_cell_alignment(cell)
                    
                    row_indentation.append({
                        "col": col_idx,
                        "value": cell_value,
                        "indent_level": indent_level,
                        "alignment": alignment,
                        "has_indent": indent_level > 0
                    })
            
            # 分析整行的缩进模式
            if row_indentation:
                indent_analysis = self._analyze_row_indentation(row_indentation)
                if indent_analysis["has_pattern"]:
                    patterns.append({
                        "row": row_idx,
                        "type": "indentation",
                        "analysis": indent_analysis,
                        "details": row_indentation
                    })
        
        return patterns
    
    def _detect_font_patterns(self, worksheet, start_row, end_row, start_col, end_col):
        """检测字体差异模式"""
        patterns = []
        
        # 收集字体信息
        font_info = []
        for row_idx in range(start_row, end_row + 1):
            row_fonts = []
            
            for col_idx in range(start_col, end_col + 1):
                cell = worksheet.cell(row_idx, col_idx)
                cell_value = cell.value
                
                if cell_value:
                    font_style = self._get_cell_font_style(cell)
                    row_fonts.append({
                        "col": col_idx,
                        "value": cell_value,
                        "font_style": font_style
                    })
            
            if row_fonts:
                font_info.append({
                    "row": row_idx,
                    "fonts": row_fonts
                })
        
        # 分析字体模式
        for i, row_info in enumerate(font_info):
            row_idx = row_info["row"]
            fonts = row_info["fonts"]
            
            # 检测字体变化模式
            font_analysis = self._analyze_font_patterns(fonts, font_info, i)
            if font_analysis["has_pattern"]:
                patterns.append({
                    "row": row_idx,
                    "type": "font",
                    "analysis": font_analysis,
                    "details": fonts
                })
        
        return patterns
    
    def _detect_repetition_patterns_enhanced(self, worksheet, start_row, end_row, start_col, end_col):
        """检测增强的重复模式（如每N行一个子表）"""
        patterns = []
        
        # 检测行级别的重复模式
        row_patterns = self._detect_row_repetition_patterns(worksheet, start_row, end_row, start_col, end_col)
        patterns.extend(row_patterns)
        
        # 检测列级别的重复模式
        col_patterns = self._detect_col_repetition_patterns(worksheet, start_row, end_row, start_col, end_col)
        patterns.extend(col_patterns)
        
        # 检测周期性模式
        periodic_patterns = self._detect_periodic_patterns(worksheet, start_row, end_row, start_col, end_col)
        patterns.extend(periodic_patterns)
        
        return patterns
    
    def _detect_separator_patterns(self, worksheet, start_row, end_row, start_col, end_col):
        """检测分隔符模式"""
        patterns = []
        
        for row_idx in range(start_row, end_row + 1):
            row_values = []
            for col_idx in range(start_col, end_col + 1):
                cell_value = worksheet.cell(row_idx, col_idx).value
                row_values.append(cell_value)
            
            # 判断是否为分隔行
            if self._is_separator_row(row_values):
                separator_type = self._classify_separator_type(row_values)
                patterns.append({
                    "row": row_idx,
                    "type": "separator",
                    "separator_type": separator_type,
                    "values": row_values
                })
        
        return patterns
    
    def _calculate_indent_level(self, text):
        """计算文本的缩进级别"""
        if not text:
            return 0
        
        # 计算前导空格数
        leading_spaces = len(text) - len(text.lstrip())
        
        # 计算前导制表符数
        leading_tabs = 0
        for char in text:
            if char == '\t':
                leading_tabs += 1
            else:
                break
        
        # 综合计算缩进级别
        return leading_spaces + leading_tabs * 4  # 假设制表符等于4个空格
    
    def _get_cell_alignment(self, cell):
        """获取单元格对齐方式"""
        try:
            alignment = cell.alignment
            return {
                "horizontal": alignment.horizontal if alignment else None,
                "vertical": alignment.vertical if alignment else None,
                "indent": alignment.indent if alignment else 0
            }
        except:
            return {"horizontal": None, "vertical": None, "indent": 0}
    
    def _get_cell_font_style(self, cell):
        """获取单元格字体样式"""
        try:
            font = cell.font
            return {
                "bold": font.bold if font else False,
                "italic": font.italic if font else False,
                "size": font.size if font else None,
                "color": font.color.rgb if font and font.color else None,
                "name": font.name if font else None
            }
        except:
            return {"bold": False, "italic": False, "size": None, "color": None, "name": None}
    
    def _analyze_row_indentation(self, row_indentation):
        """分析行的缩进模式"""
        analysis = {
            "has_pattern": False,
            "indent_levels": [],
            "alignment_patterns": [],
            "is_nested": False
        }
        
        # 收集缩进级别
        indent_levels = [item["indent_level"] for item in row_indentation if item["has_indent"]]
        if indent_levels:
            analysis["indent_levels"] = indent_levels
            analysis["has_pattern"] = True
            
            # 判断是否为嵌套结构
            if len(set(indent_levels)) > 1:  # 有多个不同的缩进级别
                analysis["is_nested"] = True
        
        # 收集对齐模式
        alignments = [item["alignment"] for item in row_indentation]
        if alignments:
            analysis["alignment_patterns"] = alignments
        
        return analysis
    
    def _analyze_font_patterns(self, fonts, all_font_info, current_index):
        """分析字体模式"""
        analysis = {
            "has_pattern": False,
            "font_variations": [],
            "is_header": False,
            "is_subheader": False
        }
        
        # 分析当前行的字体特征
        current_fonts = [f["font_style"] for f in fonts]
        
        # 检测粗体模式（可能是表头）
        bold_count = sum(1 for f in current_fonts if f.get("bold", False))
        if bold_count > len(current_fonts) * 0.5:  # 超过50%是粗体
            analysis["is_header"] = True
            analysis["has_pattern"] = True
        
        # 检测字体大小变化
        font_sizes = [f.get("size") for f in current_fonts if f.get("size")]
        if font_sizes and len(set(font_sizes)) > 1:
            analysis["font_variations"] = font_sizes
            analysis["has_pattern"] = True
        
        # 与前后行比较
        if current_index > 0 and current_index < len(all_font_info) - 1:
            prev_fonts = [f["font_style"] for f in all_font_info[current_index - 1]["fonts"]]
            next_fonts = [f["font_style"] for f in all_font_info[current_index + 1]["fonts"]]
            
            # 检测子表头模式
            if (any(f.get("bold", False) for f in current_fonts) and
                not any(f.get("bold", False) for f in prev_fonts) and
                not any(f.get("bold", False) for f in next_fonts)):
                analysis["is_subheader"] = True
                analysis["has_pattern"] = True
        
        return analysis
    
    def _detect_row_repetition_patterns(self, worksheet, start_row, end_row, start_col, end_col):
        """检测行级别的重复模式"""
        patterns = []
        
        # 检测每N行重复的模式
        for period in range(2, min(10, (end_row - start_row + 1) // 2)):
            matches = 0
            for row_idx in range(start_row, end_row - period + 1):
                current_row = []
                next_row = []
                
                for col_idx in range(start_col, end_col + 1):
                    current_row.append(worksheet.cell(row_idx, col_idx).value)
                    next_row.append(worksheet.cell(row_idx + period, col_idx).value)
                
                if current_row == next_row:
                    matches += 1
            
            if matches > 0:
                patterns.append({
                    "type": "row_repetition",
                    "period": period,
                    "matches": matches,
                    "start_row": start_row,
                    "end_row": end_row
                })
        
        return patterns
    
    def _detect_col_repetition_patterns(self, worksheet, start_row, end_row, start_col, end_col):
        """检测列级别的重复模式"""
        patterns = []
        
        # 检测每N列重复的模式
        for period in range(2, min(10, (end_col - start_col + 1) // 2)):
            matches = 0
            for col_idx in range(start_col, end_col - period + 1):
                current_col = []
                next_col = []
                
                for row_idx in range(start_row, end_row + 1):
                    current_col.append(worksheet.cell(row_idx, col_idx).value)
                    next_col.append(worksheet.cell(row_idx, col_idx + period).value)
                
                if current_col == next_col:
                    matches += 1
            
            if matches > 0:
                patterns.append({
                    "type": "col_repetition",
                    "period": period,
                    "matches": matches,
                    "start_col": start_col,
                    "end_col": end_col
                })
        
        return patterns
    
    def _detect_periodic_patterns(self, worksheet, start_row, end_row, start_col, end_col):
        """检测周期性模式"""
        patterns = []
        
        # 检测周期性数据模式
        for row_idx in range(start_row, end_row - 2):
            row_values = []
            for col_idx in range(start_col, end_col + 1):
                row_values.append(worksheet.cell(row_idx, col_idx).value)
            
            # 检查后续行是否有相似模式
            for period in range(2, min(5, end_row - row_idx)):
                if row_idx + period <= end_row:
                    next_row_values = []
                    for col_idx in range(start_col, end_col + 1):
                        next_row_values.append(worksheet.cell(row_idx + period, col_idx).value)
                    
                    # 计算相似度
                    similarity = self._calculate_row_similarity(row_values, next_row_values)
                    if similarity > 0.7:  # 70%相似度阈值
                        patterns.append({
                            "type": "periodic",
                            "period": period,
                            "start_row": row_idx,
                            "similarity": similarity,
                            "end_row": end_row
                        })
        
        return patterns
    
    def _is_separator_row(self, row_values):
        """判断是否为分隔行"""
        # 检查是否大部分为空值
        empty_count = sum(1 for value in row_values if value is None or value == "")
        if empty_count > len(row_values) * 0.8:  # 80%以上为空
            return True
        
        # 检查是否为重复字符（如---、===等）
        non_empty_values = [str(value) for value in row_values if value is not None and value != ""]
        if non_empty_values:
            for value_str in non_empty_values:
                if len(value_str) > 1 and len(set(value_str)) == 1:
                    return True
        
        return False
    
    def _classify_separator_type(self, row_values):
        """分类分隔符类型"""
        non_empty_values = [str(value) for value in row_values if value is not None and value != ""]
        
        if not non_empty_values:
            return "empty"
        
        # 检查重复字符类型
        for value_str in non_empty_values:
            if len(value_str) > 1 and len(set(value_str)) == 1:
                char = value_str[0]
                if char == '-':
                    return "dash"
                elif char == '=':
                    return "equals"
                elif char == '_':
                    return "underscore"
                else:
                    return "repeated_char"
        
        return "mixed"
    
    def _calculate_row_similarity(self, row1, row2):
        """计算两行的相似度"""
        if len(row1) != len(row2):
            return 0.0
        
        matches = 0
        total = len(row1)
        
        for i in range(total):
            if row1[i] == row2[i]:
                matches += 1
        
        return matches / total if total > 0 else 0.0


class TableBoundaryFinder(BaseTool):
    """表格边界查找器"""
    
    def __init__(self):
        super().__init__(
            name="TableBoundaryFinder",
            description="查找表格的精确边界",
            tool_type="finder"
        )
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, region: Dict[str, Any]) -> Dict[str, Any]:
        """
        查找表格边界
        
        Args:
            worksheet: 工作表对象
            region: 初始区域
            
        Returns:
            精确的表格边界
        """
        # 基于工作表内容计算紧致边界（非空单元格的最小包围矩形）
        boundary = dict(region) if isinstance(region, dict) else {}
        max_row = worksheet.max_row or 0
        max_col = worksheet.max_column or 0
        # 初始化
        min_r = boundary.get("min_row") or max_row
        min_c = boundary.get("min_col") or max_col
        max_r = boundary.get("max_row") or 1
        max_c = boundary.get("max_col") or 1

        def _nonempty(val) -> bool:
            if val is None:
                return False
            if isinstance(val, str):
                return bool(val.strip())
            return True

        for r in range(1, max_row + 1):
            row_has_val = False
            for c in range(1, max_col + 1):
                v = worksheet.cell(r, c).value
                if _nonempty(v):
                    row_has_val = True
                    if c < min_c:
                        min_c = c
                    if c > max_c:
                        max_c = c
            if row_has_val:
                if r < min_r:
                    min_r = r
                if r > max_r:
                    max_r = r

        # 合并边界（若仍未更新则回退为全表）
        if min_r > max_r or min_c > max_c:
            min_r, min_c, max_r, max_c = 1, 1, max_row, max_col

        boundary.update({
            "min_row": int(min_r),
            "min_col": int(min_c),
            "max_row": int(max_r),
            "max_col": int(max_c)
        })
        return boundary


# ============ 3. 数据清洗工具 ============

class MergedCellSplitter(BaseTool):
    """合并单元格拆分器"""
    
    def __init__(self):
        super().__init__(
            name="MergedCellSplitter",
            description="拆分合并单元格并填充值",
            tool_type="cleaner"
        )
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Any]:
        """
        拆分合并单元格
        
        Args:
            worksheet: 工作表对象
            
        Returns:
            拆分后的单元格信息
        """
        merged_cells_info = []
        
        for merged_range in worksheet.merged_cells.ranges:
            # 获取合并单元格的值
            min_row = merged_range.min_row
            min_col = merged_range.min_col
            value = worksheet.cell(min_row, min_col).value
            
            merged_cells_info.append({
                "range": str(merged_range),
                "value": value,
                "min_row": min_row,
                "max_row": merged_range.max_row,
                "min_col": min_col,
                "max_col": merged_range.max_col
            })
        
        return {
            "merged_cells": merged_cells_info,
            "count": len(merged_cells_info)
        }


class NullValueFiller(BaseTool):
    """空值填充器"""
    
    def __init__(self):
        super().__init__(
            name="NullValueFiller",
            description="填充空值（前向填充、后向填充、默认值）",
            tool_type="cleaner"
        )
    
    def execute(self, df: pd.DataFrame, method: str = "ffill") -> pd.DataFrame:
        """
        填充空值
        
        Args:
            df: DataFrame对象
            method: 填充方法 (ffill/bfill/value)
            
        Returns:
            填充后的DataFrame
        """
        if method == "ffill":
            return df.fillna(method="ffill")
        elif method == "bfill":
            return df.fillna(method="bfill")
        else:
            return df.fillna("")


class DataTypeNormalizer(BaseTool):
    """数据类型规范化器"""
    
    def __init__(self):
        super().__init__(
            name="DataTypeNormalizer",
            description="规范化数据类型（数字、日期、文本）",
            tool_type="normalizer"
        )
    
    def execute(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        规范化数据类型
        
        Args:
            df: DataFrame对象
            
        Returns:
            规范化后的DataFrame
        """
        # 尝试转换数字类型
        for col in df.columns:
            try:
                df[col] = pd.to_numeric(df[col])
            except:
                pass
        
        return df


class WhitespaceRemover(BaseTool):
    """空白字符移除器"""
    
    def __init__(self):
        super().__init__(
            name="WhitespaceRemover",
            description="移除字符串前后的空白字符",
            tool_type="cleaner"
        )
    
    def execute(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        移除空白字符
        
        Args:
            df: DataFrame对象
            
        Returns:
            清洗后的DataFrame
        """
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].str.strip()
        
        return df


# ============ 4. 表头识别工具 ============

class ExcelToMarkdown(BaseTool):
    """将 Excel 工作表区域转换为 Markdown 表格文本（可截断/带合并单元格标注）"""
    
    def __init__(self):
        super().__init__(
            name="ExcelToMarkdown",
            description="Convert Excel sheet region to Markdown string",
            tool_type="converter"
        )
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                region: Optional[Dict[str, Any]] = None,
                max_rows: int = 999999,
                max_cols: int = 999999) -> Dict[str, Any]:
        # resolve region bounds
        if region:
            min_row = int(region.get("min_row", 1))
            max_row = int(region.get("max_row", worksheet.max_row))
            min_col = int(region.get("min_col", 1))
            max_col = int(region.get("max_col", worksheet.max_column))
        else:
            min_row = 1
            max_row = worksheet.max_row or 0
            min_col = 1
            max_col = worksheet.max_column or 0
        # load into matrix
        num_rows = max_row - min_row + 1
        num_cols = max_col - min_col + 1
        matrix = [[worksheet.cell(row=r, column=c).value if (1 <= r <= worksheet.max_row and 1 <= c <= worksheet.max_column) else ''
                   for c in range(min_col, max_col + 1)]
                  for r in range(min_row, max_row + 1)]
        # normalize None
        for r in range(len(matrix)):
            for c in range(len(matrix[r])):
                if matrix[r][c] is None:
                    matrix[r][c] = ''
        # annotate merged cells
        try:
            merged_ranges = list(worksheet.merged_cells.ranges)
        except Exception:
            merged_ranges = []
        for mrange in merged_ranges:
            try:
                base_val = worksheet.cell(row=mrange.min_row, column=mrange.min_col).value or ''
                coords = [(rr, cc) for rr in range(mrange.min_row, mrange.max_row + 1)
                          for cc in range(mrange.min_col, mrange.max_col + 1)]
                total = len(coords)
                for idx, (rr, cc) in enumerate(coords, start=1):
                    r0 = rr - min_row
                    c0 = cc - min_col
                    if 0 <= r0 < len(matrix) and 0 <= c0 < len(matrix[r0]):
                        suffix = f" [merge cell {idx}/{total}]"
                        display = f"{base_val}{suffix}" if base_val != '' else suffix
                        matrix[r0][c0] = display
            except Exception:
                continue
        # truncate if needed
        truncated_rows = False
        truncated_cols = False
        if len(matrix) > max_rows:
            matrix = matrix[:max_rows]
            truncated_rows = True
        if matrix and len(matrix[0]) > max_cols:
            matrix = [row[:max_cols] for row in matrix]
            truncated_cols = True
        # build markdown
        if not matrix or not any(any(str(cell) for cell in row) for row in matrix):
            return {"markdown": "*工作表为空*\n", "stats": {"rows": 0, "cols": 0}}
        header = [str(x) for x in matrix[0]]
        headers = "| " + " | ".join(h.replace('|', '\\|').replace('\n', ' ') for h in header) + " |\n"
        separator = "| " + " | ".join(["---"] * len(header)) + " |\n"
        body_rows = ""
        for row in matrix[1:]:
            formatted = [str(cell).replace('|', '\\|').replace('\n', ' ') for cell in row]
            body_rows += "| " + " | ".join(formatted) + " |\n"
        md = headers + separator + body_rows
        if truncated_rows or truncated_cols:
            note = []
            if truncated_rows:
                note.append(f"截断至前{max_rows}行")
            if truncated_cols:
                note.append(f"前{max_cols}列")
            md += "\n*注意: " + "，".join(note) + "*\n"
        return {"markdown": md, "stats": {"rows": num_rows, "cols": num_cols}}


class MarkdownNormalizer(BaseTool):
    """对 Markdown 文本进行标准化（空白折叠、统一百分号、去千分位等）"""
    def __init__(self):
        super().__init__(
            name="MarkdownNormalizer",
            description="Normalize markdown text for stable matching",
            tool_type="normalizer"
        )
    
    def execute(self, markdown_text: str) -> str:
        if not isinstance(markdown_text, str):
            return ""
        text = markdown_text.replace('\r\n', '\n')
        # collapse spaces
        text = re.sub(r"[\t ]+", " ", text)
        # unify percent spacing
        text = re.sub(r"\s*%", "%", text)
        # remove thousand separators inside back-to-back digits (basic)
        text = re.sub(r"(?<=\d),(?=\d)", "", text)
        # collapse multiple newlines
        text = re.sub(r"\n{3,}", "\n\n", text)
        return text

class ComprehensiveTableExtractor(BaseTool):
    """综合表格提取器 - 结合规则和DataFrame的混合方法"""
    
    def __init__(self):
        super().__init__(
            name="ComprehensiveTableExtractor",
            description="综合表格提取器，结合规则和DataFrame方法",
            tool_type="extractor"
        )
        self.rule_based_detector = RuleBasedHeaderDetector()
        self.df_based_detector = DataFrameBasedHeaderDetector()
        self.merged_cell_analyzer = MergedCellAnalyzer()
        self.nested_table_detector = EnhancedNestedTableDetector()
        self.validation_engine = HeaderValidationEngine()
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                region: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        综合提取表格信息
        
        Args:
            worksheet: 工作表对象
            region: 表格区域
            
        Returns:
            综合表格信息
        """
        # 1. 基础信息提取
        basic_info = self._extract_basic_info(worksheet, region)
        
        # 2. 合并单元格分析
        merged_cell_info = self.merged_cell_analyzer.execute(worksheet, region)
        
        # 3. 规则方法检测
        rule_based_result = self.rule_based_detector.execute(worksheet, region)
        
        # 4. DataFrame方法检测
        df_based_result = self.df_based_detector.execute(worksheet, region)
        
        # 5. 嵌套表检测
        nested_tables = self.nested_table_detector.execute(worksheet, region)
        
        # 6. 结果验证和融合
        validated_result = self.validation_engine.execute(
            rule_based_result, df_based_result, merged_cell_info, nested_tables
        )
        
        # 7. 综合信息构建
        comprehensive_info = self._build_comprehensive_info(
            basic_info, merged_cell_info, validated_result, nested_tables
        )
        
        return comprehensive_info
    
    def _extract_basic_info(self, worksheet, region):
        """提取基础信息"""
        if region:
            min_row = region.get("min_row", 1)
            max_row = region.get("max_row", worksheet.max_row)
            min_col = region.get("min_col", 1)
            max_col = region.get("max_col", worksheet.max_column)
        else:
            min_row = 1
            max_row = worksheet.max_row
            min_col = 1
            max_col = worksheet.max_column
        
        return {
            "dimensions": {
                "rows": max_row - min_row + 1,
                "cols": max_col - min_col + 1,
                "min_row": min_row,
                "max_row": max_row,
                "min_col": min_col,
                "max_col": max_col
            },
            "data_density": self._calculate_data_density(worksheet, min_row, max_row, min_col, max_col),
            "cell_types": self._analyze_cell_types(worksheet, min_row, max_row, min_col, max_col)
        }
    
    def _calculate_data_density(self, worksheet, min_row, max_row, min_col, max_col):
        """计算数据密度"""
        total_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
        non_empty_cells = 0
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value is not None and str(cell_value).strip():
                    non_empty_cells += 1
        
        return non_empty_cells / total_cells if total_cells > 0 else 0.0
    
    def _analyze_cell_types(self, worksheet, min_row, max_row, min_col, max_col):
        """分析单元格类型分布"""
        cell_types = {"text": 0, "numeric": 0, "date": 0, "empty": 0}
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value is None or str(cell_value).strip() == "":
                    cell_types["empty"] += 1
                elif isinstance(cell_value, (int, float)):
                    cell_types["numeric"] += 1
                elif isinstance(cell_value, str):
                    cell_types["text"] += 1
        
        return cell_types
    
    def _build_comprehensive_info(self, basic_info, merged_cell_info, validated_result, nested_tables):
        """构建综合信息"""
        return {
            "basic_info": basic_info,
            "merged_cells": merged_cell_info,
            "headers": validated_result,
            "nested_tables": nested_tables,
            "extraction_confidence": self._calculate_extraction_confidence(validated_result),
            "complexity_score": self._calculate_complexity_score(basic_info, merged_cell_info, nested_tables)
        }
    
    def _calculate_extraction_confidence(self, validated_result):
        """计算提取置信度"""
        confidence_factors = []
        
        # 表头检测置信度
        if validated_result.get("column_headers", {}).get("confidence", 0) > 0.7:
            confidence_factors.append(0.3)
        if validated_result.get("row_headers", {}).get("confidence", 0) > 0.7:
            confidence_factors.append(0.3)
        
        # 层级结构置信度
        if validated_result.get("hierarchy", {}).get("levels", 0) > 0:
            confidence_factors.append(0.2)
        
        # 合并单元格处理置信度
        if validated_result.get("merged_cell_handling", {}).get("success", False):
            confidence_factors.append(0.2)
        
        return sum(confidence_factors) if confidence_factors else 0.0
    
    def _calculate_complexity_score(self, basic_info, merged_cell_info, nested_tables):
        """计算表格复杂度分数"""
        score = 0.0
        
        # 数据密度影响
        density = basic_info.get("data_density", 0)
        score += density * 0.2
        
        # 合并单元格复杂度
        merged_count = len(merged_cell_info.get("merged_cells", []))
        if merged_count > 0:
            score += min(merged_count / 10, 0.3)
        
        # 嵌套表复杂度
        nested_count = len(nested_tables.get("tables", []))
        if nested_count > 1:
            score += min(nested_count / 5, 0.3)
        
        # 单元格类型多样性
        cell_types = basic_info.get("cell_types", {})
        type_diversity = len([t for t in cell_types.values() if t > 0])
        score += (type_diversity / 4) * 0.2
        
        return min(score, 1.0)


class RuleBasedHeaderDetector(BaseTool):
    """基于规则的表头检测器"""
    
    def __init__(self):
        super().__init__(
            name="RuleBasedHeaderDetector",
            description="基于规则的表头检测器",
            tool_type="detector"
        )
        self.header_keywords = [
            "总计", "合计", "小计", "分类", "类别", "类型", "名称", "项目",
            "数量", "金额", "比例", "百分比", "时间", "日期", "年份", "月份",
            "地区", "区域", "省份", "城市", "国家", "单位", "部门", "机构"
        ]
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                region: Dict[str, Any] = None) -> Dict[str, Any]:
        """基于规则的表头检测"""
        if region:
            min_row = region.get("min_row", 1)
            max_row = region.get("max_row", worksheet.max_row)
            min_col = region.get("min_col", 1)
            max_col = region.get("max_col", worksheet.max_column)
        else:
            min_row = 1
            max_row = worksheet.max_row
            min_col = 1
            max_col = worksheet.max_column
        
        # 检测列表头
        col_headers = self._detect_column_headers_rule_based(worksheet, min_row, max_row, min_col, max_col)
        
        # 检测行表头
        row_headers = self._detect_row_headers_rule_based(worksheet, min_row, max_row, min_col, max_col)
        
        return {
            "column_headers": col_headers,
            "row_headers": row_headers,
            "method": "rule_based"
        }
    
    def _detect_column_headers_rule_based(self, worksheet, min_row, max_row, min_col, max_col):
        """基于规则检测列表头"""
        best_row = min_row
        best_score = 0.0
        
        # 检查前几行
        for row in range(min_row, min(max_row, min_row + 5)):
            score = self._calculate_header_score(worksheet, row, min_col, max_col, "row")
            if score > best_score:
                best_score = score
                best_row = row
        
        # 提取表头内容
        headers = []
        for col in range(min_col, max_col + 1):
            cell_value = worksheet.cell(best_row, col).value
            headers.append(cell_value)
        
        return {
            "header_row": best_row,
            "headers": headers,
            "score": best_score,
            "confidence": best_score
        }
    
    def _detect_row_headers_rule_based(self, worksheet, min_row, max_row, min_col, max_col):
        """基于规则检测行表头"""
        best_col = min_col
        best_score = 0.0
        
        # 检查前几列
        for col in range(min_col, min(max_col, min_col + 5)):
            score = self._calculate_header_score(worksheet, col, min_row, max_row, "col")
            if score > best_score:
                best_score = score
                best_col = col
        
        # 提取表头内容
        headers = []
        for row in range(min_row, max_row + 1):
            cell_value = worksheet.cell(row, best_col).value
            headers.append(cell_value)
        
        return {
            "header_col": best_col,
            "headers": headers,
            "score": best_score,
            "confidence": best_score
        }
    
    def _calculate_header_score(self, worksheet, index, start, end, direction):
        """计算表头得分"""
        score = 0.0
        
        if direction == "row":
            # 检查行
            for col in range(start, end + 1):
                cell_value = worksheet.cell(index, col).value
                if cell_value and isinstance(cell_value, str):
                    # 关键词匹配
                    for keyword in self.header_keywords:
                        if keyword in str(cell_value):
                            score += 0.1
                    # 文本特征
                    if len(str(cell_value)) > 2:
                        score += 0.05
        else:
            # 检查列
            for row in range(start, end + 1):
                cell_value = worksheet.cell(row, index).value
                if cell_value and isinstance(cell_value, str):
                    # 关键词匹配
                    for keyword in self.header_keywords:
                        if keyword in str(cell_value):
                            score += 0.1
                    # 文本特征
                    if len(str(cell_value)) > 2:
                        score += 0.05
        
        return min(score, 1.0)


class DataFrameBasedHeaderDetector(BaseTool):
    """基于DataFrame的表头检测器（改进版）"""
    
    def __init__(self):
        super().__init__(
            name="DataFrameBasedHeaderDetector",
            description="基于DataFrame的表头检测器",
            tool_type="detector"
        )
        self.min_header_density = 0.3
        self.max_search_rows = 15
        self.max_search_cols = 10
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                region: Dict[str, Any] = None) -> Dict[str, Any]:
        """基于DataFrame的表头检测"""
        return self.detect_headers(worksheet, region)
    
    def detect_headers(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                      region: Dict[str, Any] = None) -> Dict[str, Any]:
        """检测表头信息（兼容性方法）"""
        # 转换为DataFrame
        df = self._worksheet_to_dataframe(worksheet, region)
        
        # 检测列表头
        col_headers = self._detect_column_headers_df(df)
        
        # 检测行表头
        row_headers = self._detect_row_headers_df(df)
        
        # 检测嵌套表
        nested_tables = self._detect_nested_tables_df(df)
        
        return {
            "column_headers": col_headers,
            "row_headers": row_headers,
            "nested_tables": nested_tables,
            "method": "dataframe_based"
        }
    
    def _worksheet_to_dataframe(self, worksheet, region):
        """将工作表转换为DataFrame"""
        if region:
            min_row = region.get("min_row", 1)
            max_row = region.get("max_row", worksheet.max_row)
            min_col = region.get("min_col", 1)
            max_col = region.get("max_col", worksheet.max_column)
        else:
            min_row = 1
            max_row = worksheet.max_row
            min_col = 1
            max_col = worksheet.max_column
        
        data = []
        for row in range(min_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell_value = worksheet.cell(row, col).value
                row_data.append(cell_value)
            data.append(row_data)
        
        return pd.DataFrame(data)
    
    def _detect_column_headers_df(self, df):
        """基于DataFrame检测列表头"""
        header_scores = []
        
        for row_idx in range(min(self.max_search_rows, len(df))):
            row = df.iloc[row_idx]
            
            # 计算非空值密度
            non_null_count = row.notna().sum()
            total_count = len(row)
            density = non_null_count / total_count if total_count > 0 else 0
            
            # 计算文本比例
            text_count = 0
            for val in row:
                if pd.notna(val) and isinstance(val, str) and val.strip():
                    text_count += 1
            text_ratio = text_count / non_null_count if non_null_count > 0 else 0
            
            # 计算综合得分
            score = density * 0.6 + text_ratio * 0.4
            header_scores.append((row_idx, score, density, text_ratio))
        
        # 选择得分最高的行
        best_row_idx = max(header_scores, key=lambda x: x[1])[0]
        best_score = max(header_scores, key=lambda x: x[1])[1]
        
        # 提取表头内容
        header_row = df.iloc[best_row_idx]
        headers = []
        for col_idx, val in enumerate(header_row):
            if pd.notna(val):
                headers.append(str(val).strip())
            else:
                headers.append(None)
        
        return {
            "header_row": best_row_idx,
            "headers": headers,
            "score": best_score,
            "density": header_scores[best_row_idx][2],
            "text_ratio": header_scores[best_row_idx][3],
            "confidence": best_score
        }
    
    def _detect_row_headers_df(self, df):
        """基于DataFrame检测行表头"""
        header_scores = []
        
        for col_idx in range(min(self.max_search_cols, len(df.columns))):
            col = df.iloc[:, col_idx]
            
            # 计算非空值密度
            non_null_count = col.notna().sum()
            total_count = len(col)
            density = non_null_count / total_count if total_count > 0 else 0
            
            # 计算文本比例
            text_count = 0
            for val in col:
                if pd.notna(val) and isinstance(val, str) and val.strip():
                    text_count += 1
            text_ratio = text_count / non_null_count if non_null_count > 0 else 0
            
            # 计算综合得分
            score = density * 0.6 + text_ratio * 0.4
            header_scores.append((col_idx, score, density, text_ratio))
        
        # 选择得分最高的列
        best_col_idx = max(header_scores, key=lambda x: x[1])[0]
        best_score = max(header_scores, key=lambda x: x[1])[1]
        
        # 提取表头内容
        header_col = df.iloc[:, best_col_idx]
        headers = []
        for row_idx, val in enumerate(header_col):
            if pd.notna(val):
                headers.append(str(val).strip())
            else:
                headers.append(None)
        
        return {
            "header_col": best_col_idx,
            "headers": headers,
            "score": best_score,
            "density": header_scores[best_col_idx][2],
            "text_ratio": header_scores[best_col_idx][3],
            "confidence": best_score
        }
    
    def _detect_nested_tables_df(self, df):
        """基于DataFrame检测嵌套表"""
        nested_tables = []
        
        # 检测空行分割的表格
        empty_rows = []
        for row_idx in range(len(df)):
            if df.iloc[row_idx].isna().all():
                empty_rows.append(row_idx)
        
        # 根据空行分割表格
        if empty_rows:
            start_row = 0
            for empty_row in empty_rows:
                if empty_row > start_row:
                    table_region = {
                        "start_row": start_row,
                        "end_row": empty_row - 1,
                        "start_col": 0,
                        "end_col": len(df.columns) - 1,
                        "type": "row_separated"
                    }
                    nested_tables.append(table_region)
                start_row = empty_row + 1
            
            # 添加最后一个表格
            if start_row < len(df):
                table_region = {
                    "start_row": start_row,
                    "end_row": len(df) - 1,
                    "start_col": 0,
                    "end_col": len(df.columns) - 1,
                    "type": "row_separated"
                }
                nested_tables.append(table_region)
        else:
            # 没有空行，整个表格作为一个整体
            nested_tables.append({
                "start_row": 0,
                "end_row": len(df) - 1,
                "start_col": 0,
                "end_col": len(df.columns) - 1,
                "type": "single_table"
            })
        
        return {"tables": nested_tables}


class MergedCellAnalyzer(BaseTool):
    """合并单元格分析器"""
    
    def __init__(self):
        super().__init__(
            name="MergedCellAnalyzer",
            description="分析合并单元格",
            tool_type="analyzer"
        )
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                region: Dict[str, Any] = None) -> Dict[str, Any]:
        """分析合并单元格"""
        merged_cells_info = []
        
        for merged_range in worksheet.merged_cells.ranges:
            # 获取合并单元格的值
            min_row = merged_range.min_row
            min_col = merged_range.min_col
            value = worksheet.cell(min_row, min_col).value
            
            merged_cells_info.append({
                "range": str(merged_range),
                "value": value,
                "min_row": min_row,
                "max_row": merged_range.max_row,
                "min_col": min_col,
                "max_col": merged_range.max_col,
                "span_rows": merged_range.max_row - min_row + 1,
                "span_cols": merged_range.max_col - min_col + 1,
                "area": (merged_range.max_row - min_row + 1) * (merged_range.max_col - min_col + 1)
            })
        
        return {
            "merged_cells": merged_cells_info,
            "count": len(merged_cells_info),
            "total_area": sum(mc["area"] for mc in merged_cells_info),
            "complexity": self._calculate_merged_cell_complexity(merged_cells_info)
        }
    
    def _calculate_merged_cell_complexity(self, merged_cells):
        """计算合并单元格复杂度"""
        if not merged_cells:
            return 0.0
        
        # 基于合并单元格数量和跨度计算复杂度
        total_span = sum(mc["span_rows"] * mc["span_cols"] for mc in merged_cells)
        max_span = max(mc["span_rows"] * mc["span_cols"] for mc in merged_cells)
        
        complexity = (len(merged_cells) * 0.1) + (total_span * 0.01) + (max_span * 0.05)
        return min(complexity, 1.0)


class EnhancedNestedTableDetector(BaseTool):
    """增强的嵌套表检测器"""
    
    def __init__(self):
        super().__init__(
            name="EnhancedNestedTableDetector",
            description="检测嵌套表结构",
            tool_type="detector"
        )
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                region: Dict[str, Any] = None) -> Dict[str, Any]:
        """检测嵌套表"""
        if region:
            min_row = region.get("min_row", 1)
            max_row = region.get("max_row", worksheet.max_row)
            min_col = region.get("min_col", 1)
            max_col = region.get("max_col", worksheet.max_column)
        else:
            min_row = 1
            max_row = worksheet.max_row
            min_col = 1
            max_col = worksheet.max_column
        
        # 检测空行分割的嵌套表
        empty_row_tables = self._detect_empty_row_separated_tables(worksheet, min_row, max_row, min_col, max_col)
        
        # 检测模式重复的嵌套表
        pattern_tables = self._detect_pattern_based_tables(worksheet, min_row, max_row, min_col, max_col)
        
        # 检测合并单元格分割的嵌套表
        merged_cell_tables = self._detect_merged_cell_separated_tables(worksheet, min_row, max_row, min_col, max_col)
        
        return {
            "empty_row_tables": empty_row_tables,
            "pattern_tables": pattern_tables,
            "merged_cell_tables": merged_cell_tables,
            "total_tables": len(empty_row_tables) + len(pattern_tables) + len(merged_cell_tables)
        }
    
    def _detect_empty_row_separated_tables(self, worksheet, min_row, max_row, min_col, max_col):
        """检测空行分割的嵌套表"""
        tables = []
        empty_rows = []
        
        # 查找空行
        for row in range(min_row, max_row + 1):
            is_empty = True
            for col in range(min_col, max_col + 1):
                cell_value = worksheet.cell(row, col).value
                if cell_value is not None and str(cell_value).strip():
                    is_empty = False
                    break
            if is_empty:
                empty_rows.append(row)
        
        # 根据空行分割表格
        if empty_rows:
            start_row = min_row
            for empty_row in empty_rows:
                if empty_row > start_row:
                    tables.append({
                        "start_row": start_row,
                        "end_row": empty_row - 1,
                        "start_col": min_col,
                        "end_col": max_col,
                        "type": "empty_row_separated"
                    })
                start_row = empty_row + 1
            
            # 添加最后一个表格
            if start_row <= max_row:
                tables.append({
                    "start_row": start_row,
                    "end_row": max_row,
                    "start_col": min_col,
                    "end_col": max_col,
                    "type": "empty_row_separated"
                })
        else:
            # 没有空行，整个区域作为一个表格
            tables.append({
                "start_row": min_row,
                "end_row": max_row,
                "start_col": min_col,
                "end_col": max_col,
                "type": "single_table"
            })
        
        return tables
    
    def _detect_pattern_based_tables(self, worksheet, min_row, max_row, min_col, max_col):
        """检测基于模式的嵌套表"""
        tables = []
        
        # 检测重复模式
        for period in range(2, min(10, (max_row - min_row + 1) // 2)):
            if self._is_periodic_pattern(worksheet, min_row, max_row, min_col, max_col, period):
                # 按周期分割表格
                current_start = min_row
                while current_start <= max_row:
                    current_end = min(current_start + period - 1, max_row)
                    tables.append({
                        "start_row": current_start,
                        "end_row": current_end,
                        "start_col": min_col,
                        "end_col": max_col,
                        "type": "pattern_based",
                        "period": period
                    })
                    current_start = current_end + 1
        
        return tables
    
    def _detect_merged_cell_separated_tables(self, worksheet, min_row, max_row, min_col, max_col):
        """检测合并单元格分割的嵌套表"""
        tables = []
        
        # 分析合并单元格的分布
        merged_ranges = list(worksheet.merged_cells.ranges)
        if merged_ranges:
            # 按合并单元格分割表格
            separator_rows = []
            for merged_range in merged_ranges:
                if merged_range.min_row > min_row and merged_range.min_row < max_row:
                    separator_rows.append(merged_range.min_row)
            
            if separator_rows:
                separator_rows.sort()
                start_row = min_row
                for sep_row in separator_rows:
                    if sep_row > start_row:
                        tables.append({
                            "start_row": start_row,
                            "end_row": sep_row - 1,
                            "start_col": min_col,
                            "end_col": max_col,
                            "type": "merged_cell_separated"
                        })
                    start_row = sep_row
                
                # 添加最后一个表格
                if start_row <= max_row:
                    tables.append({
                        "start_row": start_row,
                        "end_row": max_row,
                        "start_col": min_col,
                        "end_col": max_col,
                        "type": "merged_cell_separated"
                    })
        
        return tables
    
    def _is_periodic_pattern(self, worksheet, min_row, max_row, min_col, max_col, period):
        """判断是否为周期性模式"""
        # 检查每period行是否相似
        for base_row in range(min_row, min(max_row - period + 1, min_row + 10)):
            base_data = []
            for col in range(min_col, max_col + 1):
                cell_value = worksheet.cell(base_row, col).value
                base_data.append(str(cell_value) if cell_value else "")
            
            # 检查后续的period行是否与base_row相似
            is_periodic = True
            for offset in range(period, min(max_row - base_row + 1, 20), period):
                if base_row + offset > max_row:
                    break
                
                compare_data = []
                for col in range(min_col, max_col + 1):
                    cell_value = worksheet.cell(base_row + offset, col).value
                    compare_data.append(str(cell_value) if cell_value else "")
                
                if not self._is_data_similar(base_data, compare_data):
                    is_periodic = False
                    break
            
            if is_periodic:
                return True
        
        return False
    
    def _is_data_similar(self, data1, data2):
        """判断两组数据是否相似"""
        if len(data1) != len(data2):
            return False
        
        matches = sum(1 for d1, d2 in zip(data1, data2) if d1 == d2)
        similarity = matches / len(data1)
        return similarity > 0.7


class HeaderValidationEngine(BaseTool):
    """表头验证引擎"""
    
    def __init__(self):
        super().__init__(
            name="HeaderValidationEngine",
            description="验证和融合表头检测结果",
            tool_type="validator"
        )
    
    def execute(self, rule_result: Dict[str, Any], df_result: Dict[str, Any], 
                merged_cell_info: Dict[str, Any], nested_tables: Dict[str, Any]) -> Dict[str, Any]:
        """验证和融合结果"""
        # 1. 表头验证
        validated_headers = self._validate_headers(rule_result, df_result)
        
        # 2. 合并单元格处理
        merged_cell_handling = self._handle_merged_cells(merged_cell_info, validated_headers)
        
        # 3. 嵌套表处理
        nested_table_handling = self._handle_nested_tables(nested_tables, validated_headers)
        
        # 4. 综合结果
        return {
            "column_headers": validated_headers["column_headers"],
            "row_headers": validated_headers["row_headers"],
            "hierarchy": self._build_hierarchy(validated_headers, merged_cell_handling),
            "merged_cell_handling": merged_cell_handling,
            "nested_table_handling": nested_table_handling,
            "validation_confidence": self._calculate_validation_confidence(validated_headers)
        }
    
    def _validate_headers(self, rule_result, df_result):
        """验证表头检测结果"""
        # 比较两种方法的结果
        rule_col = rule_result.get("column_headers", {})
        df_col = df_result.get("column_headers", {})
        
        rule_row = rule_result.get("row_headers", {})
        df_row = df_result.get("row_headers", {})
        
        # 选择置信度更高的结果
        if rule_col.get("confidence", 0) > df_col.get("confidence", 0):
            validated_col = rule_col
            validated_col["method"] = "rule_based"
        else:
            validated_col = df_col
            validated_col["method"] = "dataframe_based"
        
        if rule_row.get("confidence", 0) > df_row.get("confidence", 0):
            validated_row = rule_row
            validated_row["method"] = "rule_based"
        else:
            validated_row = df_row
            validated_row["method"] = "dataframe_based"
        
        return {
            "column_headers": validated_col,
            "row_headers": validated_row
        }
    
    def _handle_merged_cells(self, merged_cell_info, headers):
        """处理合并单元格"""
        if not merged_cell_info.get("merged_cells"):
            return {"success": True, "message": "No merged cells found"}
        
        # 分析合并单元格与表头的关系
        col_header_row = headers["column_headers"].get("header_row", 0)
        row_header_col = headers["row_headers"].get("header_col", 0)
        
        merged_cell_analysis = {
            "header_merged_cells": [],
            "data_merged_cells": [],
            "complexity": merged_cell_info.get("complexity", 0)
        }
        
        for mc in merged_cell_info["merged_cells"]:
            if mc["min_row"] <= col_header_row <= mc["max_row"]:
                merged_cell_analysis["header_merged_cells"].append(mc)
            else:
                merged_cell_analysis["data_merged_cells"].append(mc)
        
        return {
            "success": True,
            "analysis": merged_cell_analysis,
            "recommendations": self._generate_merged_cell_recommendations(merged_cell_analysis)
        }
    
    def _handle_nested_tables(self, nested_tables, headers):
        """处理嵌套表"""
        if not nested_tables.get("total_tables", 0) > 1:
            return {"success": True, "message": "No nested tables detected"}
        
        # 分析嵌套表结构
        nested_analysis = {
            "table_count": nested_tables.get("total_tables", 0),
            "table_types": {},
            "complexity": 0
        }
        
        for table_type in ["empty_row_tables", "pattern_tables", "merged_cell_tables"]:
            tables = nested_tables.get(table_type, [])
            if tables:
                nested_analysis["table_types"][table_type] = len(tables)
                nested_analysis["complexity"] += len(tables) * 0.1
        
        return {
            "success": True,
            "analysis": nested_analysis,
            "recommendations": self._generate_nested_table_recommendations(nested_analysis)
        }
    
    def _build_hierarchy(self, headers, merged_cell_handling):
        """构建层级结构"""
        hierarchy = {
            "levels": 1,
            "column_hierarchy": [],
            "row_hierarchy": []
        }
        
        # 分析列表头层级
        col_headers = headers["column_headers"].get("headers", [])
        if col_headers:
            hierarchy["column_hierarchy"] = self._analyze_header_hierarchy(col_headers)
        
        # 分析行表头层级
        row_headers = headers["row_headers"].get("headers", [])
        if row_headers:
            hierarchy["row_hierarchy"] = self._analyze_header_hierarchy(row_headers)
        
        # 确定最大层级
        max_col_level = max(hierarchy["column_hierarchy"], default=0)
        max_row_level = max(hierarchy["row_hierarchy"], default=0)
        hierarchy["levels"] = max(max_col_level, max_row_level, 1)
        
        return hierarchy
    
    def _analyze_header_hierarchy(self, headers):
        """分析表头层级"""
        hierarchy_levels = []
        
        for i, header in enumerate(headers):
            if header and isinstance(header, str):
                # 基于缩进判断层级
                indent_level = len(header) - len(header.lstrip())
                hierarchy_levels.append(max(indent_level // 2, 0))
            else:
                hierarchy_levels.append(0)
        
        return hierarchy_levels
    
    def _calculate_validation_confidence(self, headers):
        """计算验证置信度"""
        col_confidence = headers["column_headers"].get("confidence", 0)
        row_confidence = headers["row_headers"].get("confidence", 0)
        
        # 综合置信度
        return (col_confidence + row_confidence) / 2
    
    def _generate_merged_cell_recommendations(self, analysis):
        """生成合并单元格处理建议"""
        recommendations = []
        
        if analysis["complexity"] > 0.5:
            recommendations.append("High merged cell complexity detected. Consider splitting merged cells.")
        
        if len(analysis["header_merged_cells"]) > 0:
            recommendations.append("Merged cells found in header area. May affect header detection.")
        
        if len(analysis["data_merged_cells"]) > 0:
            recommendations.append("Merged cells found in data area. May affect data extraction.")
        
        return recommendations
    
    def _generate_nested_table_recommendations(self, analysis):
        """生成嵌套表处理建议"""
        recommendations = []
        
        if analysis["table_count"] > 1:
            recommendations.append(f"Multiple tables detected ({analysis['table_count']}). Consider processing separately.")
        
        if analysis["complexity"] > 0.3:
            recommendations.append("High table complexity detected. May require specialized processing.")
        
        return recommendations


# ============ 5. 增强的表头检测工具 ============

class RowHeaderDetector(BaseTool):
    """行表头检测器（基于DataFrame改进版）"""
    
    def __init__(self):
        super().__init__(
            name="RowHeaderDetector",
            description="检测和识别行表头（基于DataFrame智能检测）",
            tool_type="detector"
        )
        # 使用内置的检测器
        self.df_detector = DataFrameBasedHeaderDetector()
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, region: Dict[str, Any]) -> Dict[str, Any]:
        """
        检测行表头（基于DataFrame智能识别）
        
        Args:
            worksheet: 工作表对象
            region: 表格区域
            
        Returns:
            行表头信息
        """
        # 使用改进的DataFrame检测器
        detection_result = self.df_detector.detect_headers(worksheet, region)
        row_info = detection_result["row_headers"]
        
        row_headers = {
            "start_col": row_info["header_col"] + 1,  # 转换为1-based索引
            "end_col": row_info["header_col"] + 1,
            "headers": row_info["headers"],
            "detection_score": row_info["score"],
            "density": row_info["density"],
            "text_ratio": row_info["text_ratio"]
        }
        
        return row_headers


class ColumnHeaderDetector(BaseTool):
    """列表头检测器（基于DataFrame改进版）"""
    
    def __init__(self):
        super().__init__(
            name="ColumnHeaderDetector",
            description="检测和识别列表头（基于DataFrame智能检测）",
            tool_type="detector"
        )
        # 使用内置的检测器
        self.df_detector = DataFrameBasedHeaderDetector()
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, region: Dict[str, Any]) -> Dict[str, Any]:
        """
        检测列表头（基于DataFrame智能识别）
        
        Args:
            worksheet: 工作表对象
            region: 表格区域
            
        Returns:
            列表头信息
        """
        # 使用改进的DataFrame检测器
        detection_result = self.df_detector.detect_headers(worksheet, region)
        col_info = detection_result["column_headers"]
        
        col_headers = {
            "start_row": col_info["header_row"] + 1,  # 转换为1-based索引
            "end_row": col_info["header_row"] + 1,
            "headers": col_info["headers"],
            "detection_score": col_info["score"],
            "density": col_info["density"],
            "text_ratio": col_info["text_ratio"]
        }
        
        return col_headers


class HierarchicalHeaderParser(BaseTool):
    """层级表头解析器（智能识别）"""
    
    def __init__(self):
        super().__init__(
            name="HierarchicalHeaderParser",
            description="智能解析多级层次表头结构",
            tool_type="parser"
        )
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                merged_cells: List[Dict[str, Any]],
                region: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        智能解析层级表头
        
        使用多种启发式方法：
        1. 合并单元格模式
        2. 空值分布
        3. 数据类型变化
        4. 样式信息（粗体、背景色）
        
        Args:
            worksheet: 工作表对象
            merged_cells: 合并单元格信息
            
        Returns:
            层级表头结构
        """
        hierarchy = {
            "row_header_levels": [],
            "col_header_levels": [],
            "row_header_end_col": 0,
            "col_header_end_row": 0,
            "hierarchy_tree": {}
        }
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        min_row = 1
        min_col = 1
        if region:
            min_row = max(1, int(region.get("min_row", 1)))
            min_col = max(1, int(region.get("min_col", 1)))
            max_row = min(max_row, int(region.get("max_row", max_row)))
            max_col = min(max_col, int(region.get("max_col", max_col)))
        
        # 1. 检测列表头的结束行（数据开始行）
        col_header_end_row = self._detect_col_header_end(worksheet, merged_cells, min_row, max_row, min_col, max_col)
        hierarchy["col_header_end_row"] = col_header_end_row
        
        # 2. 检测行表头的结束列（数据开始列）
        row_header_end_col = self._detect_row_header_end(worksheet, merged_cells, min_row, max_row, min_col, max_col, col_header_end_row)
        hierarchy["row_header_end_col"] = row_header_end_col
        
        # 3. 解析列表头层级
        col_levels = self._parse_col_header_hierarchy(
            worksheet, merged_cells, col_header_end_row, row_header_end_col, min_row, max_col
        )
        hierarchy["col_header_levels"] = col_levels
        
        # 4. 解析行表头层级
        row_levels = self._parse_row_header_hierarchy(
            worksheet, merged_cells, col_header_end_row, row_header_end_col, max_row
        )
        hierarchy["row_header_levels"] = row_levels
        
        # 5. 检测主表格之后的独立小表格（如评估等级说明表）
        additional_tables = self._detect_additional_tables(
            worksheet, merged_cells, col_header_end_row, max_row, min_col, max_col
        )
        if additional_tables:
            # 将小表格的表头添加到列表头层级中
            for table_info in additional_tables:
                table_col_levels = table_info.get("col_header_levels", [])
                if table_col_levels:
                    # 合并到现有的列表头层级
                    for level_idx, level_headers in enumerate(table_col_levels):
                        if level_idx < len(col_levels):
                            col_levels[level_idx].extend(level_headers)
                        else:
                            col_levels.append(level_headers)
            hierarchy["col_header_levels"] = col_levels
            hierarchy["additional_tables"] = additional_tables
        
        return hierarchy
    
    def _detect_additional_tables(self, worksheet, merged_cells, main_header_end_row, max_row, min_col, max_col):
        """检测主表格之后的独立小表格（如评估等级说明表）
        
        Args:
            worksheet: 工作表对象
            merged_cells: 合并单元格信息
            main_header_end_row: 主表格的表头结束行
            max_row: 最大行号
            min_col: 最小列号
            max_col: 最大列号
            
        Returns:
            额外表格列表，每个包含表头层级信息
        """
        additional_tables = []
        
        # 从主表格结束后很多行开始查找（跳过数据区）
        search_start = main_header_end_row + 20
        if search_start > max_row - 5:  # 至少需要5行空间
            return additional_tables
        
        # 查找可能的小表格表头行
        # 特征：某一行有多个短文本单元格（像表头），下面几行有数据
        processed_rows = set()
        
        row_idx = search_start
        while row_idx <= max_row - 3:
            # 跳过已处理的行
            if row_idx in processed_rows:
                row_idx += 1
                continue
            
            # 检查这一行是否像表头
            header_cells = []
            for col_idx in range(2, min(max_col + 1, 10)):  # 检查前几列
                cell = worksheet.cell(row_idx, col_idx)
                val = cell.value
                if val and isinstance(val, str):
                    val_str = val.strip()
                    # 表头特征：短文本（<20字符），包含关键词
                    if len(val_str) < 20 and any(kw in val_str for kw in ['定义', '等级', '评估', '说明', '标准', '名称', '类型']):
                        header_cells.append({
                            "row": row_idx,
                            "col": col_idx,
                            "value": val_str,
                            "span": 1,
                            "merged": False,
                            "source": "additional_table"
                        })
            
            # 如果找到至少2个像表头的单元格
            if len(header_cells) >= 2:
                # 检查下面几行是否有数据
                has_data_below = False
                for check_row in range(row_idx + 1, min(row_idx + 6, max_row + 1)):
                    data_count = 0
                    for col_idx in range(2, min(max_col + 1, 10)):
                        cell = worksheet.cell(check_row, col_idx)
                        if cell.value is not None and str(cell.value).strip():
                            data_count += 1
                    if data_count >= 2:
                        has_data_below = True
                        break
                
                if has_data_below:
                    # 找到一个小表格
                    table_info = {
                        "start_row": row_idx,
                        "header_row": row_idx,
                        "col_header_levels": [header_cells],  # 只有一层表头
                        "type": "additional_table"
                    }
                    additional_tables.append(table_info)
                    # 标记这个表格的行为已处理
                    for r in range(row_idx, min(row_idx + 6, max_row + 1)):
                        processed_rows.add(r)
                    row_idx += 6
                else:
                    row_idx += 1
            else:
                row_idx += 1
        
        return additional_tables
    
    def _detect_col_header_end(self, worksheet, merged_cells, min_row, max_row, min_col, max_col):
        """检测列表头结束行（改进版：考虑多层表头和说明行）"""
        # 启发式1: 查找第一行完全数值型的行（在有效区域内）
        # 但要排除说明行（通常包含长文本）
        for row_idx in range(min_row, min(min_row + 12, max_row) + 1):
            numeric_count = 0
            total_count = 0
            has_long_text = False
            
            for col_idx in range(max(min_col, 2), max_col + 1):  # 从col 2开始，跳过可能的行号列
                cell = worksheet.cell(row_idx, col_idx)
                if cell.value is not None:
                    total_count += 1
                    if isinstance(cell.value, (int, float)):
                        numeric_count += 1
                    # 检查是否有长文本（说明行特征）
                    elif isinstance(cell.value, str) and len(cell.value) > 50:
                        has_long_text = True
            
            # 如果超过70%是数字，且不是说明行，认为是数据行
            if total_count > 0 and numeric_count / total_count > 0.7 and not has_long_text:
                return row_idx - 1
        
        # 启发式2: 基于合并单元格（表头区域通常有合并单元格）
        if merged_cells:
            # 找到前几行的最大合并行
            header_merged = [mc for mc in merged_cells if mc["min_row"] <= (min_row + 6)]
            if header_merged:
                max_merged_row = max(mc["max_row"] for mc in header_merged if mc["min_row"] <= 5)
                # 表头通常在前4-5行
                return min(max_merged_row, min_row + 4)
        
        # 启发式3: 查找包含大量文本的连续行（可能是多层表头+说明行）
        # 通常前1-4行是表头，第4行可能是说明行
        text_rows = []
        for row_idx in range(min_row, min(min_row + 6, max_row) + 1):
            text_count = 0
            total_count = 0
            for col_idx in range(max(min_col, 2), min(max_col + 1, 15)):  # 检查前几列
                cell = worksheet.cell(row_idx, col_idx)
                if cell.value is not None and str(cell.value).strip():
                    total_count += 1
                    if isinstance(cell.value, str):
                        text_count += 1
            
            if total_count > 0 and text_count / total_count > 0.5:
                text_rows.append(row_idx)
        
        # 如果有连续的文本行，最后一个可能是说明行
        if len(text_rows) >= 3:
            return text_rows[-1]
        
        # 默认：前3行是表头
        return min_row + 2
    
    def _detect_row_header_end(self, worksheet, merged_cells, min_row, max_row, min_col, max_col, col_header_end_row):
        """检测行表头结束列（改进：正确识别无行表头的情况）"""
        # 启发式0: 如果min_col > 1，说明Excel的Col 1完全空白被排除了
        # 这种情况通常是纯列表头的表格（没有行表头）
        if min_col > 1:
            # 检查实际Excel的Col 1 (openpyxl col 1)是否全空
            excel_col1_empty = True
            for row_idx in range(col_header_end_row + 1, min(col_header_end_row + 20, max_row + 1)):
                cell = worksheet.cell(row_idx, 1)  # 强制检查Excel Col 1
                if cell.value is not None and str(cell.value).strip():
                    excel_col1_empty = False
                    break
            
            if excel_col1_empty:
                return 0  # Excel Col 1空，说明没有行表头
        
        # 启发式1: 检查表格区域的第一列是否全空
        first_col_values = []
        for row_idx in range(col_header_end_row + 1, min(col_header_end_row + 20, max_row + 1)):
            cell = worksheet.cell(row_idx, min_col)
            if cell.value is not None and str(cell.value).strip():
                first_col_values.append(cell.value)
        
        # 如果第一列基本全空（<5%有值），说明没有行表头
        if len(first_col_values) < 1:
            return 0
        
        # 启发式1: 查找第一列完全数值型的列（在有效区域内）
        for col_idx in range(min_col, min(min_col + 9, max_col) + 1):
            numeric_count = 0
            total_count = 0
            
            for row_idx in range(col_header_end_row + 1, max_row + 1):
                cell = worksheet.cell(row_idx, col_idx)
                if cell.value is not None:
                    total_count += 1
                    if isinstance(cell.value, (int, float)):
                        numeric_count += 1
            
            # 如果超过70%是数字，认为是数据列
            if total_count > 0 and numeric_count / total_count > 0.7:
                # 如果第一列就是数值列，说明没有行表头
                if col_idx == min_col:
                    return 0
                return col_idx - 1
        
        # 启发式2: 基于合并单元格的垂直跨度
        if merged_cells:
            row_merged = [mc for mc in merged_cells 
                         if mc["min_col"] <= 3 and mc["max_row"] - mc["min_row"] > 0 
                         and mc["min_row"] > col_header_end_row]  # 确保在数据区
            if row_merged:
                max_merged_col = max(mc["max_col"] for mc in row_merged)
                return max_merged_col
        
        # 检查：如果前几列都是文本，才认为有行表头
        text_cols = 0
        for col_idx in range(min_col, min(min_col + 3, max_col + 1)):
            text_count = 0
            total_count = 0
            for row_idx in range(col_header_end_row + 1, min(col_header_end_row + 10, max_row + 1)):
                cell = worksheet.cell(row_idx, col_idx)
                if cell.value is not None:
                    total_count += 1
                    if isinstance(cell.value, str):
                        text_count += 1
            if total_count > 0 and text_count / total_count > 0.5:
                text_cols += 1
        
        # 如果有连续的文本列，才认为是行表头
        if text_cols >= 2:
            return min_col + text_cols - 1
        
        # 默认：没有行表头
        return 0
    
    def _parse_col_header_hierarchy(self, worksheet, merged_cells, header_end_row, start_col, min_row, max_col):
        """解析列表头的层级结构（增强版：支持非合并单元格嵌套表头）"""
        levels = []
        
        # 首先尝试基于合并单元格的解析
        merged_levels = self._parse_merged_col_headers(worksheet, merged_cells, header_end_row, start_col, max_col)
        
        # 然后尝试基于模式识别的解析
        pattern_levels = self._parse_pattern_based_col_headers(worksheet, header_end_row, start_col, min_row, max_col)
        
        # 合并两种解析结果
        levels = self._merge_col_header_levels(worksheet, merged_levels, pattern_levels, header_end_row, start_col, max_col)
        
        return levels
    
    def _parse_merged_col_headers(self, worksheet, merged_cells, header_end_row, start_col, max_col):
        """基于合并单元格解析列表头（改进：包含行表头区域的表头层）"""
        levels = []
        
        for row_idx in range(1, header_end_row + 1):
            level_headers = []
            
            # 【改进】在表头行中，从Col 2开始解析（包含行表头区域上方的列表头）
            # start_col 是行表头结束列，但在表头行中，这些列也可能包含列表头
            start_parse_col = 2 if row_idx <= header_end_row else (start_col + 1)
            
            for col_idx in range(start_parse_col, max_col + 1):
                cell = worksheet.cell(row_idx, col_idx)
                
                # 查找是否在合并单元格中
                merged_info = None
                for mc in merged_cells:
                    if (mc["min_row"] <= row_idx <= mc["max_row"] and
                        mc["min_col"] <= col_idx <= mc["max_col"]):
                        merged_info = mc
                        break
                
                header_info = {
                    "row": row_idx,
                    "col": col_idx,
                    "value": cell.value,
                    "merged": merged_info is not None,
                    "span": 1,
                    "source": "merged"
                }
                
                if merged_info:
                    header_info["span"] = merged_info["max_col"] - merged_info["min_col"] + 1
                
                level_headers.append(header_info)
            
            levels.append(level_headers)
        
        return levels
    
    def _parse_pattern_based_col_headers(self, worksheet, header_end_row, start_col, min_row, max_col):
        """基于模式识别解析列表头（改进：包含行表头区域的表头层）"""
        levels = []
        
        # 1. 检测表头模式
        header_patterns = self._detect_header_patterns(worksheet, header_end_row, start_col, min_row, max_col)
        
        # 2. 基于模式构建层级结构
        for row_idx in range(min_row, header_end_row + 1):
            level_headers = []
            
            # 【改进】在表头行中，从Col 2开始解析
            start_parse_col = 2 if row_idx <= header_end_row else (start_col + 1)
            
            for col_idx in range(start_parse_col, max_col + 1):
                cell = worksheet.cell(row_idx, col_idx)
                
                # 分析表头特征
                header_features = self._analyze_header_features(worksheet, row_idx, col_idx, header_patterns)
                
                header_info = {
                    "row": row_idx,
                    "col": col_idx,
                    "value": cell.value,
                    "merged": False,
                    "span": header_features.get("span", 1),
                    "source": "pattern",
                    "features": header_features
                }
                
                level_headers.append(header_info)
            
            levels.append(level_headers)
        
        return levels
    
    def _detect_header_patterns(self, worksheet, header_end_row, start_col, min_row, max_col):
        """检测表头模式"""
        patterns = {
            "repetition_patterns": [],
            "grouping_patterns": [],
            "hierarchical_patterns": [],
            "semantic_patterns": []
        }
        
        # 1. 检测重复模式
        patterns["repetition_patterns"] = self._detect_repetition_patterns(worksheet, header_end_row, start_col, min_row, max_col)
        
        # 2. 检测分组模式
        patterns["grouping_patterns"] = self._detect_grouping_patterns(worksheet, header_end_row, start_col, min_row, max_col)
        
        # 3. 检测层级模式
        patterns["hierarchical_patterns"] = self._detect_hierarchical_patterns(worksheet, header_end_row, start_col, min_row, max_col)
        
        # 4. 检测语义模式
        patterns["semantic_patterns"] = self._detect_semantic_patterns(worksheet, header_end_row, start_col, min_row, max_col)
        
        return patterns
    
    def _detect_repetition_patterns(self, worksheet, header_end_row, start_col, min_row, max_col):
        """检测重复模式"""
        patterns = []
        
        # 检测列级别的重复
        for row_idx in range(min_row, header_end_row + 1):
            row_values = []
            for col_idx in range(start_col + 1, max_col + 1):
                cell_value = worksheet.cell(row_idx, col_idx).value
                row_values.append(str(cell_value) if cell_value else "")
            
            # 查找重复的子序列
            repetition_groups = self._find_repetition_groups(row_values)
            if repetition_groups:
                patterns.append({
                    "row": row_idx,
                    "type": "repetition",
                    "groups": repetition_groups
                })
        
        return patterns
    
    def _detect_grouping_patterns(self, worksheet, header_end_row, start_col, min_row, max_col):
        """检测分组模式"""
        patterns = []
        
        # 检测基于空值或分隔符的分组
        for row_idx in range(min_row, header_end_row + 1):
            groups = []
            current_group = []
            
            for col_idx in range(start_col + 1, max_col + 1):
                cell_value = worksheet.cell(row_idx, col_idx).value
                
                if cell_value is None or str(cell_value).strip() == "":
                    if current_group:
                        groups.append(current_group)
                        current_group = []
                else:
                    current_group.append(col_idx)
            
            if current_group:
                groups.append(current_group)
            
            if len(groups) > 1:
                patterns.append({
                    "row": row_idx,
                    "type": "grouping",
                    "groups": groups
                })
        
        return patterns
    
    def _detect_hierarchical_patterns(self, worksheet, header_end_row, start_col, min_row, max_col):
        """检测层级模式"""
        patterns = []
        
        # 检测多行表头的层级关系
        for row_idx in range(min_row, header_end_row):
            current_row_values = []
            next_row_values = []
            
            for col_idx in range(start_col + 1, max_col + 1):
                current_value = worksheet.cell(row_idx, col_idx).value
                next_value = worksheet.cell(row_idx + 1, col_idx).value
                current_row_values.append(str(current_value) if current_value else "")
                next_row_values.append(str(next_value) if next_value else "")
            
            # 分析层级关系
            hierarchy_relation = self._analyze_hierarchy_relation(current_row_values, next_row_values)
            if hierarchy_relation:
                patterns.append({
                    "row": row_idx,
                    "type": "hierarchical",
                    "relation": hierarchy_relation
                })
        
        return patterns
    
    def _detect_semantic_patterns(self, worksheet, header_end_row, start_col, min_row, max_col):
        """检测语义模式"""
        patterns = []
        
        # 检测常见的表头语义模式
        semantic_keywords = {
            "category": ["类别", "分类", "类型", "Category", "Type"],
            "subcategory": ["子类", "细分", "Subcategory", "Sub"],
            "measure": ["数量", "金额", "比例", "Measure", "Count", "Amount"],
            "time": ["时间", "日期", "年", "月", "Time", "Date", "Year", "Month"],
            "region": ["地区", "区域", "Region", "Area", "Location"]
        }
        
        for row_idx in range(min_row, header_end_row + 1):
            row_semantics = []
            
            for col_idx in range(start_col + 1, max_col + 1):
                cell_value = worksheet.cell(row_idx, col_idx).value
                if cell_value:
                    cell_text = str(cell_value).lower()
                    detected_semantics = []
                    
                    for semantic_type, keywords in semantic_keywords.items():
                        for keyword in keywords:
                            if keyword.lower() in cell_text:
                                detected_semantics.append(semantic_type)
                                break
                    
                    row_semantics.append({
                        "col": col_idx,
                        "value": cell_value,
                        "semantics": detected_semantics
                    })
            
            if any(item["semantics"] for item in row_semantics):
                patterns.append({
                    "row": row_idx,
                    "type": "semantic",
                    "semantics": row_semantics
                })
        
        return patterns
    
    def _parse_row_header_hierarchy(self, worksheet, merged_cells, start_row, header_end_col, max_row):
        """解析行表头的层级结构（改进：正确处理无行表头的情况）"""
        # 如果没有行表头（header_end_col = 0），直接返回空列表
        if header_end_col == 0:
            return []
        
        levels = []
        
        for col_idx in range(1, header_end_col + 1):
            level_headers = []
            
            for row_idx in range(start_row + 1, max_row + 1):
                cell = worksheet.cell(row_idx, col_idx)
                
                # 查找是否在合并单元格中
                merged_info = None
                for mc in merged_cells:
                    if (mc["min_row"] <= row_idx <= mc["max_row"] and
                        mc["min_col"] <= col_idx <= mc["max_col"]):
                        merged_info = mc
                        break
                
                header_info = {
                    "row": row_idx,
                    "col": col_idx,
                    "value": cell.value,
                    "merged": merged_info is not None,
                    "span": 1
                }
                
                if merged_info:
                    header_info["span"] = merged_info["max_row"] - merged_info["min_row"] + 1
                
                level_headers.append(header_info)
            
            levels.append(level_headers)
        
        return levels
    
    def _detect_font_patterns(self, worksheet, start_row, end_row, start_col, end_col):
        """检测字体差异模式"""
        patterns = []
        
        # 收集字体信息
        font_info = []
        for row_idx in range(start_row, end_row + 1):
            row_fonts = []
            
            for col_idx in range(start_col, end_col + 1):
                cell = worksheet.cell(row_idx, col_idx)
                cell_value = cell.value
                
                if cell_value:
                    font_style = self._get_cell_font_style(cell)
                    row_fonts.append({
                        "col": col_idx,
                        "value": cell_value,
                        "font_style": font_style
                    })
            
            if row_fonts:
                font_info.append({
                    "row": row_idx,
                    "fonts": row_fonts
                })
        
        # 分析字体模式
        for i, row_info in enumerate(font_info):
            row_idx = row_info["row"]
            fonts = row_info["fonts"]
            
            # 检测字体变化模式
            font_analysis = self._analyze_font_patterns(fonts, font_info, i)
            if font_analysis["has_pattern"]:
                patterns.append({
                    "row": row_idx,
                    "type": "font",
                    "analysis": font_analysis,
                    "details": fonts
                })
        
        return patterns
    
    def _detect_repetition_patterns_enhanced(self, worksheet, start_row, end_row, start_col, end_col):
        """检测增强的重复模式（如每N行一个子表）"""
        patterns = []
        
        # 检测行级别的重复模式
        row_patterns = self._detect_row_repetition_patterns(worksheet, start_row, end_row, start_col, end_col)
        patterns.extend(row_patterns)
        
        # 检测列级别的重复模式
        col_patterns = self._detect_col_repetition_patterns(worksheet, start_row, end_row, start_col, end_col)
        patterns.extend(col_patterns)
        
        # 检测周期性模式
        periodic_patterns = self._detect_periodic_patterns(worksheet, start_row, end_row, start_col, end_col)
        patterns.extend(periodic_patterns)
        
        return patterns
    
    def _detect_separator_patterns(self, worksheet, start_row, end_row, start_col, end_col):
        """检测分隔符模式"""
        patterns = []
        
        for row_idx in range(start_row, end_row + 1):
            row_values = []
            for col_idx in range(start_col, end_col + 1):
                cell_value = worksheet.cell(row_idx, col_idx).value
                row_values.append(cell_value)
            
            # 判断是否为分隔行
            if self._is_separator_row(row_values):
                separator_type = self._classify_separator_type(row_values)
                patterns.append({
                    "row": row_idx,
                    "type": "separator",
                    "separator_type": separator_type,
                    "values": row_values
                })
        
        return patterns
    
    def _calculate_indent_level(self, text):
        """计算文本的缩进级别"""
        if not text:
            return 0
        
        # 计算前导空格数
        leading_spaces = len(text) - len(text.lstrip())
        
        # 计算前导制表符数
        leading_tabs = 0
        for char in text:
            if char == '\t':
                leading_tabs += 1
            else:
                break
        
        # 综合计算缩进级别
        return leading_spaces + leading_tabs * 4  # 假设制表符等于4个空格
    
    def _get_cell_alignment(self, cell):
        """获取单元格对齐方式"""
        try:
            alignment = cell.alignment
            return {
                "horizontal": alignment.horizontal if alignment else None,
                "vertical": alignment.vertical if alignment else None,
                "indent": alignment.indent if alignment else 0
            }
        except:
            return {"horizontal": None, "vertical": None, "indent": 0}
    
    def _get_cell_font_style(self, cell):
        """获取单元格字体样式"""
        try:
            font = cell.font
            return {
                "bold": font.bold if font else False,
                "italic": font.italic if font else False,
                "size": font.size if font else None,
                "color": font.color.rgb if font and font.color else None,
                "name": font.name if font else None
            }
        except:
            return {"bold": False, "italic": False, "size": None, "color": None, "name": None}
    
    def _analyze_row_indentation(self, row_indentation):
        """分析行的缩进模式"""
        analysis = {
            "has_pattern": False,
            "indent_levels": [],
            "alignment_patterns": [],
            "is_nested": False
        }
        
        # 收集缩进级别
        indent_levels = [item["indent_level"] for item in row_indentation if item["has_indent"]]
        if indent_levels:
            analysis["indent_levels"] = indent_levels
            analysis["has_pattern"] = True
            
            # 判断是否为嵌套结构
            if len(set(indent_levels)) > 1:  # 有多个不同的缩进级别
                analysis["is_nested"] = True
        
        # 收集对齐模式
        alignments = [item["alignment"] for item in row_indentation]
        if alignments:
            analysis["alignment_patterns"] = alignments
        
        return analysis
    
    def _analyze_font_patterns(self, fonts, all_font_info, current_index):
        """分析字体模式"""
        analysis = {
            "has_pattern": False,
            "font_variations": [],
            "is_header": False,
            "is_subheader": False
        }
        
        # 分析当前行的字体特征
        current_fonts = [f["font_style"] for f in fonts]
        
        # 检测粗体模式（可能是表头）
        bold_count = sum(1 for f in current_fonts if f.get("bold", False))
        if bold_count > len(current_fonts) * 0.5:  # 超过50%是粗体
            analysis["is_header"] = True
            analysis["has_pattern"] = True
        
        # 检测字体大小变化
        font_sizes = [f.get("size") for f in current_fonts if f.get("size")]
        if font_sizes and len(set(font_sizes)) > 1:
            analysis["font_variations"] = font_sizes
            analysis["has_pattern"] = True
        
        # 与前后行比较
        if current_index > 0 and current_index < len(all_font_info) - 1:
            prev_fonts = [f["font_style"] for f in all_font_info[current_index - 1]["fonts"]]
            next_fonts = [f["font_style"] for f in all_font_info[current_index + 1]["fonts"]]
            
            # 检测子表头模式（当前行字体与前后行不同）
            if self._is_font_different(current_fonts, prev_fonts) and self._is_font_different(current_fonts, next_fonts):
                analysis["is_subheader"] = True
                analysis["has_pattern"] = True
        
        return analysis
    
    def _detect_row_repetition_patterns(self, worksheet, start_row, end_row, start_col, end_col):
        """检测行级别的重复模式"""
        patterns = []
        
        # 收集所有行的特征
        row_features = []
        for row_idx in range(start_row, end_row + 1):
            row_data = []
            for col_idx in range(start_col, end_col + 1):
                cell_value = worksheet.cell(row_idx, col_idx).value
                row_data.append(str(cell_value) if cell_value else "")
            
            # 计算行特征
            features = self._calculate_row_features(row_data)
            row_features.append({
                "row": row_idx,
                "features": features,
                "data": row_data
            })
        
        # 检测重复模式
        repetition_groups = self._find_row_repetition_groups(row_features)
        for group in repetition_groups:
            patterns.append({
                "type": "row_repetition",
                "pattern_length": group["pattern_length"],
                "repetitions": group["repetitions"],
                "start_row": group["start_row"],
                "end_row": group["end_row"]
            })
        
        return patterns
    
    def _detect_col_repetition_patterns(self, worksheet, start_row, end_row, start_col, end_col):
        """检测列级别的重复模式"""
        patterns = []
        
        # 收集所有列的特征
        col_features = []
        for col_idx in range(start_col, end_col + 1):
            col_data = []
            for row_idx in range(start_row, end_row + 1):
                cell_value = worksheet.cell(row_idx, col_idx).value
                col_data.append(str(cell_value) if cell_value else "")
            
            # 计算列特征
            features = self._calculate_col_features(col_data)
            col_features.append({
                "col": col_idx,
                "features": features,
                "data": col_data
            })
        
        # 检测重复模式
        repetition_groups = self._find_col_repetition_groups(col_features)
        for group in repetition_groups:
            patterns.append({
                "type": "col_repetition",
                "pattern_length": group["pattern_length"],
                "repetitions": group["repetitions"],
                "start_col": group["start_col"],
                "end_col": group["end_col"]
            })
        
        return patterns
    
    def _detect_periodic_patterns(self, worksheet, start_row, end_row, start_col, end_col):
        """检测周期性模式"""
        patterns = []
        
        # 检测每N行重复的模式
        for period in range(2, min(20, (end_row - start_row + 1) // 2)):
            if self._is_periodic_pattern(worksheet, start_row, end_row, start_col, end_col, period):
                patterns.append({
                    "type": "periodic",
                    "period": period,
                    "start_row": start_row,
                    "end_row": end_row
                })
        
        return patterns
    
    def _is_separator_row(self, row_values):
        """判断是否为分隔行"""
        # 检查是否为空行
        if all(v is None or str(v).strip() == "" for v in row_values):
            return True
        
        # 检查是否为分隔符行（如：---, ===, ***等）
        separator_chars = ["-", "=", "*", "_", "|"]
        for value in row_values:
            if value and isinstance(value, str):
                value_str = str(value).strip()
                if len(value_str) > 2 and all(c in separator_chars for c in value_str):
                    return True
        
        # 检查是否为重复字符行
        for value in row_values:
            if value and isinstance(value, str):
                value_str = str(value).strip()
                if len(value_str) > 1 and len(set(value_str)) == 1:
                    return True
        
        return False
    
    def _classify_separator_type(self, row_values):
        """分类分隔符类型"""
        if all(v is None or str(v).strip() == "" for v in row_values):
            return "empty"
        
        # 检查分隔符字符
        separator_chars = ["-", "=", "*", "_", "|"]
        for value in row_values:
            if value and isinstance(value, str):
                value_str = str(value).strip()
                for char in separator_chars:
                    if char in value_str:
                        return f"separator_{char}"
        
        return "unknown"
    
    def _calculate_row_features(self, row_data):
        """计算行特征"""
        features = {
            "non_empty_count": sum(1 for v in row_data if v and str(v).strip()),
            "empty_count": sum(1 for v in row_data if not v or str(v).strip() == ""),
            "numeric_count": sum(1 for v in row_data if v and str(v).replace(".", "").replace("-", "").isdigit()),
            "text_count": sum(1 for v in row_data if v and isinstance(v, str) and not str(v).replace(".", "").replace("-", "").isdigit()),
            "first_non_empty": next((i for i, v in enumerate(row_data) if v and str(v).strip()), -1),
            "last_non_empty": next((i for i, v in enumerate(reversed(row_data)) if v and str(v).strip()), -1)
        }
        
        return features
    
    def _calculate_col_features(self, col_data):
        """计算列特征"""
        features = {
            "non_empty_count": sum(1 for v in col_data if v and str(v).strip()),
            "empty_count": sum(1 for v in col_data if not v or str(v).strip() == ""),
            "numeric_count": sum(1 for v in col_data if v and str(v).replace(".", "").replace("-", "").isdigit()),
            "text_count": sum(1 for v in col_data if v and isinstance(v, str) and not str(v).replace(".", "").replace("-", "").isdigit()),
            "first_non_empty": next((i for i, v in enumerate(col_data) if v and str(v).strip()), -1),
            "last_non_empty": next((i for i, v in enumerate(reversed(col_data)) if v and str(v).strip()), -1)
        }
        
        return features
    
    def _find_row_repetition_groups(self, row_features):
        """查找行重复组"""
        groups = []
        n = len(row_features)
        
        # 查找重复模式
        for pattern_len in range(2, n // 2 + 1):
            for start in range(n - pattern_len + 1):
                pattern = row_features[start:start + pattern_len]
                
                # 检查这个模式是否重复
                repetitions = [start]
                for i in range(start + pattern_len, n - pattern_len + 1, pattern_len):
                    if self._is_row_pattern_match(pattern, row_features[i:i + pattern_len]):
                        repetitions.append(i)
                    else:
                        break
                
                if len(repetitions) > 1:
                    groups.append({
                        "pattern_length": pattern_len,
                        "repetitions": repetitions,
                        "start_row": row_features[start]["row"],
                        "end_row": row_features[start + pattern_len - 1]["row"]
                    })
        
        return groups
    
    def _find_col_repetition_groups(self, col_features):
        """查找列重复组"""
        groups = []
        n = len(col_features)
        
        # 查找重复模式
        for pattern_len in range(2, n // 2 + 1):
            for start in range(n - pattern_len + 1):
                pattern = col_features[start:start + pattern_len]
                
                # 检查这个模式是否重复
                repetitions = [start]
                for i in range(start + pattern_len, n - pattern_len + 1, pattern_len):
                    if self._is_col_pattern_match(pattern, col_features[i:i + pattern_len]):
                        repetitions.append(i)
                    else:
                        break
                
                if len(repetitions) > 1:
                    groups.append({
                        "pattern_length": pattern_len,
                        "repetitions": repetitions,
                        "start_col": col_features[start]["col"],
                        "end_col": col_features[start + pattern_len - 1]["col"]
                    })
        
        return groups
    
    def _is_periodic_pattern(self, worksheet, start_row, end_row, start_col, end_col, period):
        """判断是否为周期性模式"""
        # 检查每period行是否相似
        for base_row in range(start_row, end_row - period + 1):
            base_data = []
            for col_idx in range(start_col, end_col + 1):
                cell_value = worksheet.cell(base_row, col_idx).value
                base_data.append(str(cell_value) if cell_value else "")
            
            # 检查后续的period行是否与base_row相似
            is_periodic = True
            for offset in range(period, end_row - base_row + 1, period):
                if base_row + offset > end_row:
                    break
                
                compare_data = []
                for col_idx in range(start_col, end_col + 1):
                    cell_value = worksheet.cell(base_row + offset, col_idx).value
                    compare_data.append(str(cell_value) if cell_value else "")
                
                if not self._is_row_data_similar(base_data, compare_data):
                    is_periodic = False
                    break
            
            if is_periodic:
                return True
        
        return False
    
    def _is_row_pattern_match(self, pattern1, pattern2):
        """判断两个行模式是否匹配"""
        if len(pattern1) != len(pattern2):
            return False
        
        for p1, p2 in zip(pattern1, pattern2):
            if not self._is_row_features_similar(p1["features"], p2["features"]):
                return False
        
        return True
    
    def _is_col_pattern_match(self, pattern1, pattern2):
        """判断两个列模式是否匹配"""
        if len(pattern1) != len(pattern2):
            return False
        
        for p1, p2 in zip(pattern1, pattern2):
            if not self._is_col_features_similar(p1["features"], p2["features"]):
                return False
        
        return True
    
    def _is_row_features_similar(self, features1, features2):
        """判断两个行特征是否相似"""
        # 比较关键特征
        key_features = ["non_empty_count", "empty_count", "numeric_count", "text_count"]
        
        for feature in key_features:
            if abs(features1[feature] - features2[feature]) > 1:  # 允许1个单元格的差异
                return False
        
        return True
    
    def _is_col_features_similar(self, features1, features2):
        """判断两个列特征是否相似"""
        # 比较关键特征
        key_features = ["non_empty_count", "empty_count", "numeric_count", "text_count"]
        
        for feature in key_features:
            if abs(features1[feature] - features2[feature]) > 1:  # 允许1个单元格的差异
                return False
        
        return True
    
    def _is_row_data_similar(self, data1, data2):
        """判断两行数据是否相似"""
        if len(data1) != len(data2):
            return False
        
        # 计算相似度
        matches = sum(1 for d1, d2 in zip(data1, data2) if d1 == d2)
        similarity = matches / len(data1)
        
        return similarity > 0.7  # 70%以上相似度认为是相似的
    
    def _is_font_different(self, fonts1, fonts2):
        """判断两组字体是否不同"""
        if len(fonts1) != len(fonts2):
            return True
        
        for f1, f2 in zip(fonts1, fonts2):
            if (f1.get("bold") != f2.get("bold") or 
                f1.get("size") != f2.get("size") or
                f1.get("name") != f2.get("name")):
                return True
        
        return False
    
    def _find_repetition_groups(self, values):
        """查找重复的子序列"""
        groups = []
        n = len(values)
        
        # 查找重复模式
        for pattern_len in range(2, n // 2 + 1):
            for start in range(n - pattern_len + 1):
                pattern = values[start:start + pattern_len]
                if pattern.count("") < len(pattern) * 0.5:  # 避免空值模式
                    # 检查这个模式是否重复
                    repetitions = [start]
                    for i in range(start + pattern_len, n - pattern_len + 1, pattern_len):
                        if values[i:i + pattern_len] == pattern:
                            repetitions.append(i)
                        else:
                            break
                    
                    if len(repetitions) > 1:
                        groups.append({
                            "pattern": pattern,
                            "positions": repetitions,
                            "length": pattern_len
                        })
        
        return groups
    
    def _analyze_hierarchy_relation(self, parent_values, child_values):
        """分析层级关系"""
        # 检测父子关系
        parent_child_pairs = []
        
        for i, (parent_val, child_val) in enumerate(zip(parent_values, child_values)):
            if parent_val and child_val and parent_val != child_val:
                # 检查是否是层级关系
                if self._is_hierarchical_relation(parent_val, child_val):
                    parent_child_pairs.append({
                        "col": i,
                        "parent": parent_val,
                        "child": child_val
                    })
        
        return parent_child_pairs if parent_child_pairs else None
    
    def _is_hierarchical_relation(self, parent, child):
        """判断是否是层级关系"""
        parent_lower = str(parent).lower()
        child_lower = str(child).lower()
        
        # 1. 包含关系
        if parent_lower in child_lower or child_lower in parent_lower:
            return True
        
        # 2. 语义关系
        semantic_relations = [
            ("总计", "小计"), ("总计", "合计"), ("总计", "分项"),
            ("类别", "子类"), ("分类", "细分"), ("类型", "子类型"),
            ("地区", "城市"), ("省份", "城市"), ("国家", "省份"),
            ("年份", "月份"), ("年度", "季度"), ("季度", "月份")
        ]
        
        for parent_keyword, child_keyword in semantic_relations:
            if parent_keyword in parent_lower and child_keyword in child_lower:
                return True
        
        return False
    
    def _analyze_header_features(self, worksheet, row_idx, col_idx, patterns):
        """分析表头特征"""
        features = {
            "span": 1,
            "is_group_header": False,
            "is_category_header": False,
            "hierarchy_level": 0,
            "semantic_type": None
        }
        
        # 基于模式分析特征
        for pattern_type, pattern_list in patterns.items():
            for pattern in pattern_list:
                if pattern["row"] == row_idx:
                    if pattern_type == "repetition_patterns":
                        # 检查是否在重复组中
                        for group in pattern["groups"]:
                            if col_idx - 1 in group["positions"]:
                                features["span"] = group["length"]
                                features["is_group_header"] = True
                    
                    elif pattern_type == "grouping_patterns":
                        # 检查是否在分组中
                        for group in pattern["groups"]:
                            if col_idx in group:
                                features["is_group_header"] = True
                    
                    elif pattern_type == "hierarchical_patterns":
                        # 检查层级关系
                        for relation in pattern["relation"]:
                            if relation["col"] == col_idx - 1:
                                features["hierarchy_level"] = 1
                    
                    elif pattern_type == "semantic_patterns":
                        # 检查语义类型
                        for semantic in pattern["semantics"]:
                            if semantic["col"] == col_idx:
                                features["semantic_type"] = semantic["semantics"]
                                if "category" in semantic["semantics"]:
                                    features["is_category_header"] = True
        
        return features
    
    def _merge_col_header_levels(self, worksheet, merged_levels, pattern_levels, header_end_row, start_col, max_col):
        """合并两种解析结果"""
        merged_result = []
        
        for row_idx in range(header_end_row):
            merged_row = []
            
            for col_idx in range(start_col + 1, max_col + 1):
                # 获取合并单元格信息
                # 计算在merged_levels中的索引（考虑start_col = 0的情况）
                merged_col_offset = col_idx - (start_col + 1) if start_col > 0 else col_idx - 2
                merged_info = merged_levels[row_idx][merged_col_offset] if (row_idx < len(merged_levels) and merged_col_offset >= 0 and merged_col_offset < len(merged_levels[row_idx])) else None
                # 获取模式信息
                pattern_col_offset = col_idx - (start_col + 1) if start_col > 0 else col_idx - 2
                pattern_info = pattern_levels[row_idx][pattern_col_offset] if (row_idx < len(pattern_levels) and pattern_col_offset >= 0 and pattern_col_offset < len(pattern_levels[row_idx])) else None
                
                # 合并信息
                if merged_info and merged_info["merged"]:
                    # 优先使用合并单元格信息
                    final_info = merged_info.copy()
                    final_info["source"] = "merged"
                elif pattern_info and pattern_info["features"]["span"] > 1:
                    # 使用模式识别的信息
                    final_info = pattern_info.copy()
                    final_info["source"] = "pattern"
                else:
                    # 使用基础信息
                    cell = worksheet.cell(row_idx + 1, col_idx)
                    final_info = {
                        "row": row_idx + 1,
                        "col": col_idx,
                        "value": cell.value,
                        "merged": False,
                        "span": 1,
                        "source": "basic"
                    }
                
                merged_row.append(final_info)
            
            merged_result.append(merged_row)
        
        return merged_result
    
    def _parse_row_header_hierarchy(self, worksheet, merged_cells, start_row, header_end_col, max_row):
        """解析行表头的层级结构（改进：正确处理无行表头的情况）"""
        # 如果没有行表头（header_end_col = 0），直接返回空列表
        if header_end_col == 0:
            return []
        
        levels = []
        
        for col_idx in range(1, header_end_col + 1):
            level_headers = []
            
            for row_idx in range(start_row + 1, max_row + 1):
                cell = worksheet.cell(row_idx, col_idx)
                
                # 查找是否在合并单元格中
                merged_info = None
                for mc in merged_cells:
                    if (mc["min_row"] <= row_idx <= mc["max_row"] and
                        mc["min_col"] <= col_idx <= mc["max_col"]):
                        merged_info = mc
                        break
                
                header_info = {
                    "row": row_idx,
                    "col": col_idx,
                    "value": cell.value,
                    "merged": merged_info is not None,
                    "span": 1
                }
                
                if merged_info:
                    header_info["span"] = merged_info["max_row"] - merged_info["min_row"] + 1
                
                level_headers.append(header_info)
            
            levels.append(level_headers)
        
        return levels


class MetaInfoExtractor(BaseTool):
    """Meta信息提取器"""
    
    def __init__(self):
        super().__init__(
            name="MetaInfoExtractor",
            description="提取表格的Meta元信息",
            tool_type="extractor"
        )
    
    def execute(self, row_headers: Dict[str, Any], 
                col_headers: Dict[str, Any],
                hierarchy: Dict[str, Any]) -> Dict[str, Any]:
        """
        提取Meta信息
        
        Args:
            row_headers: 行表头信息
            col_headers: 列表头信息
            hierarchy: 层级信息
            
        Returns:
            Meta信息字典
        """
        meta_info = {
            "row_header_names": row_headers.get("headers", []),
            "col_header_names": col_headers.get("headers", []),
            "hierarchy_levels": hierarchy.get("max_level", 1),
            "table_structure": {
                "has_row_headers": len(row_headers.get("headers", [])) > 0,
                "has_col_headers": len(col_headers.get("headers", [])) > 0,
                "is_hierarchical": hierarchy.get("max_level", 1) > 1
            }
        }
        
        return meta_info


# ============ 5. DataFrame提取工具 ============

class DataFrameBuilder(BaseTool):
    """DataFrame构建器"""
    
    def __init__(self):
        super().__init__(
            name="DataFrameBuilder",
            description="将Excel数据转换为DataFrame",
            tool_type="builder"
        )
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                region: Dict[str, Any],
                col_headers: Dict[str, Any]) -> pd.DataFrame:
        """
        构建DataFrame
        
        Args:
            worksheet: 工作表对象
            region: 表格区域
            col_headers: 列表头信息
            
        Returns:
            DataFrame对象
        """
        data = []
        header_row = col_headers.get("start_row", 1)
        
        for row in range(header_row + 1, region.get("max_row", 1) + 1):
            row_data = []
            for col in range(region.get("min_col", 1), region.get("max_col", 1) + 1):
                cell_value = worksheet.cell(row, col).value
                row_data.append(cell_value)
            data.append(row_data)
        
        df = pd.DataFrame(data, columns=col_headers.get("headers", []))
        return df


class FieldExtractor(BaseTool):
    """字段提取器"""
    
    def __init__(self):
        super().__init__(
            name="FieldExtractor",
            description="提取表格中的各个字段",
            tool_type="extractor"
        )
    
    def execute(self, df: pd.DataFrame) -> Dict[str, Any]:
        """
        提取字段信息
        
        Args:
            df: DataFrame对象
            
        Returns:
            字段信息字典
        """
        fields = {
            "field_names": df.columns.tolist(),
            "field_types": {col: str(df[col].dtype) for col in df.columns},
            "field_count": len(df.columns),
            "record_count": len(df)
        }
        
        return fields


class SchemaGenerator(BaseTool):
    """Schema生成器"""
    
    def __init__(self):
        super().__init__(
            name="SchemaGenerator",
            description="生成表格的Schema定义",
            tool_type="generator"
        )
    
    def execute(self, df: pd.DataFrame, meta_info: Dict[str, Any]) -> Dict[str, Any]:
        """
        生成Schema
        
        Args:
            df: DataFrame对象
            meta_info: Meta信息
            
        Returns:
            Schema定义
        """
        schema = {
            "table_name": "excel_table",
            "fields": [],
            "meta": meta_info
        }
        
        for col in df.columns:
            field_schema = {
                "name": col,
                "type": str(df[col].dtype),
                "nullable": df[col].isnull().any(),
                "unique_count": df[col].nunique()
            }
            schema["fields"].append(field_schema)
        
        return schema


# ============ 6. 图谱构建工具 ============

class GraphBuilder(BaseTool):
    """图谱构建器（支持层级关系）"""
    
    def __init__(self):
        super().__init__(
            name="GraphBuilder",
            description="构建带层级关系的知识图谱",
            tool_type="builder"
        )
    
    def execute(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                hierarchy: Dict[str, Any], 
                meta_info: Dict[str, Any],
                region: Optional[Dict[str, Any]] = None,
                excel_filename: Optional[str] = None) -> nx.MultiDiGraph:
        """
        构建图谱，支持层级表头关系
        
        核心逻辑：
        1. 构建Meta节点
        2. 构建列表头层级树（多级列表头）
        3. 构建行表头层级树（多级行表头）
        4. 构建数据单元格节点
        5. 将每个单元格连接到对应的行父节点和列父节点
        
        Args:
            worksheet: 工作表对象
            hierarchy: 层级结构信息
            meta_info: Meta信息
            region: 区域信息（可选）
            excel_filename: Excel文件名（不带后缀），用于设置META节点名称（可选）
            
        Returns:
            NetworkX图谱对象
        """
        graph = nx.MultiDiGraph()
        
        # 1. 添加Meta节点，使用文件名（如果提供）作为节点名称
        meta_node_id = excel_filename if excel_filename else "meta"
        graph.add_node(meta_node_id, 
                      label="meta",
                      properties=meta_info,
                      level=0)
        
        row_header_end_col = hierarchy.get("row_header_end_col", 1)
        col_header_end_row = hierarchy.get("col_header_end_row", 1)
        col_header_levels = hierarchy.get("col_header_levels", [])
        row_header_levels = hierarchy.get("row_header_levels", [])
        
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        min_row = 1
        min_col = 1
        if region:
            # 允许对区域进行轻度扩展，避免边界误差导致单元格遗漏
            min_row = max(1, int(region.get("min_row", 1)) - 1)
            min_col = max(1, int(region.get("min_col", 1)) - 1)
            max_row = min(max_row, int(region.get("max_row", max_row)) + 1)
            max_col = min(max_col, int(region.get("max_col", max_col)) + 1)
        
        # 2. 构建列表头层级树
        col_header_map = {}  # (row, col) -> node_id
        col_leaf_map = {}    # col_idx -> leaf_node_id (叶子节点，用于连接数据)
        
        self._build_col_header_tree(graph, col_header_levels, row_header_end_col, 
                                    max_col, col_header_map, col_leaf_map, meta_node_id)
        
        # 3. 构建行表头层级树
        row_header_map = {}  # (row, col) -> node_id
        row_leaf_map = {}    # row_idx -> leaf_node_id (叶子节点，用于连接数据)
        
        self._build_row_header_tree(graph, row_header_levels, col_header_end_row,
                                    max_row, row_header_map, row_leaf_map, meta_node_id)
        
        # 3.1 添加META root节点，连接所有行表头和列表头，使图连通
        try:
            # 创建META root节点
            meta_root_id = "META_ROOT"
            graph.add_node(meta_root_id,
                          label="META_ROOT",
                          properties={
                              "description": "Root node connecting all headers",
                              "type": "connector"
                          },
                          level=0)
            
            # 连接META root到原始meta节点
            graph.add_edge(meta_root_id, meta_node_id, relation="contains")
            
            # 连接所有第一级列表头到META root
            for (row, col), node_id in col_header_map.items():
                if graph.nodes[node_id].get('properties', {}).get('level', 0) == 0:
                    graph.add_edge(meta_root_id, node_id, relation="has_col_header")
            
            # 连接所有第一级行表头到META root
            for (row, col), node_id in row_header_map.items():
                if graph.nodes[node_id].get('properties', {}).get('level', 0) == 0:
                    graph.add_edge(meta_root_id, node_id, relation="has_row_header")
                    
        except Exception:
            pass
        
        # 4. 构建数据单元格节点，并连接到行列父节点
        start_data_row = max(col_header_end_row + 1, min_row)
        start_data_col = max(row_header_end_col + 1, min_col)
        def _collapse_spaces(s: str) -> str:
            return re.sub(r"\s+", " ", s).strip()
        def _format_display(cell, val) -> str:
            try:
                fmt = getattr(cell, 'number_format', '') or ''
            except Exception:
                fmt = ''
            # percentages
            try:
                if isinstance(val, (int, float)) and ('%' in str(fmt)):
                    return f"{float(val) * 100:.2f}%"
            except Exception:
                pass
            # thousands separator
            try:
                if isinstance(val, (int, float)) and (',' in str(fmt)):
                    return f"{val:,.0f}" if float(val).is_integer() else f"{val:,.2f}"
            except Exception:
                pass
            s = str(val).replace('\n', ' ')
            return _collapse_spaces(s)
        def _format_norm(cell, val) -> str:
            # canonical form for matching: numbers as plain decimals; strings stripped of commas/spaces/percent
            try:
                if isinstance(val, (int, float)):
                    f = float(val)
                    s = ('{:.12f}'.format(f)).rstrip('0').rstrip('.')
                    if s == '-0':
                        s = '0'
                    return s
            except Exception:
                pass
            s = str(val).strip()
            # parentheses negative
            try:
                import re as _re
                m = _re.fullmatch(r"\(([^)]+)\)", s)
                if m:
                    s = '-' + m.group(1)
            except Exception:
                pass
            # remove thousand separators and spaces
            s = s.replace(',', '').replace('\u00a0', '').replace(' ', '')
            # unify percent by removing the sign
            s = s.replace('%', '')
            return s
        for row_idx in range(start_data_row, max_row + 1):
            for col_idx in range(start_data_col, max_col + 1):
                cell = worksheet.cell(row_idx, col_idx)
                cell_value = cell.value
                # 跳过空单元格，避免构图中大量None
                # 尽量保留更多有效值：允许数字0、"0"等；仅过滤纯空白
                if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
                    continue
                cell_id = f"cell_{row_idx}_{col_idx}"
                graph.add_node(cell_id,
                              label="cell",
                              properties={
                                  "value": cell_value,
                                  "value_text": _format_display(cell, cell_value),
                                  "value_norm": _format_norm(cell, cell_value),
                                  "row": row_idx,
                                  "col": col_idx,
                                  "data_type": type(cell_value).__name__
                              },
                              level=100)  # 数据节点层级最深
                
                # 连接到列父节点（列表头的叶子节点）
                col_parent = col_leaf_map.get(col_idx)
                if col_parent:
                    graph.add_edge(col_parent, cell_id, relation="has_cell")
                
                # 连接到行父节点（行表头的叶子节点）
                row_parent = row_leaf_map.get(row_idx)
                if row_parent:
                    graph.add_edge(row_parent, cell_id, relation="has_cell")
        
        return graph
    
    def _build_col_header_tree(self, graph, col_header_levels, start_col, max_col, 
                                col_header_map, col_leaf_map, meta_node_id="meta"):
        """【方案一+方案二改进】构建列表头层级树（去重+合并单元格优化）"""
        if not col_header_levels:
            return
        
        # 【方案一】去重字典：记录相同值和层级的表头，只创建一个节点
        value_to_node_map = {}  # (level_idx, value, row) -> node_id
        
        # 从上到下遍历每一级列表头
        for level_idx, level_headers in enumerate(col_header_levels):
            for header_info in level_headers:
                row = header_info["row"]
                col = header_info["col"]
                value = header_info["value"]
                span = header_info.get("span", 1)
                
                # 处理空值和清理空白字符
                is_empty = False
                if value is None or (isinstance(value, str) and not value.strip()):
                    value = f"[EMPTY_HEADER_{row}_{col}]"
                    is_empty = True
                elif isinstance(value, str):
                    value = value.strip()
                
                # 【关键修复】如果当前位置已经被合并单元格映射占用，跳过创建空表头节点
                # 避免空表头覆盖合并单元格的映射
                if is_empty and (row, col) in col_header_map:
                    # 这个位置已经被合并单元格占用，跳过
                    continue
                
                # 【方案二】检查是否是合并单元格的一部分
                # 如果span > 1，说明这是合并单元格的起始位置
                # 对于同一行、同一值的连续列，只创建一个节点
                dedup_key = (level_idx, value, row)
                
                if span > 1:
                    # 【方案二改进】合并单元格：只为起始位置创建一个节点
                    if dedup_key not in value_to_node_map:
                        node_id = f"col_header_L{level_idx}_R{row}_C{col}_span{span}"
                        
                        graph.add_node(node_id,
                                      label="column_header",
                                      properties={
                                          "value": value,
                                          "row": row,
                                          "col": col,
                                          "col_end": col + span - 1,
                                          "level": level_idx,
                                          "span": span,
                                          "is_merged": True,
                                          "is_placeholder": is_empty
                                      },
                                      level=level_idx + 1)
                        
                        # 记录这个节点，为跨越的所有列建立映射
                        for c in range(col, col + span):
                            col_header_map[(row, c)] = node_id
                            # 如果是最后一级，所有跨越的列都映射到这个叶子节点
                            if level_idx == len(col_header_levels) - 1:
                                col_leaf_map[c] = node_id
                        
                        value_to_node_map[dedup_key] = node_id
                        
                        # 连接到Meta节点（第一级）
                        if level_idx == 0:
                            graph.add_edge(meta_node_id, node_id, relation="has_col_header")
                        
                        # 【修复】连接到父节点（上一级）- 合并单元格作为多个子节点的父节点
                        if level_idx > 0:
                            parent_row = row - 1
                            parent_node = col_header_map.get((parent_row, col))
                            if parent_node:
                                graph.add_edge(parent_node, node_id, relation="has_child")
                        
                        # 【关键修复】如果这是合并单元格的父节点，需要在下一轮连接所有子节点
                        # 这会在处理下一级表头时，通过col_header_map自动建立连接
                else:
                    # 非合并单元格，正常处理
                    if (row, col) not in col_header_map:
                        node_id = f"col_header_L{level_idx}_R{row}_C{col}"
                        
                        graph.add_node(node_id,
                                      label="column_header",
                                      properties={
                                          "value": value,
                                          "row": row,
                                          "col": col,
                                          "level": level_idx,
                                          "span": span,
                                          "is_merged": False,
                                          "is_placeholder": is_empty
                                      },
                                      level=level_idx + 1)
                        
                        col_header_map[(row, col)] = node_id
                        
                        # 连接到Meta节点（第一级）
                        if level_idx == 0:
                            graph.add_edge(meta_node_id, node_id, relation="has_col_header")
                        
                        # 【关键修复】连接到父节点（上一级）- 支持合并单元格父节点连接多个子节点
                        if level_idx > 0:
                            parent_row = row - 1
                            # 直接查找当前列的父节点（合并单元格会为所有跨越列建立映射）
                            parent_node = col_header_map.get((parent_row, col))
                            
                            if not parent_node:
                                # 如果没找到，向前搜索（处理对齐问题）
                                for prev_col in range(max(1, col - 10), col):
                                    parent_node = col_header_map.get((parent_row, prev_col))
                                    if parent_node:
                                        # 验证这个父节点的跨度是否覆盖当前列
                                        parent_props = graph.nodes[parent_node].get('properties', {})
                                        col_end = parent_props.get('col_end', parent_props.get('col'))
                                        if col_end >= col:
                                            # 父节点覆盖当前列，可以连接
                                            break
                                        else:
                                            parent_node = None
                            
                            if parent_node:
                                graph.add_edge(parent_node, node_id, relation="has_child")
                        
                        # 如果是最后一级，记录为叶子节点
                        if level_idx == len(col_header_levels) - 1:
                            col_leaf_map[col] = node_id
    
    def _build_row_header_tree(self, graph, row_header_levels, start_row, max_row,
                                row_header_map, row_leaf_map, meta_node_id="meta"):
        """【方案一+方案二改进】构建行表头层级树（去重+合并单元格优化）"""
        if not row_header_levels:
            return
        
        # 【方案一】去重字典
        value_to_node_map = {}  # (level_idx, value, col) -> node_id
        
        # 从左到右遍历每一级行表头
        for level_idx, level_headers in enumerate(row_header_levels):
            for header_info in level_headers:
                row = header_info["row"]
                col = header_info["col"]
                value = header_info["value"]
                span = header_info.get("span", 1)
                
                # 处理空值和清理空白字符
                is_empty = False
                if value is None or (isinstance(value, str) and not value.strip()):
                    value = f"[EMPTY_HEADER_{row}_{col}]"
                    is_empty = True
                elif isinstance(value, str):
                    value = value.strip()
                
                # 【方案二】合并单元格处理
                dedup_key = (level_idx, value, col)
                
                if span > 1:
                    # 【方案二改进】合并单元格：只为起始位置创建一个节点
                    if dedup_key not in value_to_node_map:
                        node_id = f"row_header_L{level_idx}_R{row}_C{col}_span{span}"
                        
                        graph.add_node(node_id,
                                      label="row_header",
                                      properties={
                                          "value": value,
                                          "row": row,
                                          "row_end": row + span - 1,
                                          "col": col,
                                          "level": level_idx,
                                          "span": span,
                                          "is_merged": True,
                                          "is_placeholder": is_empty
                                      },
                                      level=level_idx + 1)
                        
                        # 记录这个节点，为跨越的所有行建立映射
                        for r in range(row, row + span):
                            row_header_map[(r, col)] = node_id
                            # 如果是最后一级，所有跨越的行都映射到这个叶子节点
                            if level_idx == len(row_header_levels) - 1:
                                row_leaf_map[r] = node_id
                        
                        value_to_node_map[dedup_key] = node_id
                        
                        # 连接到Meta节点（第一级）
                        if level_idx == 0:
                            graph.add_edge(meta_node_id, node_id, relation="has_row_header")
                        
                        # 连接到父节点（左一列）
                        if level_idx > 0:
                            parent_col = col - 1
                            parent_node = row_header_map.get((row, parent_col))
                            if parent_node:
                                graph.add_edge(parent_node, node_id, relation="has_child")
                else:
                    # 非合并单元格，正常处理
                    if (row, col) not in row_header_map:
                        node_id = f"row_header_L{level_idx}_R{row}_C{col}"
                        
                        graph.add_node(node_id,
                                      label="row_header",
                                      properties={
                                          "value": value,
                                          "row": row,
                                          "col": col,
                                          "level": level_idx,
                                          "span": span,
                                          "is_merged": False,
                                          "is_placeholder": is_empty
                                      },
                                      level=level_idx + 1)
                        
                        row_header_map[(row, col)] = node_id
                        
                        # 连接到Meta节点（第一级）
                        if level_idx == 0:
                            graph.add_edge(meta_node_id, node_id, relation="has_row_header")
                        
                        # 连接到父节点（左一列）
                        if level_idx > 0:
                            parent_col = col - 1
                            parent_node = row_header_map.get((row, parent_col))
                            
                            if not parent_node:
                                # 查找合并单元格的父节点
                                for prev_row in range(max(start_row + 1, row - 10), row + 1):
                                    parent_node = row_header_map.get((prev_row, parent_col))
                                    if parent_node:
                                        break
                            
                            if parent_node:
                                graph.add_edge(parent_node, node_id, relation="has_child")
                    
                    # 如果是最后一级，记录为叶子节点
                    if level_idx == len(row_header_levels) - 1:
                        row_leaf_map[row] = node_id


class NodeCreator(BaseTool):
    """节点创建器"""
    
    def __init__(self):
        super().__init__(
            name="NodeCreator",
            description="创建图谱节点",
            tool_type="builder"
        )
    
    def execute(self, node_id: str, label: str, properties: Dict[str, Any], level: int = 2) -> Dict[str, Any]:
        """
        创建节点
        
        Args:
            node_id: 节点ID
            label: 节点标签
            properties: 节点属性
            level: 节点层级
            
        Returns:
            节点定义
        """
        return {
            "id": node_id,
            "label": label,
            "properties": properties,
            "level": level
        }


class EdgeCreator(BaseTool):
    """边创建器"""
    
    def __init__(self):
        super().__init__(
            name="EdgeCreator",
            description="创建图谱边（关系）",
            tool_type="builder"
        )
    
    def execute(self, from_id: str, to_id: str, relation: str) -> Dict[str, Any]:
        """
        创建边
        
        Args:
            from_id: 起始节点ID
            to_id: 目标节点ID
            relation: 关系类型
            
        Returns:
            边定义
        """
        return {
            "from": from_id,
            "to": to_id,
            "relation": relation
        }


class MetaNodeBuilder(BaseTool):
    """Meta节点构建器"""
    
    def __init__(self):
        super().__init__(
            name="MetaNodeBuilder",
            description="构建Meta信息节点",
            tool_type="builder"
        )
    
    def execute(self, meta_info: Dict[str, Any]) -> Dict[str, Any]:
        """
        构建Meta节点
        
        Args:
            meta_info: Meta信息
            
        Returns:
            Meta节点定义
        """
        return {
            "id": "meta",
            "label": "meta",
            "properties": meta_info,
            "level": 0
        }


class GraphSerializer(BaseTool):
    """图谱序列化器"""
    
    def __init__(self):
        super().__init__(
            name="GraphSerializer",
            description="将图谱序列化为JSON格式",
            tool_type="serializer"
        )
    
    def execute(self, graph: nx.MultiDiGraph, output_path: str) -> str:
        """
        序列化图谱
        
        Args:
            graph: NetworkX图谱对象
            output_path: 输出路径
            
        Returns:
            输出文件路径
        """
        import sys
        from pathlib import Path
        sys.path.append(str(Path(__file__).parent.parent))
        from utils.graph_processor import save_graph_to_json
        save_graph_to_json(graph, output_path)
        return output_path


class GraphSaver(BaseTool):
    """图谱保存器（使用graph_processor）"""
    
    def __init__(self):
        super().__init__(
            name="GraphSaver",
            description="保存图谱到JSON文件",
            tool_type="saver"
        )
    
    def execute(self, graph: nx.MultiDiGraph, output_path: str) -> str:
        """
        保存图谱（使用优化存储格式）
        
        Args:
            graph: NetworkX图谱对象
            output_path: 输出路径
            
        Returns:
            输出文件路径
        """
        import sys
        from pathlib import Path
        sys.path.append(str(Path(__file__).parent.parent))
        from utils.graph_processor import save_optimized_graph_to_json
        save_optimized_graph_to_json(graph, output_path)
        return output_path


# ============================================================
# 检索阶段工具 - Phase 2: Retrieval
# ============================================================

# ============ 7. 图谱加载工具 ============

class GraphLoader(BaseTool):
    """图谱加载器"""
    
    def __init__(self):
        super().__init__(
            name="GraphLoader",
            description="加载已构建的知识图谱",
            tool_type="loader"
        )
    
    def execute(self, graph_path: str) -> nx.MultiDiGraph:
        """
        加载图谱
        
        Args:
            graph_path: 图谱文件路径
            
        Returns:
            NetworkX图谱对象
        """
        import sys
        from pathlib import Path
        sys.path.append(str(Path(__file__).parent.parent))
        from utils.graph_processor import load_graph
        graph = load_graph(graph_path)
        return graph


class GraphReader(BaseTool):
    """图谱读取器（使用graph_processor）"""
    
    def __init__(self):
        super().__init__(
            name="GraphReader",
            description="从JSON文件读取图谱",
            tool_type="reader"
        )
    
    def execute(self, graph_path: str) -> nx.MultiDiGraph:
        """
        读取图谱
        
        Args:
            graph_path: 图谱文件路径
            
        Returns:
            NetworkX图谱对象
        """
        import sys
        from pathlib import Path
        sys.path.append(str(Path(__file__).parent.parent))
        from utils.graph_processor import load_graph_from_json
        graph = load_graph_from_json(graph_path)
        return graph


class NetworkXReader(BaseTool):
    """NetworkX读取器"""
    
    def __init__(self):
        super().__init__(
            name="NetworkXReader",
            description="使用NetworkX读取和分析图谱",
            tool_type="reader"
        )
    
    def execute(self, graph: nx.MultiDiGraph) -> Dict[str, Any]:
        """
        读取图谱信息
        
        Args:
            graph: NetworkX图谱对象
            
        Returns:
            图谱信息字典
        """
        return {
            "node_count": graph.number_of_nodes(),
            "edge_count": graph.number_of_edges(),
            "is_directed": graph.is_directed(),
            "nodes": list(graph.nodes()),
            "edges": list(graph.edges())
        }


# ============ 8. 查询输入工具 ============

class QueryParser(BaseTool):
    """查询解析器"""
    
    def __init__(self):
        super().__init__(
            name="QueryParser",
            description="解析用户查询",
            tool_type="parser"
        )
    
    def execute(self, query: str) -> Dict[str, Any]:
        """
        解析查询
        
        Args:
            query: 用户查询字符串
            
        Returns:
            解析后的查询信息
        """
        return {
            "original_query": query,
            "tokens": query.split(),
            "length": len(query),
            "query_type": "unknown"
        }


class QueryNormalizer(BaseTool):
    """查询规范化器"""
    
    def __init__(self):
        super().__init__(
            name="QueryNormalizer",
            description="规范化查询文本",
            tool_type="normalizer"
        )
    
    def execute(self, parsed_query: Dict[str, Any]) -> str:
        """
        规范化查询
        
        Args:
            parsed_query: 解析后的查询
            
        Returns:
            规范化的查询字符串
        """
        # 简单实现：转小写，去除多余空格
        query = parsed_query.get("original_query", "")
        normalized = " ".join(query.lower().split())
        return normalized


# ============ 9. Meta提取工具 ============

class MetaExtractor(BaseTool):
    """Meta信息提取器（从图谱）"""
    
    def __init__(self):
        super().__init__(
            name="MetaExtractor",
            description="从图谱中提取Meta信息",
            tool_type="extractor"
        )
    
    def execute(self, graph: nx.MultiDiGraph) -> Dict[str, Any]:
        """
        提取Meta信息
        
        Args:
            graph: NetworkX图谱对象
            
        Returns:
            Meta信息
        """
        # 查找meta节点（通过label查找，因为节点ID可能是文件名）
        meta_info = {}
        
        # 查找label为"meta"的节点
        for node_id, node_data in graph.nodes(data=True):
            if node_data.get("label") == "meta":
                meta_info = node_data.get("properties", {})
                break
        
        return meta_info


class HeaderInfoRetriever(BaseTool):
    """表头信息检索器"""
    
    def __init__(self):
        super().__init__(
            name="HeaderInfoRetriever",
            description="检索表头信息",
            tool_type="retriever"
        )
    
    def execute(self, graph: nx.MultiDiGraph) -> Dict[str, Any]:
        """
        检索表头信息
        
        Args:
            graph: NetworkX图谱对象
            
        Returns:
            表头信息
        """
        headers = {
            "column_headers": [],
            "row_headers": []
        }
        
        # 查找所有列表头节点
        for node_id, node_data in graph.nodes(data=True):
            if node_data.get("label") == "column_header":
                headers["column_headers"].append(node_data.get("properties", {}))
        
        return headers


# ============ 10. 子问题分解工具 ============

class LLMDecomposer(BaseTool):
    """LLM子问题分解器"""
    
    def __init__(self):
        super().__init__(
            name="LLMDecomposer",
            description="利用大模型分解复杂问题为子问题",
            tool_type="decomposer"
        )
    
    def execute(self, query: str, meta_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        分解子问题
        
        Args:
            query: 原始查询
            meta_info: Meta信息
            
        Returns:
            子问题列表
        """
        # 简单实现：返回原查询作为单个子问题
        # 实际应用中需要调用LLM API
        subqueries = [
            {
                "id": 0,
                "query": query,
                "type": "retrieval",
                "meta_context": meta_info
            }
        ]
        
        return subqueries


class SubqueryGenerator(BaseTool):
    """子查询生成器"""
    
    def __init__(self):
        super().__init__(
            name="SubqueryGenerator",
            description="生成子查询",
            tool_type="generator"
        )
    
    def execute(self, decomposed_queries: List[Dict[str, Any]]) -> List[str]:
        """
        生成子查询
        
        Args:
            decomposed_queries: 分解后的查询
            
        Returns:
            子查询列表
        """
        return [q.get("query", "") for q in decomposed_queries]


class IntentAnalyzer(BaseTool):
    """意图分析器"""
    
    def __init__(self):
        super().__init__(
            name="IntentAnalyzer",
            description="分析查询意图",
            tool_type="analyzer"
        )
    
    def execute(self, query: str) -> str:
        """
        分析意图
        
        Args:
            query: 查询字符串
            
        Returns:
            意图类型
        """
        query_lower = query.lower()
        
        if any(word in query_lower for word in ["sum", "总和", "合计"]):
            return "aggregation_sum"
        elif any(word in query_lower for word in ["average", "avg", "平均"]):
            return "aggregation_avg"
        elif any(word in query_lower for word in ["count", "数量", "个数"]):
            return "aggregation_count"
        elif any(word in query_lower for word in ["max", "min", "最大", "最小"]):
            return "aggregation_extrema"
        else:
            return "retrieval"


# ============ 11. 向量检索工具 ============

class FAISSRetriever(BaseTool):
    """FAISS向量检索器"""
    
    def __init__(self):
        super().__init__(
            name="FAISSRetriever",
            description="使用FAISS进行向量相似度检索",
            tool_type="retriever"
        )
    
    def execute(self, query_embedding: List[float], 
                value_embeddings: Dict[str, List[float]], 
                top_k: int = 5) -> List[Tuple[str, float]]:
        """
        向量检索
        
        Args:
            query_embedding: 查询向量
            value_embeddings: 值向量字典
            top_k: 返回top-k结果
            
        Returns:
            检索结果列表 [(value_id, score), ...]
        """
        # 简单实现：余弦相似度
        # 实际应用中使用FAISS库
        results = []
        
        return results[:top_k]


class EmbeddingGenerator(BaseTool):
    """Embedding生成器"""
    
    def __init__(self):
        super().__init__(
            name="EmbeddingGenerator",
            description="生成文本的向量表示",
            tool_type="generator"
        )
    
    def execute(self, text: str) -> List[float]:
        """
        生成Embedding
        
        Args:
            text: 输入文本
            
        Returns:
            向量表示
        """
        # 简单实现：返回空列表
        # 实际应用中使用Sentence-BERT等模型
        return []


class SimilarityCalculator(BaseTool):
    """相似度计算器"""
    
    def __init__(self):
        super().__init__(
            name="SimilarityCalculator",
            description="计算向量相似度",
            tool_type="calculator"
        )
    
    def execute(self, vec1: List[float], vec2: List[float]) -> float:
        """
        计算相似度
        
        Args:
            vec1: 向量1
            vec2: 向量2
            
        Returns:
            相似度分数
        """
        # 简单实现：余弦相似度
        if not vec1 or not vec2:
            return 0.0
        
        dot_product = sum(a * b for a, b in zip(vec1, vec2))
        norm1 = sum(a * a for a in vec1) ** 0.5
        norm2 = sum(b * b for b in vec2) ** 0.5
        
        if norm1 == 0 or norm2 == 0:
            return 0.0
        
        return dot_product / (norm1 * norm2)


# ============ 11. Meta匹配工具 ============

class MetaVectorMatcher(BaseTool):
    """Meta向量匹配器"""
    
    def __init__(self):
        super().__init__(
            name="MetaVectorMatcher",
            description="使用向量匹配Meta信息",
            tool_type="matcher"
        )
    
    def execute(self, query: str, meta_info: Dict[str, Any]) -> List[str]:
        """
        匹配Meta信息
        
        Args:
            query: 查询
            meta_info: Meta信息
            
        Returns:
            匹配的表头列表
        """
        matched_headers = []
        
        # 简单实现：字符串匹配
        col_headers = meta_info.get("col_header_names", [])
        query_lower = query.lower()
        
        for header in col_headers:
            if header and str(header).lower() in query_lower:
                matched_headers.append(str(header))
        
        return matched_headers


class HeaderMatcher(BaseTool):
    """表头匹配器"""
    
    def __init__(self):
        super().__init__(
            name="HeaderMatcher",
            description="匹配相关表头",
            tool_type="matcher"
        )
    
    def execute(self, subquery: str, headers: List[str]) -> List[str]:
        """
        匹配表头
        
        Args:
            subquery: 子查询
            headers: 表头列表
            
        Returns:
            匹配的表头
        """
        matched = []
        subquery_lower = subquery.lower()
        
        for header in headers:
            if header and str(header).lower() in subquery_lower:
                matched.append(header)
        
        return matched


class SemanticMatcher(BaseTool):
    """语义匹配器"""
    
    def __init__(self):
        super().__init__(
            name="SemanticMatcher",
            description="基于语义的匹配",
            tool_type="matcher"
        )
    
    def execute(self, query: str, candidates: List[str], threshold: float = 0.7) -> List[str]:
        """
        语义匹配
        
        Args:
            query: 查询
            candidates: 候选项列表
            threshold: 相似度阈值
            
        Returns:
            匹配的候选项
        """
        # 简单实现：返回所有候选项
        # 实际应用中使用语义相似度模型
        return candidates


# ============ 12. 子图检索工具 ============

class SubgraphExtractor(BaseTool):
    """子图提取器"""
    
    def __init__(self):
        super().__init__(
            name="SubgraphExtractor",
            description="从图谱中提取相关子图",
            tool_type="extractor"
        )
    
    def execute(self, graph: nx.MultiDiGraph, 
                seed_nodes: List[str], 
                hops: int = 2) -> nx.MultiDiGraph:
        """
        提取子图
        
        Args:
            graph: 完整图谱
            seed_nodes: 种子节点列表
            hops: 扩展跳数
            
        Returns:
            子图
        """
        subgraph_nodes = set(seed_nodes)
        
        # 扩展N跳
        for _ in range(hops):
            new_nodes = set()
            for node in subgraph_nodes:
                if node in graph:
                    # 添加后继节点
                    new_nodes.update(graph.successors(node))
                    # 添加前驱节点
                    new_nodes.update(graph.predecessors(node))
            subgraph_nodes.update(new_nodes)
        
        subgraph = graph.subgraph(subgraph_nodes).copy()
        return subgraph


class NodeTraverser(BaseTool):
    """节点遍历器"""
    
    def __init__(self):
        super().__init__(
            name="NodeTraverser",
            description="遍历图谱节点",
            tool_type="traverser"
        )
    
    def execute(self, graph: nx.MultiDiGraph, start_node: str, 
                direction: str = "successors") -> List[str]:
        """
        遍历节点
        
        Args:
            graph: 图谱对象
            start_node: 起始节点
            direction: 遍历方向 (successors/predecessors/both)
            
        Returns:
            遍历到的节点列表
        """
        nodes = []
        
        if start_node not in graph:
            return nodes
        
        if direction == "successors":
            nodes = list(graph.successors(start_node))
        elif direction == "predecessors":
            nodes = list(graph.predecessors(start_node))
        else:  # both
            nodes = list(graph.successors(start_node)) + list(graph.predecessors(start_node))
        
        return nodes


class PathFinder(BaseTool):
    """路径查找器"""
    
    def __init__(self):
        super().__init__(
            name="PathFinder",
            description="查找节点间的路径",
            tool_type="finder"
        )
    
    def execute(self, graph: nx.MultiDiGraph, 
                source: str, target: str) -> List[List[str]]:
        """
        查找路径
        
        Args:
            graph: 图谱对象
            source: 源节点
            target: 目标节点
            
        Returns:
            路径列表
        """
        try:
            paths = list(nx.all_simple_paths(graph, source, target, cutoff=5))
            return paths
        except:
            return []


class RelationshipFollower(BaseTool):
    """关系跟随器"""
    
    def __init__(self):
        super().__init__(
            name="RelationshipFollower",
            description="沿着特定关系类型遍历",
            tool_type="traverser"
        )
    
    def execute(self, graph: nx.MultiDiGraph, 
                start_node: str, 
                relation_type: str) -> List[str]:
        """
        跟随关系
        
        Args:
            graph: 图谱对象
            start_node: 起始节点
            relation_type: 关系类型
            
        Returns:
            相关节点列表
        """
        related_nodes = []
        
        if start_node not in graph:
            return related_nodes
        
        for successor in graph.successors(start_node):
            edge_data = graph.get_edge_data(start_node, successor)
            if edge_data:
                for key, data in edge_data.items():
                    if data.get("relation") == relation_type:
                        related_nodes.append(successor)
        
        return related_nodes


# ============ 13. 聚合计算工具 ============

class SumAggregator(BaseTool):
    """求和聚合器"""
    
    def __init__(self):
        super().__init__(
            name="SumAggregator",
            description="计算数值总和",
            tool_type="aggregator"
        )
    
    def execute(self, values: List[float]) -> float:
        """
        求和
        
        Args:
            values: 数值列表
            
        Returns:
            总和
        """
        return sum(values)


class AverageCalculator(BaseTool):
    """平均值计算器"""
    
    def __init__(self):
        super().__init__(
            name="AverageCalculator",
            description="计算平均值",
            tool_type="calculator"
        )
    
    def execute(self, values: List[float]) -> float:
        """
        计算平均值
        
        Args:
            values: 数值列表
            
        Returns:
            平均值
        """
        if not values:
            return 0.0
        return sum(values) / len(values)


class CountTool(BaseTool):
    """计数工具"""
    
    def __init__(self):
        super().__init__(
            name="CountTool",
            description="计算数量",
            tool_type="counter"
        )
    
    def execute(self, items: List[Any]) -> int:
        """
        计数
        
        Args:
            items: 项目列表
            
        Returns:
            数量
        """
        return len(items)


class MaxMinFinder(BaseTool):
    """最大最小值查找器"""
    
    def __init__(self):
        super().__init__(
            name="MaxMinFinder",
            description="查找最大值和最小值",
            tool_type="finder"
        )
    
    def execute(self, values: List[float]) -> Dict[str, float]:
        """
        查找最大最小值
        
        Args:
            values: 数值列表
            
        Returns:
            包含max和min的字典
        """
        if not values:
            return {"max": 0.0, "min": 0.0}
        
        return {
            "max": max(values),
            "min": min(values)
        }


class GroupByAggregator(BaseTool):
    """分组聚合器"""
    
    def __init__(self):
        super().__init__(
            name="GroupByAggregator",
            description="按字段分组聚合",
            tool_type="aggregator"
        )
    
    def execute(self, data: List[Dict[str, Any]], 
                group_by: str, 
                agg_field: str, 
                agg_func: str = "sum") -> Dict[str, float]:
        """
        分组聚合
        
        Args:
            data: 数据列表
            group_by: 分组字段
            agg_field: 聚合字段
            agg_func: 聚合函数 (sum/avg/count)
            
        Returns:
            分组聚合结果
        """
        groups = {}
        
        for item in data:
            key = item.get(group_by)
            value = item.get(agg_field, 0)
            
            if key not in groups:
                groups[key] = []
            groups[key].append(value)
        
        results = {}
        for key, values in groups.items():
            if agg_func == "sum":
                results[key] = sum(values)
            elif agg_func == "avg":
                results[key] = sum(values) / len(values) if values else 0
            elif agg_func == "count":
                results[key] = len(values)
        
        return results


# ============ 14. 答案合成工具 ============

class AnswerGenerator(BaseTool):
    """答案生成器"""
    
    def __init__(self):
        super().__init__(
            name="AnswerGenerator",
            description="生成最终答案",
            tool_type="generator"
        )
    
    def execute(self, aggregation_results: Any, 
                query: str, 
                intent: str) -> str:
        """
        生成答案
        
        Args:
            aggregation_results: 聚合结果
            query: 原始查询
            intent: 查询意图
            
        Returns:
            答案字符串
        """
        if isinstance(aggregation_results, dict):
            answer = f"查询结果: {aggregation_results}"
        else:
            answer = f"答案: {aggregation_results}"
        
        return answer


class ResultFormatter(BaseTool):
    """结果格式化器"""
    
    def __init__(self):
        super().__init__(
            name="ResultFormatter",
            description="格式化输出结果",
            tool_type="formatter"
        )
    
    def execute(self, answer: str, metadata: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        格式化结果
        
        Args:
            answer: 答案字符串
            metadata: 元数据
            
        Returns:
            格式化的结果
        """
        return {
            "answer": answer,
            "metadata": metadata or {},
            "status": "success"
        }


class ConfidenceEstimator(BaseTool):
    """置信度估计器"""
    
    def __init__(self):
        super().__init__(
            name="ConfidenceEstimator",
            description="估计答案的置信度",
            tool_type="estimator"
        )
    
    def execute(self, answer: str, 
                retrieval_results: List[Any], 
                aggregation_results: Any) -> float:
        """
        估计置信度
        
        Args:
            answer: 答案
            retrieval_results: 检索结果
            aggregation_results: 聚合结果
            
        Returns:
            置信度分数 (0-1)
        """
        # 简单实现：基于检索结果数量
        if not retrieval_results:
            return 0.3
        elif len(retrieval_results) >= 5:
            return 0.9
        else:
            return 0.6


# ============================================================
# 工具注册表
# ============================================================

class ToolRegistry:
    """工具注册表"""
    
    def __init__(self):
        self.tools: Dict[str, BaseTool] = {}
        self._register_default_tools()
    
    def _register_default_tools(self):
        """注册所有默认工具"""
        default_tools = [
            # 构建阶段工具
            # 1. Excel加载
            ExcelLoader(),
            WorkbookReader(),
            # 2. 区域识别
            NestedTableDetector(),
            RegionSegmenter(),
            TableBoundaryFinder(),
            # 3. 数据清洗
            MergedCellSplitter(),
            NullValueFiller(),
            DataTypeNormalizer(),
            WhitespaceRemover(),
            # 4. 表头识别（原有）
            RowHeaderDetector(),
            ColumnHeaderDetector(),
            HierarchicalHeaderParser(),
            MetaInfoExtractor(),
            # 4.1 综合表格提取（新增）
            ComprehensiveTableExtractor(),
            RuleBasedHeaderDetector(),
            DataFrameBasedHeaderDetector(),
            MergedCellAnalyzer(),
            EnhancedNestedTableDetector(),
            HeaderValidationEngine(),
            # 5. DataFrame提取
            DataFrameBuilder(),
            FieldExtractor(),
            SchemaGenerator(),
            # 6. 图谱构建
            GraphBuilder(),
            NodeCreator(),
            EdgeCreator(),
            MetaNodeBuilder(),
            GraphSerializer(),
            GraphSaver(),
            
            # 检索阶段工具
            # 7. 图谱加载
            GraphLoader(),
            GraphReader(),
            NetworkXReader(),
            # 8. 查询输入
            QueryParser(),
            QueryNormalizer(),
            # 9. Meta提取
            MetaExtractor(),
            HeaderInfoRetriever(),
            # 10. 子问题分解
            LLMDecomposer(),
            SubqueryGenerator(),
            IntentAnalyzer(),
            # 11. 向量检索
            FAISSRetriever(),
            EmbeddingGenerator(),
            SimilarityCalculator(),
            # 11. Meta匹配
            MetaVectorMatcher(),
            HeaderMatcher(),
            SemanticMatcher(),
            # 12. 子图检索
            SubgraphExtractor(),
            NodeTraverser(),
            PathFinder(),
            RelationshipFollower(),
            # 13. 聚合计算
            SumAggregator(),
            AverageCalculator(),
            CountTool(),
            MaxMinFinder(),
            GroupByAggregator(),
            # 14. 答案合成
            AnswerGenerator(),
            ResultFormatter(),
            ConfidenceEstimator()
        ]
        
        for tool in default_tools:
            self.register(tool)
    
    def register(self, tool: BaseTool):
        """注册工具"""
        self.tools[tool.name] = tool
    
    def get(self, name: str) -> Optional[BaseTool]:
        """获取工具"""
        return self.tools.get(name)
    
    def get_all(self) -> Dict[str, BaseTool]:
        """获取所有工具"""
        return self.tools
    
    def list_tools(self) -> List[str]:
        """列出所有工具名称"""
        return list(self.tools.keys())
    
    def get_tools_by_type(self, tool_type: str) -> List[BaseTool]:
        """按类型获取工具"""
        return [tool for tool in self.tools.values() if tool.tool_type == tool_type]


# 创建全局工具注册表实例
tool_registry = ToolRegistry()
