"""Knowledge base configuration service.

Handles business logic for configuration updates, building, and validation.

⚠️ SQL Injection Prevention:
- All database operations use SQLAlchemy ORM with parameter binding
- User inputs are properly validated before database queries

⚠️ Security Note:
- Database passwords are stored in plaintext in the config field
- In production, consider using encrypted storage or secrets management
"""

import io
import json
import logging
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional, List, Dict

import openpyxl
import yaml
from sqlalchemy.orm import Session

from ..database import (
    KnowledgeBase,
    KBSourceConfig,
    KBBuildConfig,
    KBBuildLog,
)
from ..minio_client import minio_client
from ..utils import load_yaml_config
from ...knowledge_builder.agent import KnowledgeBuilderAgent, BuildRequest, SourceConfig
from ...knowledge_builder.cleanup_manager import KnowledgeCleanupManager
from ...storage.implementations.chroma_store import ChromaVectorStore
from ...config import VectorStoreConfig

logger = logging.getLogger(__name__)


class KBConfigService:
    """Service interface for knowledge base configuration."""
    
    # Import from utils to maintain backward-compatible interface
    load_yaml_config = staticmethod(load_yaml_config)
    
    @staticmethod
    def parse_tables_created(tables_created_raw: Any) -> Optional[List[str]]:
        """Parse the tables_created field (could be JSON string, comma-separated string, or list).
        
        Args:
            tables_created_raw: Raw tables_created data.
            
        Returns:
            List of table names, or None if empty.
        """
        if not tables_created_raw:
            return None
            
        if isinstance(tables_created_raw, str):
            try:
                # Try parsing as JSON
                return json.loads(tables_created_raw)
            except json.JSONDecodeError:
                # Try comma-separated
                return [t.strip() for t in tables_created_raw.split(',') if t.strip()]
        elif isinstance(tables_created_raw, list):
            return tables_created_raw
            
        return None
    
    @staticmethod
    async def cleanup_removed_sources(
        removed_sources: List[KBSourceConfig],
        collection_name: str,
        vector_store_path: str = None,
        relational_db_path: str = None
    ) -> Dict[str, int]:
        """Cleanup removed sources (vector data, SQLite tables, column-level embeddings).
        
        Args:
            removed_sources: List of removed source configurations.
            collection_name: Vector store collection name.
            vector_store_path: Vector store path (optional).
            relational_db_path: Relational database path (optional).
            
        Returns:
            A dictionary of statistics of the cleanup process.
        """
        if not removed_sources:
            return {"total_deleted_chunks": 0, "total_deleted_tables": 0}
        
        logger.info(f"🗑️  Detected {len(removed_sources)} removed sources, cleaning up...")
        
        vector_config = VectorStoreConfig(
            collection_name=collection_name,
            persist_directory=vector_store_path or os.getenv("VECTOR_STORE_PATH", "rag_data/vector_store")
        )
        vector_store = ChromaVectorStore(config=vector_config)
        
        cleanup_manager = KnowledgeCleanupManager(
            vector_store=vector_store,
            relational_db_path=relational_db_path or os.getenv("RELATIONAL_DB_PATH", "rag_data/relational_database/rag_demo.sqlite")
        )
        
        total_deleted_chunks = 0
        total_deleted_tables = 0
        
        # Cleanup each removed source
        for removed_source in removed_sources:
            try:
                # Parse tables_created
                tables_created = KBConfigService.parse_tables_created(removed_source.tables_created)
                
                # Full cleanup
                cleanup_stats = await cleanup_manager.cleanup_source(
                    source_identifier=removed_source.source_identifier,
                    source_type=removed_source.source_type,
                    tables_created=tables_created,
                    kb_id=removed_source.knowledge_base_id
                )
                
                total_deleted_chunks += cleanup_stats["vector_chunks_deleted"]
                total_deleted_tables += cleanup_stats["sqlite_tables_deleted"]
                
                logger.info(
                    f"🗑️  Cleaned up {removed_source.source_identifier}: "
                    f"{cleanup_stats['vector_chunks_deleted']} vectors, "
                    f"{cleanup_stats['sqlite_tables_deleted']} tables"
                )
                
                if cleanup_stats["errors"]:
                    for error in cleanup_stats["errors"]:
                        logger.warning(f"  ⚠️  {error}")
                
            except Exception as e:
                logger.error(f"Failed to cleanup {removed_source.source_identifier}: {e}")
        
        logger.info(
            f"✅ Cleanup completed: "
            f"{total_deleted_chunks} vectors and {total_deleted_tables} SQLite tables removed "
            f"from {len(removed_sources)} sources"
        )
        
        return {
            "total_deleted_chunks": total_deleted_chunks,
            "total_deleted_tables": total_deleted_tables
        }
    
    @staticmethod
    def _determine_source_status(
        old_config: Optional[Dict[str, Any]],
        current_etag: Optional[str] = None
    ) -> Dict[str, Any]:
        """Determine source status from old configuration and current ETag (if provided).
        
        Args:
            old_config: Old source configuration.
            current_etag: Current file ETag (optional).
            
        Returns:
            A dictionary containing status, chunks_created, and error_message.
        """
        if old_config and old_config["status"] == "completed":
            # If ETag is present, compare ETag to decide if rebuild is needed.
            # Maintain completed status if unchanged.
            # Rebuild if changed.
            if current_etag and old_config.get("source_etag"):
                if current_etag == old_config["source_etag"]:
                    return {
                        "status": "completed",
                        "chunks_created": old_config.get("chunks_created", 0),
                        "error_message": None
                    }
                else:
                    return {
                        "status": "pending",
                        "chunks_created": 0,
                        "error_message": None
                    }
            # No ETag information, maintain original status
            return {
                "status": "completed",
                "chunks_created": old_config.get("chunks_created", 0),
                "error_message": None
            }
        elif old_config:
            # Maintain original non-completed status
            return {
                "status": old_config["status"],
                "chunks_created": old_config.get("chunks_created", 0),
                "error_message": old_config.get("error_message")
            }
        else:
            # New file, set pending status
            return {
                "status": "pending",
                "chunks_created": 0,
                "error_message": None
            }

    @staticmethod
    async def update_configuration(
        kb_id: int,
        tools_config: Dict[str, Any],
        selected_files: List[str],
        selected_qa_files: List[str],
        db_connections: List[Dict[str, Any]],
        db: Session
    ) -> Dict[str, Any]:
        """Update knowledge base configuration.
        
        Args:
            kb_id: Knowledge base ID.
            tools_config: Tools configuration.
            selected_files: List of selected files.
            selected_qa_files: List of selected QA files.
            db_connections: List of database connection configurations.
            db: Database session.
            
        Returns:
            Configuration update results.
        """
        kb = db.query(KnowledgeBase).filter(KnowledgeBase.id == kb_id).first()
        if not kb:
            raise ValueError("Knowledge base not found")

        # 1. Save old source configurations
        old_sources = db.query(KBSourceConfig).filter(
            KBSourceConfig.knowledge_base_id == kb_id
        ).all()
        
        old_sources_map = {
            src.source_identifier: {
                "status": src.status,
                "source_etag": src.source_etag,
                "chunks_created": src.chunks_created,
                "tables_created": src.tables_created,
                "error_message": src.error_message
            }
            for src in old_sources
        }
        
        # 2. Identify removed files and clean up their vector data
        new_source_identifiers = set()
        new_source_identifiers.update(selected_files)
        new_source_identifiers.update(selected_qa_files)
        for db_conn in db_connections:
            connection_str = db_conn.get("connectionString", "")
            for table_name in db_conn.get("tables", []):
                new_source_identifiers.add(f"{connection_str}:{table_name}")
        
        removed_sources = [src for src in old_sources if src.source_identifier not in new_source_identifiers]
        
        # 2.1. Clean up removed sources
        cleanup_stats = await KBConfigService.cleanup_removed_sources(
            removed_sources=removed_sources,
            collection_name=kb.collection_name
        )
        total_deleted_chunks = cleanup_stats["total_deleted_chunks"]
        total_deleted_tables = cleanup_stats["total_deleted_tables"]
        
        # 2.2. Delete configuration records of removed sources
        if removed_sources:
            for removed_source in removed_sources:
                db.delete(removed_source)
        
        # 3. UPSERT MinIO file sources
        for filename in selected_files:
            file_ext = filename.split('.')[-1].lower() if '.' in filename else 'txt'
            
            etag = None
            try:
                stat = minio_client.get_file_stat(filename)
                if stat:
                    etag = stat.etag.strip('"') if stat.etag else None
            except Exception as e:
                logger.warning(f"Failed to get ETag for {filename}: {e}")
            
            existing_source = db.query(KBSourceConfig).filter(
                KBSourceConfig.knowledge_base_id == kb_id,
                KBSourceConfig.source_type == "minio_file",
                KBSourceConfig.source_identifier == filename
            ).first()
            
            old_config = old_sources_map.get(filename)
            source_status = KBConfigService._determine_source_status(old_config, etag)
            
            if existing_source:  # Update if exists
                existing_source.source_etag = etag
                existing_source.config = {
                    "file_path": filename,
                    "file_type": file_ext,
                    "kb_id": kb_id,
                    "collection_name": kb.collection_name
                }
                existing_source.status = source_status["status"]
                existing_source.chunks_created = source_status["chunks_created"]
                existing_source.error_message = source_status["error_message"]
                existing_source.updated_at = datetime.now(timezone.utc)
            else:  # Insert new source
                source = KBSourceConfig(
                    knowledge_base_id=kb_id,
                    source_type="minio_file",
                    source_identifier=filename,
                    source_etag=etag,
                    config={
                        "file_path": filename,
                        "file_type": file_ext,
                        "kb_id": kb_id,
                        "collection_name": kb.collection_name
                    },
                    status=source_status["status"],
                    chunks_created=source_status["chunks_created"],
                    error_message=source_status["error_message"]
                )
                db.add(source)
        
        # 4. UPSERT QA file sources
        for qa_filename in selected_qa_files:
            etag = None
            try:
                stat = minio_client.get_file_stat(qa_filename)
                if stat:
                    etag = stat.etag.strip('"') if stat.etag else None
            except Exception as e:
                logger.warning(f"Failed to get ETag for QA file {qa_filename}: {e}")
            
            existing_source = db.query(KBSourceConfig).filter(
                KBSourceConfig.knowledge_base_id == kb_id,
                KBSourceConfig.source_type == "qa_file",
                KBSourceConfig.source_identifier == qa_filename
            ).first()
            
            old_config = old_sources_map.get(qa_filename)
            source_status = KBConfigService._determine_source_status(old_config, etag)
            
            if existing_source:  # Update if exists
                existing_source.source_etag = etag
                existing_source.config = {
                    "file_path": qa_filename,
                    "sheet_name": "example",
                    "kb_id": kb_id,
                    "collection_name": kb.collection_name
                }
                existing_source.status = source_status["status"]
                existing_source.chunks_created = source_status["chunks_created"]
                existing_source.error_message = source_status["error_message"]
                existing_source.updated_at = datetime.now(timezone.utc)
            else:  # Insert new source
                source = KBSourceConfig(
                    knowledge_base_id=kb_id,
                    source_type="qa_file",
                    source_identifier=qa_filename,
                    source_etag=etag,
                    config={
                        "file_path": qa_filename,
                        "sheet_name": "example",
                        "kb_id": kb_id,
                        "collection_name": kb.collection_name
                    },
                    status=source_status["status"],
                    chunks_created=source_status["chunks_created"],
                    error_message=source_status["error_message"]
                )
                db.add(source)
        
        # 5. UPSERT database sources
        # ⚠️ SECURITY WARNING: Database passwords are stored in plaintext
        for db_conn in db_connections:
            connection_str = db_conn.get("connectionString", "")
            db_type = db_conn.get("type", "mysql")

            for table_name in db_conn.get("tables", []):
                config_data = {
                    "connection_string": connection_str,
                    "table_name": table_name,
                    "db_type": db_type,
                    "kb_id": kb_id,
                    "collection_name": kb.collection_name
                }

                if db_type == "sqlite":
                    config_data["file_path"] = db_conn.get("file_path", "")
                else: # mysql
                    host_from_conn = ""
                    username_from_conn = ""
                    password_from_conn = ""
                    port_from_conn = 3306
                    database_from_conn = ""

                    if connection_str:
                        from urllib.parse import urlparse
                        try:
                            parsed = urlparse(connection_str)
                            host_from_conn = parsed.hostname or ""
                            username_from_conn = parsed.username or ""
                            password_from_conn = parsed.password or ""
                            port_from_conn = parsed.port or 3306
                            if parsed.path:
                                database_from_conn = parsed.path.lstrip('/').split(':')[0]
                        except Exception as e:
                            logger.warning(f"Failed to parse connection string: {e}")

   
                    host = db_conn.get("host", host_from_conn) if db_conn.get("host") else host_from_conn
                    username = db_conn.get("username", username_from_conn) if db_conn.get("username") else username_from_conn

                    password = db_conn.get("password", password_from_conn)
                    port = db_conn.get("port", port_from_conn) if db_conn.get("port") else port_from_conn
                    database = db_conn.get("database", database_from_conn) if db_conn.get("database") else database_from_conn

                    config_data["host"] = host
                    config_data["port"] = port
                    config_data["database"] = database
                    config_data["username"] = username
                    config_data["password"] = password  # ⚠️ SECURITY WARNING: Stored in plaintext
                
                source_identifier = f"{connection_str}:{table_name}"
                existing_source = db.query(KBSourceConfig).filter(
                    KBSourceConfig.knowledge_base_id == kb_id,
                    KBSourceConfig.source_type == "database",
                    KBSourceConfig.source_identifier == source_identifier
                ).first()
                
                # 数据库源不使用ETag，直接保留原状态
                old_config = old_sources_map.get(source_identifier)
                if old_config:
                    status = old_config["status"]
                    chunks_created = old_config.get("chunks_created", 0)
                    tables_created = old_config.get("tables_created", "")
                    error_message = old_config.get("error_message")
                else:
                    status = "pending"
                    chunks_created = 0
                    tables_created = ""
                    error_message = None
                
                if existing_source:
                    existing_source.config = config_data
                    existing_source.status = status
                    existing_source.chunks_created = chunks_created
                    existing_source.tables_created = tables_created
                    existing_source.error_message = error_message
                    existing_source.updated_at = datetime.now(timezone.utc)
                else:
                    source = KBSourceConfig(
                        knowledge_base_id=kb_id,
                        source_type="database",
                        source_identifier=source_identifier,
                        config=config_data,
                        status=status,
                        chunks_created=chunks_created,
                        tables_created=tables_created,
                        error_message=error_message
                    )
                    db.add(source)
        
        # 6. Save or update build configuration
        build_config = db.query(KBBuildConfig).filter(
            KBBuildConfig.knowledge_base_id == kb_id
        ).first()
        
        if build_config:
            build_config.tools_config = tools_config
            build_config.updated_at = datetime.now(timezone.utc)
        else:
            build_config = KBBuildConfig(
                knowledge_base_id=kb_id,
                tools_config=tools_config,
                build_options={"parallel_processing": True, "max_workers": 4}
            )
            db.add(build_config)
        
        # 7. Update updated_at and commit
        kb.updated_at = datetime.now(timezone.utc)

        db.commit()
        
        return {
            "message": "Configuration updated successfully",
            "kb_id": kb_id,
            "kb_name": kb.name,
            "file_count": len(selected_files),
            "qa_file_count": len(selected_qa_files),
            "db_connections": len(db_connections),
            "cleanup_stats": {
                "removed_sources": len(removed_sources),
                "deleted_chunks": total_deleted_chunks,
                "deleted_tables": total_deleted_tables
            }
        }
    
    @staticmethod
    async def validate_qa_file(filename: str) -> Dict[str, Any]:
        try:
 
            file_data = minio_client.download_file(filename)
            if file_data is None:
                return {
                    "valid": False,
                    "filename": filename,
                    "errors": ["File not found in MinIO"]
                }

            workbook = openpyxl.load_workbook(file_data, read_only=True)

            sheet = None
            sheet_name = None

            possible_sheet_names = ["example"]
            for name in possible_sheet_names:
                if name in workbook.sheetnames:
                    sheet = workbook[name]
                    sheet_name = name
                    break

            if sheet is None:
                available_sheets = workbook.sheetnames if len(workbook.sheetnames) > 0 else []
                return {
                    "valid": False,
                    "filename": filename,
                    "errors": [
                        f"Required sheet 'example' not found",
                        f"Available sheets: {available_sheets}" if available_sheets else "No sheets found in the Excel file"
                    ]
                }

            required_headers_lower = ["question", "answer", "howtofind"]
            actual_headers = [str(cell.value).lower().strip() if cell.value else "" for cell in sheet[1]]

            header_mapping = {
                "question": ["question", "问题", "题目"],
                "answer": ["answer", "答案", "回答"],
                "howtofind": ["howtofind", "how to find", "查找方式"]
            }

            column_indices = {}  # {required_name: column_index}
            for i, actual_header in enumerate(actual_headers[:10]):  
                for required, aliases in header_mapping.items():
                    if actual_header in aliases and required not in column_indices:
                        column_indices[required] = i
                        break

            missing = [h for h in required_headers_lower if h not in column_indices]
            if missing:
                return {
                    "valid": False,
                    "filename": filename,
                    "sheet_name": sheet_name,
                    "errors": [
                        f"Missing required columns: {missing}",
                        f"Expected columns (one of): question(问题), answer(答案), howtofind(查找方式)",
                        f"Got headers: {actual_headers[:5]}"
                    ]
                }

            row_count = sum(1 for row in sheet.iter_rows(min_row=2, values_only=True) if any(row))


            sample_data = []
            for row in sheet.iter_rows(min_row=2, max_row=4, values_only=True):
                if any(row):
                    sample_data.append({
                        "question": str(row[column_indices["question"]]) if len(row) > column_indices["question"] and row[column_indices["question"]] else "",
                        "answer": str(row[column_indices["answer"]]) if len(row) > column_indices["answer"] and row[column_indices["answer"]] else "",
                        "howtofind": str(row[column_indices["howtofind"]]) if len(row) > column_indices["howtofind"] and row[column_indices["howtofind"]] else ""
                    })

            return {
                "valid": True,
                "filename": filename,
                "sheet_name": sheet_name,
                "row_count": row_count,
                "columns": list(column_indices.keys()),
                "column_indices": column_indices,  
                "errors": [],
                "sample_data": sample_data
            }
            
        except Exception as e:
            logger.error(f"Error validating QA file {filename}: {e}")
            return {
                "valid": False,
                "filename": filename,
                "errors": [str(e)]
            }
    
    @staticmethod
    async def test_database_connection(
        db_type: str,
        host: Optional[str] = None,
        port: Optional[int] = None,
        database: Optional[str] = None,
        username: Optional[str] = None,
        password: Optional[str] = None,
        file_path: Optional[str] = None
    ) -> Dict[str, Any]:
        """Test database connection.
        
        Args:
            db_type: Database type ("mysql" or "sqlite").
            host: MySQL host.
            port: MySQL port.
            database: MySQL database name.
            username: MySQL username.
            password: MySQL password.
            file_path: SQLite file path.
            
        Returns:
            A dictionary of connection test results.
        """
        try:
            tables = []
            
            if db_type == "sqlite":
                if not file_path:
                    return {
                        "success": False,
                        "message": "SQLite file path is required"
                    }
                
                import sqlite3
                conn = sqlite3.connect(file_path)
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tables = [row[0] for row in cursor.fetchall()]
                conn.close()
                
            elif db_type == "mysql":
                if not all([host, database, username]):
                    return {
                        "success": False,
                        "message": "MySQL connection requires host, database, and username"
                    }
                
                import pymysql
                conn = pymysql.connect(
                    host=host,
                    port=port or 3306,
                    database=database,
                    user=username,
                    password=password or ""
                )
                cursor = conn.cursor()
                cursor.execute("SHOW TABLES")
                tables = [row[0] for row in cursor.fetchall()]
                conn.close()
                
            else:
                return {
                    "success": False,
                    "message": f"Unsupported database type: {db_type}"
                }
            
            return {
                "success": True,
                "message": "Connection successful",
                "tables": tables
            }
            
        except Exception as e:
            logger.error(f"Database connection test failed: {e}")
            return {
                "success": False,
                "message": str(e)
            }
