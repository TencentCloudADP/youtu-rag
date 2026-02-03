"""Knowledge Base Configuration API routes."""

import io
import logging
import os
import re
from pathlib import Path
from typing import Any, Optional

import openpyxl
import yaml
from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.orm import Session
from datetime import datetime, timezone

from .database import (
    KnowledgeBase,
    KBSourceConfig,
    KBBuildConfig,
    KBBuildLog,
    get_db,
)
from .minio_client import MinIOClient
from .models.kb_config import KBBuildRequest, KBBuildResponse, QAValidationResult
from .utils import load_yaml_config
from ..knowledge_builder.agent import KnowledgeBuilderAgent, BuildRequest, SourceConfig

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/knowledge", tags=["knowledge-configuration"])


# Helper functions are now imported from utils (load_yaml_config)


class ToolConfig(BaseModel):
    """Configuration for a single tool."""
    enabled: bool
    settings: dict[str, Any]


class KBConfiguration(BaseModel):
    """Knowledge base configuration model."""
    tools: dict[str, ToolConfig]
    selectedFiles: list[str] = []
    selectedQAFiles: list[str] = []
    dbConnections: list[dict[str, Any]] = []


class KBConfigurationUpdate(BaseModel):
    """Request model for updating KB configuration."""
    configuration: KBConfiguration


@router.put("/{kb_id}/configuration")
async def update_kb_configuration(
    kb_id: int,
    config_update: KBConfigurationUpdate,
    db: Session = Depends(get_db)
):
    """
    Update knowledge base configuration (tools, files, connections).

    Args:
        kb_id: Knowledge base ID
        config_update: Configuration update data

    Returns:
        Updated configuration confirmation

    Example:
        ```
        PUT /api/knowledge/1/configuration
        {
            "configuration": {
                "tools": {
                    "embedding": {
                        "enabled": true,
                        "settings": {
                            "backend": "openai",
                            "model": "text-embedding-3-small"
                        }
                    },
                    "chunk": {
                        "enabled": true,
                        "settings": {
                            "chunk_size": 500,
                            "overlap": 50
                        }
                    }
                },
                "selectedFiles": ["doc1.pdf", "doc2.txt"],
                "selectedQAFiles": ["qa_examples.xlsx"],
                "dbConnections": []
            }
        }
        ```
    """
    try:
        from datetime import datetime, timezone

        kb = db.query(KnowledgeBase).filter(KnowledgeBase.id == kb_id).first()
        if not kb:
            raise HTTPException(status_code=404, detail="Knowledge base not found")

        config_summary = {
            "file_count": len(config_update.configuration.selectedFiles),
            "qa_file_count": len(config_update.configuration.selectedQAFiles),
            "db_connections": len(config_update.configuration.dbConnections)
        }

        # Used to preserve status and statistics.
        old_sources = db.query(KBSourceConfig).filter(
            KBSourceConfig.knowledge_base_id == kb_id
        ).all()

        old_sources_map = {
            src.source_identifier: {
                "status": src.status,
                "source_etag": src.source_etag,
                "chunks_created": src.chunks_created,
                "tables_created": src.tables_created,
                "error_message": src.error_message
            }
            for src in old_sources
        }

        # Find removed sources and cleanup
        new_source_identifiers = set()
        new_source_identifiers.update(config_update.configuration.selectedFiles)
        new_source_identifiers.update(config_update.configuration.selectedQAFiles)
        for db_conn in config_update.configuration.dbConnections:
            connection_str = db_conn.get("connectionString", "")
            for table_name in db_conn.get("tables", []):
                new_source_identifiers.add(f"{connection_str}:{table_name}")

        removed_sources = [src for src in old_sources if src.source_identifier not in new_source_identifiers]

        if removed_sources:
            from .services.kb_config_service import KBConfigService
            
            cleanup_stats = await KBConfigService.cleanup_removed_sources(
                removed_sources=removed_sources,
                collection_name=kb.collection_name
            )
            total_deleted_chunks = cleanup_stats["total_deleted_chunks"]
            total_deleted_tables = cleanup_stats["total_deleted_tables"]

            for removed_source in removed_sources:
                db.delete(removed_source)

        # UPSERT MinIO file sources
        minio_client = MinIOClient(
            endpoint=os.getenv("MINIO_ENDPOINT", "localhost:9000"),
            access_key=os.getenv("MINIO_ACCESS_KEY", "minioadmin"),
            secret_key=os.getenv("MINIO_SECRET_KEY", "minioadmin"),
            bucket_name=os.getenv("MINIO_BUCKET", "rag-documents"),
            secure=os.getenv("MINIO_SECURE", "false").lower() == "true"
        )

        for filename in config_update.configuration.selectedFiles:
            file_ext = filename.split('.')[-1].lower() if '.' in filename else 'txt'

            etag = None
            try:
                stat = minio_client.get_file_stat(filename)
                if stat:
                    etag = stat.etag.strip('"') if stat.etag else None
            except Exception as e:
                logger.warning(f"Failed to get ETag for {filename}: {e}")

            existing_source = db.query(KBSourceConfig).filter(
                KBSourceConfig.knowledge_base_id == kb_id,
                KBSourceConfig.source_type == "minio_file",
                KBSourceConfig.source_identifier == filename
            ).first()

            old_config = old_sources_map.get(filename)
            if old_config and old_config["status"] == "completed":
                if etag and old_config["source_etag"] and etag == old_config["source_etag"]:  # Content changes are captured by ETag
                    status = "completed"
                    chunks_created = old_config.get("chunks_created", 0)
                    error_message = None
                    logger.info(f"💾 Preserving completed status for {filename} (ETag unchanged)")
                else:
                    status = "pending"
                    chunks_created = 0
                    error_message = None
                    logger.info(f"🔄 Resetting status for {filename} (ETag changed)")
            else:  # New or previously failed
                status = old_config["status"] if old_config else "pending"
                chunks_created = old_config.get("chunks_created", 0) if old_config else 0
                error_message = old_config.get("error_message") if old_config else None

            if existing_source:  # Update existing record to maintain ID consistency
                existing_source.source_etag = etag
                existing_source.config = {
                    "file_path": filename,
                    "file_type": file_ext,
                    "kb_id": kb_id,
                    "collection_name": kb.collection_name
                }
                existing_source.status = status
                existing_source.chunks_created = chunks_created
                existing_source.error_message = error_message
                existing_source.updated_at = datetime.now(timezone.utc)
                logger.debug(f"Updated existing source config for {filename} (id={existing_source.id})")
            else:
                source = KBSourceConfig(
                    knowledge_base_id=kb_id,
                    source_type="minio_file",
                    source_identifier=filename,
                    source_etag=etag,
                    config={
                        "file_path": filename,
                        "file_type": file_ext,
                        "kb_id": kb_id,
                        "collection_name": kb.collection_name
                    },
                    status=status,
                    chunks_created=chunks_created,
                    error_message=error_message
                )
                db.add(source)
                logger.debug(f"Created new source config for {filename}")

        # UPSERT QA file sources
        for qa_filename in config_update.configuration.selectedQAFiles:
            etag = None
            try:
                stat = minio_client.get_file_stat(qa_filename)
                if stat:
                    etag = stat.etag.strip('"') if stat.etag else None
            except Exception as e:
                logger.warning(f"Failed to get ETag for QA file {qa_filename}: {e}")

            existing_source = db.query(KBSourceConfig).filter(
                KBSourceConfig.knowledge_base_id == kb_id,
                KBSourceConfig.source_type == "qa_file",
                KBSourceConfig.source_identifier == qa_filename
            ).first()

            old_config = old_sources_map.get(qa_filename)
            if old_config and old_config["status"] == "completed":
                if etag and old_config["source_etag"] and etag == old_config["source_etag"]:  # Content changes are captured by ETag
                    status = "completed"
                    chunks_created = old_config.get("chunks_created", 0)
                    error_message = None
                    logger.info(f"💾 Preserving completed status for QA file {qa_filename} (ETag unchanged)")
                else:
                    status = "pending"
                    chunks_created = 0
                    error_message = None
                    logger.info(f"🔄 Resetting status for QA file {qa_filename} (ETag changed)")
            else:  # New or previously failed
                status = old_config["status"] if old_config else "pending"
                chunks_created = old_config.get("chunks_created", 0) if old_config else 0
                error_message = old_config.get("error_message") if old_config else None

            if existing_source:  # Update existing record to maintain ID consistency
                existing_source.source_etag = etag
                existing_source.config = {
                    "file_path": qa_filename,
                    "sheet_name": "example",
                    "kb_id": kb_id,
                    "collection_name": kb.collection_name
                }
                existing_source.status = status
                existing_source.chunks_created = chunks_created
                existing_source.error_message = error_message
                existing_source.updated_at = datetime.now(timezone.utc)
                logger.debug(f"Updated existing QA source config for {qa_filename} (id={existing_source.id})")
            else:
                source = KBSourceConfig(
                    knowledge_base_id=kb_id,
                    source_type="qa_file",
                    source_identifier=qa_filename,
                    source_etag=etag,
                    config={
                        "file_path": qa_filename,
                        "sheet_name": "example",
                        "kb_id": kb_id,
                        "collection_name": kb.collection_name
                    },
                    status=status,
                    chunks_created=chunks_created,
                    error_message=error_message
                )
                db.add(source)
                logger.debug(f"Created new QA source config for {qa_filename}")

        # UPSERT database sources
        # ⚠️ SECURITY WARNING: Database passwords are stored in plaintext in the config field.
        # This is a security risk. In production environments, consider:
        # - Using encrypted storage (e.g., Fernet encryption)
        # - Storing passwords in environment variables
        # - Using a secrets management service
        for db_conn in config_update.configuration.dbConnections:
            logger.info(f"🔍 Processing db_conn: {db_conn}")
            logger.info(f"🔍 Password field: '{db_conn.get('password', 'KEY_NOT_FOUND')}'")

            connection_str = db_conn.get("connectionString", "")
            db_type = db_conn.get("type", "mysql")

            for table_name in db_conn.get("tables", []):
                config_data = {
                    "connection_string": connection_str,
                    "table_name": table_name,
                    "db_type": db_type
                }

                if db_type == "sqlite":  # For SQLite, store the file path
                    config_data["file_path"] = db_conn.get("file_path", "")
                else:  # For MySQL, store full connection info (including password)
                    config_data["host"] = db_conn.get("host", "")
                    config_data["port"] = db_conn.get("port", 3306)
                    config_data["database"] = db_conn.get("database", "")
                    config_data["username"] = db_conn.get("username", "")
                    config_data["password"] = db_conn.get("password", "")  # ⚠️ Plaintext storage

                config_data["kb_id"] = kb_id
                config_data["collection_name"] = kb.collection_name

                source_identifier = f"{connection_str}:{table_name}"
                existing_source = db.query(KBSourceConfig).filter(
                    KBSourceConfig.knowledge_base_id == kb_id,
                    KBSourceConfig.source_type == "database",
                    KBSourceConfig.source_identifier == source_identifier
                ).first()

                old_config = old_sources_map.get(source_identifier)
                if old_config and old_config["status"] == "completed":
                    # Notice that we don't check for data changes here.
                    # Use force_rebuild to force reprocess.
                    status = "completed"
                    chunks_created = old_config.get("chunks_created", 0)
                    tables_created = old_config.get("tables_created", "")
                    error_message = None
                    logger.info(f"💾 Preserving completed status for database source {source_identifier}")
                else:  # New or previously failed
                    status = old_config["status"] if old_config else "pending"
                    chunks_created = old_config.get("chunks_created", 0) if old_config else 0
                    tables_created = old_config.get("tables_created", "") if old_config else ""
                    error_message = old_config.get("error_message") if old_config else None

                if existing_source:  # Update existing record to maintain ID consistency
                    existing_source.config = config_data
                    existing_source.status = status
                    existing_source.chunks_created = chunks_created
                    existing_source.tables_created = tables_created
                    existing_source.error_message = error_message
                    existing_source.updated_at = datetime.now(timezone.utc)
                    logger.debug(f"Updated existing database source config for {source_identifier} (id={existing_source.id})")
                else:
                    source = KBSourceConfig(
                        knowledge_base_id=kb_id,
                        source_type="database",
                        source_identifier=source_identifier,
                        config=config_data,
                        status=status,
                        chunks_created=chunks_created,
                        tables_created=tables_created,
                        error_message=error_message
                    )
                    db.add(source)
                    logger.debug(f"Created new database source config for {source_identifier}")

        # Store or update tools configuration
        tools_config_dict = {}
        for tool_name, tool_config in config_update.configuration.tools.items():
            tools_config_dict[tool_name] = {
                "enabled": tool_config.enabled,
                "settings": tool_config.settings
            }

        build_config = db.query(KBBuildConfig).filter(
            KBBuildConfig.knowledge_base_id == kb_id
        ).first()

        if build_config:
            build_config.tools_config = tools_config_dict
            build_config.updated_at = datetime.now(timezone.utc)
        else:
            build_config = KBBuildConfig(
                knowledge_base_id=kb_id,
                tools_config=tools_config_dict,
                build_options={"parallel_processing": True, "max_workers": 4}
            )
            db.add(build_config)

        kb.updated_at = datetime.now(timezone.utc)

        db.commit()  # Finish update knowledge base configuration

        logger.info(f"Updated configuration for KB {kb_id}: {config_summary}")

        return {
            "message": "Configuration updated successfully",
            "kb_id": kb_id,
            "kb_name": kb.name,
            "configuration": config_update.configuration.dict(),
            "summary": config_summary
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Update configuration error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{kb_id}/build", response_model=KBBuildResponse)
async def build_knowledge_base(
    kb_id: int,
    build_request: KBBuildRequest,
    db: Session = Depends(get_db)
):
    """
    Build/rebuild knowledge base with configured tools and sources.

    This endpoint:
    1. Loads sources from configuration (MinIO files, databases, QA files)
    2. Processes sources in parallel using KnowledgeBuilderAgent
    3. Stores vectors in ChromaDB and structured data in SQLite
    4. Updates build status and logs

    Args:
        kb_id: Knowledge base ID
        build_request: Build options

    Returns:
        Build result with statistics

    Example:
        ```
        POST /api/knowledge/1/build
        {
            "force_rebuild": false,
            "file_filter": ["doc1.pdf"]
        }
        ```
    """
    from datetime import datetime

    logger.info("=" * 100)
    logger.info(f"🚀 BUILD API CALLED - KB ID: {kb_id}, force_rebuild: {build_request.force_rebuild}")
    logger.info("=" * 100)

    try:
        minio_client = MinIOClient(
            endpoint=os.getenv("MINIO_ENDPOINT", "localhost:9000"),
            access_key=os.getenv("MINIO_ACCESS_KEY", "minioadmin"),
            secret_key=os.getenv("MINIO_SECRET_KEY", "minioadmin"),
            bucket_name=os.getenv("MINIO_BUCKET", "rag-documents"),
            secure=os.getenv("MINIO_SECURE", "false").lower() == "true"
        )

        # Check if KB exists
        kb = db.query(KnowledgeBase).filter(KnowledgeBase.id == kb_id).first()
        if not kb:
            raise HTTPException(status_code=404, detail="Knowledge base not found")

        sources = db.query(KBSourceConfig).filter(
            KBSourceConfig.knowledge_base_id == kb_id
        ).all()

        if not sources:
            raise HTTPException(
                status_code=400,
                detail="No sources configured for this knowledge base. Please save configuration first."
            )

        # Notice that tools configurations are loaded from YAML instead of database
        from ..knowledge_builder.agent import ToolsConfig
        yaml_config = load_yaml_config(kb.name)

        embedding_config = yaml_config.get('embedding', {})
        reranker_config = yaml_config.get('reranker', {})
        llm_config = yaml_config.get('llm_model', {})

        tools_config = ToolsConfig(
            semantic_retrieval=True,  # Always enabled if embedding config exists
            embedding_model=embedding_config.get('model', 'text-embedding-3-small'),
            embedding_type=embedding_config.get('type', 'api'),  # "local" or "api"
            embedding_base_url=embedding_config.get('base_url'),  # Base URL for embedding service
            reranker_model=reranker_config.get('model', 'jina-reranker-v2') if reranker_config.get('enabled', True) else None,
            text2sql_enabled=True,  # Always enabled
            sql_generator_model=llm_config.get('model', 'gpt-4')
        )

        logger.info(f"Loaded tools config from YAML: embedding_model={tools_config.embedding_model}, "
                   f"embedding_type={tools_config.embedding_type}, embedding_base_url={tools_config.embedding_base_url}, "
                   f"reranker_model={tools_config.reranker_model}, sql_generator_model={tools_config.sql_generator_model}")

        # Apply file filter if specified
        if build_request.file_filter:
            sources = [s for s in sources if s.source_identifier in build_request.file_filter]

        if not sources:
            raise HTTPException(status_code=400, detail="No sources to process after applying filter")

        # Skip completed sources if not force_rebuild and ETag hasn't changed
        skipped_sources = []
        sources_to_process = []

        if not build_request.force_rebuild:
            for source in sources:
                should_skip = False

                # Only check for completed sources
                if source.status == "completed":
                    # For MinIO files and QA files, check if ETag and metadata have changed
                    if source.source_type in ["minio_file", "qa_file"]:
                        try:
                            stat = minio_client.get_file_stat(source.source_identifier)
                            current_etag = stat.etag.strip('"') if stat and stat.etag else None

                            import hashlib
                            import json
                            current_metadata = minio_client.get_file_metadata(source.source_identifier) or {}
                            current_metadata_hash = None
                            if current_metadata:
                                # MinIO files store only original fields, while kb_source_configs tables store hash
                                metadata_json = json.dumps(current_metadata, sort_keys=True, ensure_ascii=False)
                                current_metadata_hash = hashlib.md5(metadata_json.encode('utf-8')).hexdigest()

                            etag_unchanged = current_etag and source.source_etag and current_etag == source.source_etag
                            metadata_unchanged = current_metadata_hash and source.metadata_hash and current_metadata_hash == source.metadata_hash

                            # Check if derived files have changed, including OCR- and Chunk-processed files
                            derived_files_unchanged = True
                            current_derived_etags = []

                            if current_metadata.get("chunk_processed") == "chunk_success":  # Chunk-processed file
                                sys_bucket = os.getenv("MINIO_BUCKET_SYS", "sysfile")
                                from pathlib import Path
                                chunked_filename = f"{Path(source.source_identifier).stem}_chunklevel.md"

                                try:
                                    chunk_stat = minio_client.get_file_stat(chunked_filename, bucket_name=sys_bucket)
                                    if chunk_stat and hasattr(chunk_stat, 'etag'):
                                        current_derived_etags.append(chunk_stat.etag.strip('"'))
                                except Exception as e:
                                    logger.debug(f"Chunk file {chunked_filename} not found or error: {e}")

                            if current_metadata.get("ocr_processed") == "ocr_success":  # OCR-processed file
                                sys_bucket = os.getenv("MINIO_BUCKET_SYS", "sysfile")

                                _, ocr_derived_etags = minio_client.load_derived_markdown_files(
                                    source_filename=source.source_identifier,
                                    sys_bucket=sys_bucket
                                )

                                if ocr_derived_etags:
                                    current_derived_etags.extend(ocr_derived_etags)

                            # Calculate combined hash if we have any derived files
                            if current_derived_etags:
                                current_derived_hash = minio_client.calculate_derived_files_hash(current_derived_etags)
                                derived_files_unchanged = (
                                    source.derived_files_hash and
                                    current_derived_hash == source.derived_files_hash
                                )

                                if not derived_files_unchanged:
                                    logger.info(
                                        f"🔄 Derived files changed for {source.source_identifier}: "
                                        f"{source.derived_files_hash[:8] if source.derived_files_hash else 'None'}... -> {current_derived_hash[:8]}..."
                                    )

                            # Skip only if all three (ETag, metadata, derived files) are unchanged
                            if etag_unchanged and metadata_unchanged and derived_files_unchanged:
                                should_skip = True
                                logger.info(f"⏭️  Skipping {source.source_identifier} (already completed, all unchanged)")
                            else:
                                change_reasons = []
                                if not etag_unchanged:
                                    change_reasons.append(f"ETag: {source.source_etag[:8] if source.source_etag else 'None'}... -> {current_etag[:8] if current_etag else 'None'}...")
                                if not metadata_unchanged:
                                    change_reasons.append(f"Metadata: {source.metadata_hash[:8] if source.metadata_hash else 'None'}... -> {current_metadata_hash[:8] if current_metadata_hash else 'None'}...")
                                if not derived_files_unchanged:
                                    change_reasons.append(f"Derived files changed")
                                logger.info(f"🔄 Processing {source.source_identifier} ({'; '.join(change_reasons)})")
                        except Exception as e:
                            logger.warning(f"Failed to check ETag/metadata for {source.source_identifier}: {e}, will process it")

                    # For database sources, check if source_etag (checksum/row count) has changed
                    elif source.source_type == "database":
                        # For databases, we always reprocess unless we implement checksum tracking
                        # This is a simpler approach - skip if completed and no explicit rebuild requested
                        should_skip = True
                        logger.info(f"⏭️  Skipping database source {source.source_identifier} (already completed)")

                if should_skip:
                    skipped_sources.append(source)
                else:
                    sources_to_process.append(source)
        else:
            # Force rebuild - process all sources
            sources_to_process = sources
            logger.info(f"🔄 Force rebuild enabled - processing all {len(sources)} sources")

        sources = sources_to_process

        if not sources and not skipped_sources:
            raise HTTPException(status_code=400, detail="No sources to process after applying filter")

        if not sources and skipped_sources:
            # All sources were skipped - return success without processing
            logger.info(f"✅ All {len(skipped_sources)} sources already completed and unchanged - skipping build")
            return KBBuildResponse(
                status="success",
                message="All sources already completed and unchanged",
                kb_id=kb_id,
                kb_name=kb.name,
                total_files=len(skipped_sources),
                processed_files=0,
                skipped_files=len(skipped_sources),
                total_chunks=0,
                errors=[]
            )

        logger.info(f"📊 Build summary: {len(sources)} to process, {len(skipped_sources)} skipped")

        # Create build log
        build_log = KBBuildLog(
            knowledge_base_id=kb_id,
            status="running",
            total_sources=len(sources),
            processed_sources=0,
            total_chunks=0,
            total_tables=0,
            start_time=datetime.now(timezone.utc)
        )
        db.add(build_log)
        db.commit()
        db.refresh(build_log)

        logger.info(f"Starting build for KB {kb_id} with {len(sources)} sources")

        # Prepare BuildRequest for KnowledgeBuilderAgent
        source_configs = [
            SourceConfig(
                source_type=s.source_type,
                source_identifier=s.source_identifier,
                config={**(s.config or {}), "source_id": s.id},  # Add source_id to config
                source_etag=s.source_etag,
                metadata_hash=s.metadata_hash,  # Add metadata_hash
                status=s.status
            )
            for s in sources
        ]

        def on_progress(progress):
            """The callback function to update the status of kb_source_configs"""
            try:
                if progress.latest_result:
                    result = progress.latest_result

                    source = db.query(KBSourceConfig).filter(
                        KBSourceConfig.knowledge_base_id == kb_id,
                        KBSourceConfig.source_identifier == result.source_identifier
                    ).first()

                    if source:
                        source.status = result.status

                        if result.status == "failed":
                            source.error_message = result.error_message
                        elif result.status == "completed":
                            source.error_message = None

                            source.chunks_created = result.chunks_created
                            if result.tables_created:
                                source.tables_created = ','.join(result.tables_created)
                            else:
                                source.tables_created = None

                            # Update ETag for MinIO files and QA files
                            if source.source_type in ["minio_file", "qa_file"]:
                                try:
                                    stat = minio_client.get_file_stat(source.source_identifier)
                                    if stat and stat.etag:
                                        source.source_etag = stat.etag.strip('"')
                                        logger.info(f"Real-time ETag update: {source.source_identifier} -> {source.source_etag[:8]}...")
                                except Exception as e:
                                    logger.warning(f"Failed to update ETag in progress callback: {e}")

                        db.commit()
                        logger.info(f"Real-time status update: {result.source_identifier} -> {result.status}")

            except Exception as e:
                logger.error(f"Progress callback error: {e}")
                db.rollback()

        agent_request = BuildRequest(
            knowledge_base_id=kb_id,
            kb_name=kb.name,
            collection_name=kb.collection_name,
            sources=source_configs,
            tools_config=tools_config,
            force_rebuild=build_request.force_rebuild,
            progress_callback=on_progress
        )

        logger.info("=" * 80)
        logger.info("Build Request Details:")
        logger.info("=" * 80)
        logger.info(f"  KB ID: {agent_request.knowledge_base_id}")
        logger.info(f"  KB Name: {agent_request.kb_name}")
        logger.info(f"  Collection Name: {agent_request.collection_name}")
        logger.info(f"  Force Rebuild: {agent_request.force_rebuild}")
        logger.info(f"\n  Tools Configuration:")
        logger.info(f"    - Semantic Retrieval: {agent_request.tools_config.semantic_retrieval}")
        logger.info(f"    - Embedding Model: {agent_request.tools_config.embedding_model}")
        logger.info(f"    - Reranker Model: {agent_request.tools_config.reranker_model}")
        logger.info(f"    - Text2SQL Enabled: {agent_request.tools_config.text2sql_enabled}")
        logger.info(f"    - SQL Generator Model: {agent_request.tools_config.sql_generator_model}")

        # Count data sources by type
        from collections import defaultdict
        source_type_stats = defaultdict(int)
        file_type_stats = defaultdict(int)
        for src in agent_request.sources:
            source_type_stats[src.source_type] += 1
            if src.source_type == "minio_file":
                file_type = src.config.get("file_type", "unknown")
                file_type_stats[file_type] += 1

        logger.info(f"\n  Data Sources: {len(agent_request.sources)} total")
        for src_type, count in source_type_stats.items():
            logger.info(f"    - {src_type}: {count}")

        if file_type_stats:
            logger.info(f"\n  File Types:")
            for file_type, count in sorted(file_type_stats.items()):
                logger.info(f"    - {file_type.upper()}: {count}")

        logger.info(f"\n  Source Details:")
        for idx, src in enumerate(agent_request.sources, 1):
            etag_info = f" (ETag: {src.source_etag[:8]}...)" if src.source_etag else ""
            status_info = f" [Status: {src.status}]" if src.status else ""
            logger.info(f"    {idx}. [{src.source_type}] {src.source_identifier}{etag_info}{status_info}")

        logger.info("=" * 80)

        # Initialize and run KnowledgeBuilderAgent
        # Note: vector_store and relational_db_path will be initialized in Stage 0
        # based on kb_id (each KB gets independent storage)
        agent = KnowledgeBuilderAgent(
            vector_store=None,  # Will be initialized in Stage 0 as kb_{kb_id} collection
            relational_db_path=None,  # Will be initialized in Stage 0 as kb_{kb_id}.sqlite
            minio_client=minio_client  # Pass MinIO client for incremental build support
        )

        build_report = await agent.build(agent_request)

        # Update source statuses, ETags, metadata hashes, chunks_created, and tables_created
        for source in sources:
            # Find corresponding result from detailed results
            matching_result = next(
                (r for r in build_report.results if r.source_identifier == source.source_identifier),
                None
            )

            if matching_result:
                source.status = matching_result.status
                source.error_message = matching_result.error_message if matching_result.status == "failed" else None

                if matching_result.status == "completed":
                    source.chunks_created = matching_result.chunks_created
                    if matching_result.tables_created:
                        source.tables_created = ','.join(matching_result.tables_created)
                    else:
                        source.tables_created = None
            else:
                # Fallback: check errors list
                matching_errors = [r for r in build_report.errors if source.source_identifier in str(r)]
                if matching_errors:
                    source.status = "failed"
                    source.error_message = matching_errors[0]
                else:
                    # Default to completed if no error found
                    source.status = "completed"

            # Update ETag and metadata hash for MinIO files and QA files after successful processing
            if source.status == "completed" and source.source_type in ["minio_file", "qa_file"]:
                try:
                    # Update ETag
                    stat = minio_client.get_file_stat(source.source_identifier)
                    if stat and stat.etag:
                        source.source_etag = stat.etag.strip('"')
                        logger.info(f"Updated ETag for {source.source_identifier}: {source.source_etag[:8]}...")

                    # Update metadata hash
                    # Notice that MinIO files store only raw fields.
                    import hashlib
                    import json
                    current_metadata = minio_client.get_file_metadata(source.source_identifier) or {}
                    if current_metadata:
                        metadata_json = json.dumps(current_metadata, sort_keys=True, ensure_ascii=False)
                        source.metadata_hash = hashlib.md5(metadata_json.encode('utf-8')).hexdigest()
                        logger.info(f"Updated metadata hash for {source.source_identifier}: {source.metadata_hash[:8]}...")

                    # Update derived_files_hash if OCR-derived files were used
                    if matching_result and matching_result.metadata:
                        derived_etags = matching_result.metadata.get("derived_files_etags", [])
                        if derived_etags:
                            derived_hash = minio_client.calculate_derived_files_hash(derived_etags)
                            source.derived_files_hash = derived_hash
                            logger.info(f"Updated derived_files_hash for {source.source_identifier}: {derived_hash[:8]}...")
                except Exception as e:
                    logger.warning(f"Failed to update ETag/metadata hash for {source.source_identifier}: {e}")

        # Update build log
        build_log.status = build_report.status
        build_log.processed_sources = build_report.successful + build_report.failed
        build_log.total_chunks = build_report.total_chunks
        build_log.total_tables = build_report.total_tables
        build_log.end_time = datetime.now(timezone.utc)
        build_log.duration_seconds = int(build_report.duration_seconds)
        build_log.result_detail = {
            "successful": build_report.successful,
            "failed": build_report.failed,
            "skipped": len(skipped_sources),
            "errors": build_report.errors,
            "qa_validation": build_report.qa_validation
        }

        kb.updated_at = datetime.now(timezone.utc)

        db.commit()

        logger.info(f"Build completed for KB {kb_id}: {build_report.status}")

        total_skipped = len(skipped_sources) + build_report.skipped
        if skipped_sources:
            logger.info(f"Skipped {len(skipped_sources)} sources (pre-filtered)")
        if build_report.skipped > 0:
            logger.info(f"Skipped {build_report.skipped} sources (agent-skipped)")
        if total_skipped > 0:
            logger.info(f"Total skipped: {total_skipped} sources")

        return KBBuildResponse(
            status="success" if build_report.status == "completed" else "partial",
            message=f"Knowledge base build {build_report.status}. Processed: {build_report.successful + build_report.failed}, Skipped: {total_skipped}",
            kb_id=kb_id,
            kb_name=kb.name,
            total_files=build_report.total_sources + len(skipped_sources),
            processed_files=build_report.successful + build_report.failed,
            skipped_files=total_skipped,
            total_chunks=build_report.total_chunks,
            errors=build_report.errors
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Build knowledge base error: {str(e)}", exc_info=True)

        # Update build log if it exists
        if 'build_log' in locals():
            build_log.status = "failed"
            build_log.end_time = datetime.now(timezone.utc)
            build_log.result_detail = {"error": str(e)}
            db.commit()

        raise HTTPException(status_code=500, detail=str(e))


@router.get("/files/validate-qa/{filename}", response_model=QAValidationResult)
async def validate_qa_file(filename: str):
    """
    Validate QA Excel file format.

    Expected format:
    - Sheet name: "example"
    - Headers: "question", "answer", "howtofind"

    Args:
        filename: QA Excel file name in MinIO

    Returns:
        Validation result with details

    Example:
        ```
        GET /api/knowledge/files/validate-qa/qa_examples.xlsx
        ```
    """
    try:
        minio_client = MinIOClient(
            endpoint=os.getenv("MINIO_ENDPOINT", "localhost:9000"),
            access_key=os.getenv("MINIO_ACCESS_KEY", "minioadmin"),
            secret_key=os.getenv("MINIO_SECRET_KEY", "minioadmin"),
            bucket_name=os.getenv("MINIO_BUCKET", "rag-documents"),
            secure=os.getenv("MINIO_SECURE", "false").lower() == "true"
        )

        file_data = minio_client.download_file(filename)
        if file_data is None:
            raise HTTPException(status_code=404, detail=f"File '{filename}' not found in storage")

        # Read Excel file
        file_bytes = file_data.read()
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))

        errors = []
        valid = True

        # Check for sheet "example"
        if "example" not in workbook.sheetnames:
            errors.append("Missing required sheet 'example'")
            valid = False
            return QAValidationResult(
                valid=False,
                filename=filename,
                errors=errors
            )

        sheet = workbook["example"]

        # Check headers (handle both string and non-string cell values)
        required_headers = ["question", "answer", "howtofind"]
        header_row = [str(cell.value).lower().strip() if cell.value else "" for cell in sheet[1]]

        missing_headers = [h for h in required_headers if h not in header_row]
        if missing_headers:
            errors.append(f"Missing required columns: {', '.join(missing_headers)}. Found columns: {header_row}")
            valid = False

        # Get column indices
        try:
            question_col = header_row.index("question")
            answer_col = header_row.index("answer")
            howtofind_col = header_row.index("howtofind")
        except ValueError as e:
            return QAValidationResult(
                valid=False,
                filename=filename,
                sheet_name="example",
                columns=header_row,
                errors=errors
            )

        row_count = sheet.max_row - 1  # excluding header

        # Validate data rows
        empty_rows = []
        for row_idx in range(2, min(sheet.max_row + 1, 12)):  # Check first 10 data rows
            row = sheet[row_idx]
            question = row[question_col].value
            answer = row[answer_col].value

            if not question or not answer:
                empty_rows.append(row_idx)

        if empty_rows:
            errors.append(f"Empty question or answer in rows: {', '.join(map(str, empty_rows))}")
            if len(empty_rows) > 5:
                valid = False

        # Get sample data (first 3 rows)
        sample_data = []
        for row_idx in range(2, min(sheet.max_row + 1, 5)):
            row = sheet[row_idx]
            sample_data.append({
                "question": str(row[question_col].value or ""),
                "answer": str(row[answer_col].value or ""),
                "howtofind": str(row[howtofind_col].value or "")
            })

        return QAValidationResult(
            valid=valid,
            filename=filename,
            sheet_name="example",
            row_count=row_count,
            columns=header_row,
            errors=errors,
            sample_data=sample_data if valid else None
        )

    except HTTPException:
        raise
    except openpyxl.utils.exceptions.InvalidFileException:
        logger.error(f"Invalid Excel file: {filename}")
        raise HTTPException(status_code=400, detail="Invalid Excel file format")
    except Exception as e:
        logger.error(f"QA file validation error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Validation error: {str(e)}")


# Database Connection Testing Models
class DBConnectionTestRequest(BaseModel):
    """Request model for testing database connection."""
    db_type: str  # mysql, postgresql, sqlite
    host: Optional[str] = None
    port: Optional[int] = None
    database: str
    username: Optional[str] = None
    password: Optional[str] = None
    file_path: Optional[str] = None  # For SQLite


class DBConnectionTestResponse(BaseModel):
    """Response model for database connection test."""
    success: bool
    message: str
    tables: list[str] = []
    error: Optional[str] = None


@router.post("/database/test-connection", response_model=DBConnectionTestResponse)
async def test_database_connection(
    request: DBConnectionTestRequest
):
    """
    Test database connection and retrieve list of tables.

    Supports:
    - MySQL: Requires host, port, database, username, password
    - SQLite: Requires file_path

    Returns list of available tables on success.
    """
    try:
        tables = []

        if request.db_type == "sqlite":  # Handle SQLite
            if not request.file_path:
                raise HTTPException(
                    status_code=400,
                    detail="file_path is required for SQLite"
                )

            import sqlite3

            # Check if file exists
            if not os.path.exists(request.file_path):
                return DBConnectionTestResponse(
                    success=False,
                    message="SQLite file not found",
                    error=f"File does not exist: {request.file_path}"
                )

            try:
                conn = sqlite3.connect(request.file_path)
                cursor = conn.cursor()

                # Get list of tables
                cursor.execute("""
                    SELECT name FROM sqlite_master
                    WHERE type='table' AND name NOT LIKE 'sqlite_%'
                    ORDER BY name
                """)
                tables = [row[0] for row in cursor.fetchall()]

                conn.close()

                if not tables:
                    return DBConnectionTestResponse(
                        success=True,
                        message="Connected successfully, but no tables found",
                        tables=[]
                    )

                return DBConnectionTestResponse(
                    success=True,
                    message=f"Connected successfully. Found {len(tables)} table(s).",
                    tables=tables
                )

            except sqlite3.Error as e:
                return DBConnectionTestResponse(
                    success=False,
                    message="Failed to connect to SQLite database",
                    error=str(e)
                )

        elif request.db_type == "mysql":  # Handle MySQL
            if not all([request.host, request.port, request.database, request.username]):
                raise HTTPException(
                    status_code=400,
                    detail="host, port, database, and username are required for MySQL"
                )

            try:
                import pymysql
            except ImportError:
                return DBConnectionTestResponse(
                    success=False,
                    message="MySQL driver not installed",
                    error="pymysql package is not installed"
                )

            try:
                conn = pymysql.connect(
                    host=request.host,
                    port=request.port,
                    user=request.username,
                    password=request.password or "",
                    database=request.database,
                    connect_timeout=10
                )
                cursor = conn.cursor()

                # Get list of tables
                cursor.execute("SHOW TABLES")
                tables = [row[0] for row in cursor.fetchall()]

                conn.close()

                if not tables:
                    return DBConnectionTestResponse(
                        success=True,
                        message="Connected successfully, but no tables found",
                        tables=[]
                    )

                return DBConnectionTestResponse(
                    success=True,
                    message=f"Connected successfully. Found {len(tables)} table(s).",
                    tables=tables
                )

            except pymysql.Error as e:
                return DBConnectionTestResponse(
                    success=False,
                    message="Failed to connect to MySQL database",
                    error=str(e)
                )

        else:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported database type: {request.db_type}"
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Database connection test error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
