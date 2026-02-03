"""
Schema Loader
从 Meta Graph 加载表格 Schema 信息，供 ADG Workflow 使用
如果找不到预构建的 Meta Graph，会自动使用 meta_gen 生成层级信息
"""

import json
import sys
import os
from typing import Dict, Any, List, Optional
from pathlib import Path

# 修正导入路径
sys.path.append(str(Path(__file__).parent.parent.parent.parent))

from retrieval.schema_linking_table import SchemaLinking
from utils.logger import logger


class SchemaLoader:
    """Schema Loader - 支持自动生成 Meta Graph"""
    
    def __init__(self, meta_graphs_dir: Optional[str] = None, auto_generate: bool = True):
        """初始化
        
        Args:
            meta_graphs_dir: meta graph 文件所在目录
            auto_generate: 如果找不到 meta graph，是否自动生成（默认 True）
        """
        self.schema_linking = SchemaLinking(meta_graphs_dir)
        self.meta_graphs_dir = self.schema_linking.meta_graphs_dir
        self.auto_generate = auto_generate
        logger.info(f"SchemaLoader initialized with meta_graphs_dir: {self.meta_graphs_dir}, auto_generate: {auto_generate}")
        
    def _clean_column_name(self, name: str) -> str:
        """清洗列名：移除换行符，合并空格"""
        if not name or not isinstance(name, str):
            return str(name)
        # 替换换行符为空格
        clean = name.replace('\n', ' ').replace('\r', ' ')
        # 合并多余空格
        clean = ' '.join(clean.split())
        return clean.strip()

    def _get_header_value(self, props: Dict[str, Any]) -> Optional[str]:
        """从属性中提取表头值"""
        return props.get("value") or props.get("value_text") or props.get("text")

    def _is_valid_header_value(self, value: Any) -> bool:
        """检查表头值是否有效"""
        if not value:
            return False
        if isinstance(value, str):
            return bool(value.strip()) and not value.startswith('[EMPTY_')
        return True

    def _col_header_priority(self, item: tuple) -> int:
        """计算列表头的优先级"""
        _, props = item
        value = self._get_header_value(props)
        if not value or (isinstance(value, str) and not value.strip()):
            return 999
        if any(kw in str(value) for kw in ['评估', '等级', '星', '定义']):
            return 0
        return 1

    def _extract_column_and_row_headers(self, entities: List[Dict[str, Any]]) -> tuple:
        """从 entities 中分离列表头和行表头"""
        col_headers = []
        row_headers = []

        for ent in entities:
            label = ent.get("label", "")
            props = ent.get("properties", {}) or {}
            if label == "column_header":
                col_headers.append((ent.get("id"), props))
            elif label == "row_header":
                row_headers.append((ent.get("id"), props))

        return col_headers, row_headers

    def _extract_table_column_triplets(self, col_headers: List[tuple], max_headers: int) -> List[str]:
        """提取表格与列表头的关系三元组"""
        triplets = []

        # 获取 Level 0 的列表头
        level0_col_headers = [(hid, props) for hid, props in col_headers if props.get("level") == 0]
        if not level0_col_headers:
            level0_col_headers = col_headers

        sorted_col_headers = sorted(level0_col_headers, key=self._col_header_priority)

        for _, props in sorted_col_headers[:max_headers]:
            value = self._get_header_value(props)
            if value:
                value_clean = self._clean_column_name(str(value))[:60]
                triplets.append(f'(table, has_column_header, "{value_clean}")')

        return triplets

    def _build_header_parent_map(self, relationships: List[Dict[str, Any]]) -> Dict[str, str]:
        """构建表头父子关系映射"""
        header_parent_map = {}
        for rel in relationships:
            if rel.get("relation") == "has_child":
                header_parent_map[rel.get("end_entity_id")] = rel.get("start_entity_id")
        return header_parent_map

    def _extract_hierarchy_triplets_from_entities(
        self,
        col_headers: List[tuple],
        row_headers: List[tuple],
        entity_map: Dict[str, Dict[str, Any]],
        header_parent_map: Dict[str, str],
        max_hierarchy: int
    ) -> List[str]:
        """从 entities 中提取层级关系三元组"""
        triplets = []
        hierarchy_count = 0
        sorted_headers = sorted(col_headers + row_headers, key=lambda x: x[1].get("level", 99))

        for header_id, props in sorted_headers:
            if hierarchy_count >= max_hierarchy:
                break

            value = self._get_header_value(props)
            if not value:
                continue

            value_clean = self._clean_column_name(str(value))[:60]

            if header_id in header_parent_map:
                parent_id = header_parent_map[header_id]
                parent_ent = entity_map.get(parent_id)
                if parent_ent:
                    p_props = parent_ent.get("properties", {})
                    p_val = self._get_header_value(p_props)
                    if p_val:
                        p_val_clean = self._clean_column_name(str(p_val))[:60]
                        triplets.append(f'("{p_val_clean}", has_child, "{value_clean}")')
                        hierarchy_count += 1

        return triplets

    def _extract_triplets_from_entities(
        self,
        entities: List[Dict[str, Any]],
        relationships: List[Dict[str, Any]],
        max_headers: int,
        max_hierarchy: int
    ) -> List[str]:
        """从完整的 entities 和 relationships 中提取三元组"""
        triplets = []

        entity_map = {ent.get("id"): ent for ent in entities}
        col_headers, row_headers = self._extract_column_and_row_headers(entities)

        # 1. 提取表格与列表头的关系
        triplets.extend(self._extract_table_column_triplets(col_headers, max_headers))

        # 2. 提取表头层级关系
        header_parent_map = self._build_header_parent_map(relationships)
        triplets.extend(self._extract_hierarchy_triplets_from_entities(
            col_headers, row_headers, entity_map, header_parent_map, max_hierarchy
        ))

        return triplets

    def _extract_hierarchy_from_level(
        self,
        level_headers: List[Dict[str, Any]],
        parent_level: List[Dict[str, Any]],
        hierarchy_count: int,
        max_hierarchy: int
    ) -> tuple:
        """从单个层级中提取层级关系三元组"""
        triplets = []

        for header in level_headers:
            if hierarchy_count >= max_hierarchy:
                break

            value = header.get("value") or header.get("text", "")
            if not self._is_valid_header_value(value):
                continue

            value_clean = self._clean_column_name(str(value))[:60]

            # 查找父表头（通过位置关系）
            parent_value = None
            if parent_level:
                header_idx = level_headers.index(header) if header in level_headers else -1
                if 0 <= header_idx < len(parent_level):
                    parent_header = parent_level[header_idx]
                    parent_value = parent_header.get("value") or parent_header.get("text", "")

            if parent_value:
                parent_clean = self._clean_column_name(str(parent_value))[:60]
                triplets.append(f'("{parent_clean}", has_child, "{value_clean}")')
                hierarchy_count += 1

        return triplets, hierarchy_count

    def _extract_hierarchy_triplets_from_levels(
        self,
        header_levels: List[List[Dict[str, Any]]],
        max_hierarchy: int,
        hierarchy_count: int = 0
    ) -> tuple:
        """从层级列表中提取层级关系三元组"""
        triplets = []

        for level_idx, level_headers in enumerate(header_levels):
            if hierarchy_count >= max_hierarchy:
                break
            if level_idx == 0:
                continue  # 跳过第一层

            parent_level = header_levels[level_idx - 1] if level_idx > 0 else []
            level_triplets, hierarchy_count = self._extract_hierarchy_from_level(
                level_headers, parent_level, hierarchy_count, max_hierarchy
            )
            triplets.extend(level_triplets)

        return triplets, hierarchy_count

    def _extract_triplets_from_hierarchy(
        self,
        meta_graph: Dict[str, Any],
        max_headers: int,
        max_hierarchy: int
    ) -> List[str]:
        """从简化的 hierarchy 信息中提取三元组"""
        triplets = []

        meta_info = meta_graph.get("meta_info", {})
        col_header_names = meta_info.get("col_header_names", [])

        # 1. 提取表格与列表头的关系
        for col_name in col_header_names[:max_headers]:
            if self._is_valid_header_value(col_name):
                value_clean = self._clean_column_name(str(col_name))[:60]
                triplets.append(f'(table, has_column_header, "{value_clean}")')

        # 2. 从 hierarchy 中提取层级关系
        hierarchy = meta_graph.get("hierarchy", {})
        if hierarchy:
            col_header_levels = hierarchy.get("col_header_levels", [])
            row_header_levels = hierarchy.get("row_header_levels", [])

            hierarchy_count = 0

            # 处理列表头层级
            col_triplets, hierarchy_count = self._extract_hierarchy_triplets_from_levels(
                col_header_levels, max_hierarchy, hierarchy_count
            )
            triplets.extend(col_triplets)

            # 处理行表头层级
            row_triplets, hierarchy_count = self._extract_hierarchy_triplets_from_levels(
                row_header_levels, max_hierarchy, hierarchy_count
            )
            triplets.extend(row_triplets)

        return triplets

    def _extract_rich_triplets(self, meta_graph: Dict[str, Any], max_headers: int = 30, max_hierarchy: int = 20) -> List[str]:
        """从 meta graph 中提取丰富的三元组信息（支持层级）

        支持两种格式：
        1. 完整的 meta graph（有 entities 和 relationships）
        2. 简化的 meta graph（只有 hierarchy 信息）
        """
        entities = meta_graph.get("entities", [])
        relationships = meta_graph.get("relationships", [])

        if entities:
            triplets = self._extract_triplets_from_entities(entities, relationships, max_headers, max_hierarchy)
        else:
            triplets = self._extract_triplets_from_hierarchy(meta_graph, max_headers, max_hierarchy)

        return list(dict.fromkeys(triplets))

    
    def _load_excel_worksheet(self, excel_path: str):
        """加载Excel工作表

        Returns:
            worksheet 或 None 如果失败
        """
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(excel_path, data_only=False)
            worksheet = workbook.active
            return worksheet
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            return None

    def _create_region_from_worksheet(self, worksheet) -> Dict[str, int]:
        """从工作表创建区域定义"""
        return {
            "min_row": 1,
            "max_row": worksheet.max_row,
            "min_col": 1,
            "max_col": worksheet.max_column
        }

    def _extract_merged_cells(self, worksheet) -> List[Dict[str, int]]:
        """提取合并单元格信息"""
        merged_cells = []
        for merged_range in worksheet.merged_cells.ranges:
            merged_cells.append({
                "min_row": merged_range.min_row,
                "max_row": merged_range.max_row,
                "min_col": merged_range.min_col,
                "max_col": merged_range.max_col
            })
        return merged_cells

    def _execute_tool_safely(self, tool_registry, tool_name: str, *args) -> Dict[str, Any]:
        """安全执行工具，捕获异常并返回空字典"""
        tool = tool_registry.get(tool_name)
        if not tool:
            return {}

        try:
            return tool.execute(*args)
        except Exception as e:
            logger.warning(f"{tool_name} failed: {e}")
            return {}

    def _extract_headers_and_hierarchy(self, tool_registry, worksheet, region, merged_cells) -> tuple:
        """提取行表头、列表头、层级结构和元信息

        Returns:
            tuple: (row_headers, col_headers, hierarchy, meta_info)
        """
        row_headers = self._execute_tool_safely(tool_registry, "RowHeaderDetector", worksheet, region)
        col_headers = self._execute_tool_safely(tool_registry, "ColumnHeaderDetector", worksheet, region)
        hierarchy = self._execute_tool_safely(tool_registry, "HierarchicalHeaderParser", worksheet, merged_cells, region)
        meta_info = self._execute_tool_safely(tool_registry, "MetaInfoExtractor", row_headers, col_headers, hierarchy)

        return row_headers, col_headers, hierarchy, meta_info

    def _extract_entities_from_graph(self, graph) -> tuple:
        """从图中提取实体（只保留表头节点）

        Returns:
            tuple: (entities, entity_id_map)
        """
        entities = []
        entity_id_map = {}

        for node_id, node_data in graph.nodes(data=True):
            label = node_data.get("label", "")
            if label in ["column_header", "row_header"]:
                entity_id = node_id
                entity_id_map[node_id] = entity_id

                entity = {
                    "id": entity_id,
                    "label": label,
                    "properties": node_data.get("properties", {})
                }
                entities.append(entity)

        return entities, entity_id_map

    def _extract_relationships_from_graph(self, graph, entity_id_map: Dict[str, str]) -> List[Dict[str, Any]]:
        """从图中提取关系（只保留表头之间的has_child关系）"""
        relationships = []

        for u, v, data in graph.edges(data=True):
            u_label = graph.nodes[u].get("label", "")
            v_label = graph.nodes[v].get("label", "")

            if u_label in ["column_header", "row_header"] and v_label in ["column_header", "row_header"]:
                relation = data.get("relation", "")
                if relation == "has_child":
                    u_entity_id = entity_id_map.get(u)
                    v_entity_id = entity_id_map.get(v)

                    if u_entity_id and v_entity_id:
                        relationships.append({
                            "start_entity_id": u_entity_id,
                            "end_entity_id": v_entity_id,
                            "relation": relation
                        })

        return relationships

    def _extract_header_names_from_entities(self, entities: List[Dict[str, Any]]) -> tuple:
        """从实体中提取列名和行表头名称

        Returns:
            tuple: (col_header_names, row_header_names)
        """
        col_header_names = []
        row_header_names = []

        for entity in entities:
            props = entity.get("properties", {})
            value = props.get("value") or props.get("text", "")
            if value and isinstance(value, str) and not value.startswith('[EMPTY_'):
                value_clean = self._clean_column_name(str(value))
                if entity.get("label") == "column_header":
                    if value_clean not in col_header_names:
                        col_header_names.append(value_clean)
                elif entity.get("label") == "row_header":
                    if value_clean not in row_header_names:
                        row_header_names.append(value_clean)

        return col_header_names, row_header_names

    def _fallback_extract_header_names(self, col_headers: Dict, row_headers: Dict) -> tuple:
        """从col_headers/row_headers中提取列名和行表头名称（作为后备方案）

        Returns:
            tuple: (col_header_names, row_header_names)
        """
        col_header_names = []
        row_header_names = []

        if col_headers.get("headers"):
            col_header_names = [
                self._clean_column_name(str(h)) for h in col_headers["headers"]
                if h and isinstance(h, str) and not str(h).startswith('[EMPTY_')
            ]

        if row_headers.get("headers"):
            row_header_names = [
                self._clean_column_name(str(h)) for h in row_headers["headers"]
                if h and isinstance(h, str) and not str(h).startswith('[EMPTY_')
            ]

        return col_header_names, row_header_names

    def _build_simplified_meta_graph(
        self,
        col_header_names: List[str],
        row_header_names: List[str],
        worksheet,
        hierarchy: Dict,
        meta_info: Dict,
        entities: List[Dict],
        relationships: List[Dict]
    ) -> Dict[str, Any]:
        """构建简化的meta graph"""
        return {
            "meta_info": {
                "col_header_names": col_header_names,
                "row_header_names": row_header_names,
                "num_rows": worksheet.max_row - hierarchy.get("col_header_end_row", 1),
                "hierarchy_levels": hierarchy.get("max_level", 1),
                "table_structure": meta_info.get("table_structure", {})
            },
            "hierarchy": hierarchy,
            "entities": entities,
            "relationships": relationships
        }

    def _generate_hierarchy_info(self, excel_path: str, table_name: str) -> Optional[Dict[str, Any]]:
        """只提取层级信息，不生成完整的 graph

        Args:
            excel_path: Excel 文件路径
            table_name: 表格名称（不带后缀）

        Returns:
            简化的 meta graph 字典（只包含层级信息），如果失败返回 None
        """
        if not self.auto_generate:
            return None

        if not os.path.exists(excel_path):
            logger.warning(f"Excel file not found for hierarchy extraction: {excel_path}")
            return None

        try:
            logger.info(f"Extracting hierarchy info for {table_name} from {excel_path}")

            from tools.tools import tool_registry

            # 1. 加载Excel文件
            worksheet = self._load_excel_worksheet(excel_path)
            if not worksheet:
                return None

            # 2. 创建区域和提取合并单元格
            region = self._create_region_from_worksheet(worksheet)
            merged_cells = self._extract_merged_cells(worksheet)

            # 3. 提取表头和层级信息
            row_headers, col_headers, hierarchy, meta_info = self._extract_headers_and_hierarchy(
                tool_registry, worksheet, region, merged_cells
            )

            # 4. 使用 GraphBuilder 构建 graph
            graph_builder = tool_registry.get("GraphBuilder")
            if not graph_builder:
                logger.error("GraphBuilder not found in tool registry")
                return None

            try:
                graph = graph_builder.execute(worksheet, hierarchy, meta_info, region, table_name)

                # 5. 从 graph 中提取 entities 和 relationships
                entities, entity_id_map = self._extract_entities_from_graph(graph)
                relationships = self._extract_relationships_from_graph(graph, entity_id_map)

                # 6. 提取列名和行表头名称
                col_header_names, row_header_names = self._extract_header_names_from_entities(entities)

                # 7. 如果没有从entities中提取到，使用后备方案
                if not col_header_names or not row_header_names:
                    fallback_col_names, fallback_row_names = self._fallback_extract_header_names(
                        col_headers, row_headers
                    )
                    col_header_names = col_header_names or fallback_col_names
                    row_header_names = row_header_names or fallback_row_names

                # 8. 构建简化的 meta graph
                simplified_meta_graph = self._build_simplified_meta_graph(
                    col_header_names, row_header_names, worksheet, hierarchy,
                    meta_info, entities, relationships
                )

                logger.info(f"Successfully extracted hierarchy info for {table_name}: {len(col_header_names)} columns, "
                           f"{len(entities)} entities, {len(relationships)} relationships")
                return simplified_meta_graph

            except Exception as e:
                logger.error(f"GraphBuilder failed: {e}")
                import traceback
                logger.debug(traceback.format_exc())
                return None
        except Exception as e:
            logger.error(f"Error extracting hierarchy info: {e}")
            import traceback
            logger.debug(traceback.format_exc())
        return None
    
    def get_table_info(self, table_name: str, excel_path: Optional[str] = None) -> Dict[str, Any]:
        """获取表格信息（从 Meta Graph，如果不存在则自动生成）
        
        Args:
            table_name: 表格名称（不带后缀）
            excel_path: Excel 文件路径（用于自动生成 meta graph，可选）
            
        Returns:
            符合 ADG 要求的 table_info 字典
        """
        meta_graph = self.schema_linking.get_meta_graph(table_name)
        
        # 如果找不到 meta graph 且启用了自动生成，尝试提取层级信息
        if not meta_graph and self.auto_generate and excel_path:
            meta_graph = self._generate_hierarchy_info(excel_path, table_name)
        
        if not meta_graph:
            logger.warning(f"Meta graph not found for {table_name}, returning empty info")
            return {
                "file_name": f"{table_name}.xlsx",
                "column_names": [],
                "column_types": {},
                "row_count": 0,
                "summary_text": ""
            }
            
        meta_info = meta_graph.get("meta_info", {})
        entities = meta_graph.get("entities", [])
        
        # 1. 提取列名 (清洗后)
        col_header_names = meta_info.get("col_header_names", [])
        column_names = []
        for name in col_header_names:
            if name and isinstance(name, str) and not name.startswith('[EMPTY_'):
                column_names.append(self._clean_column_name(name))
        
        if not column_names:
            for ent in entities:
                if ent.get("label") == "column_header":
                    props = ent.get("properties", {})
                    val = props.get("value") or props.get("text")
                    if val and isinstance(val, str) and not val.startswith('[EMPTY_'):
                        column_names.append(self._clean_column_name(val))
                        
        # 2. 提取行数
        row_count = meta_info.get("num_rows")
        if row_count is None:
            row_header_names = meta_info.get("row_header_names", [])
            row_count = len(row_header_names)
            
        column_types = {col: "object" for col in column_names}
        
        # 3. 提取丰富的三元组 (替换原有的逻辑)
        triplets = self._extract_rich_triplets(meta_graph)
        
        # 4. 生成增强的摘要
        summary_lines = [f"表格包含 {len(column_names)} 列和 {row_count} 行数据。"]
        
        # 添加层级结构描述到摘要
        child_rels = [t for t in triplets if "has_child" in t]
        if child_rels:
            summary_lines.append("表格具有层级表头结构：")
            for t in child_rels[:10]:
                summary_lines.append(f"- {t}")
        
        summary_text = "\n".join(summary_lines)
        
        # 5. 构造 table_info
        table_info = {
            "file_name": f"{table_name}.xlsx",
            "column_names": column_names,
            "column_types": column_types,
            "row_count": row_count,
            "summary_text": summary_text,
            "meta_graph_triplets": triplets
        }
        
        logger.info(f"Loaded info for {table_name}: {len(column_names)} cols, {row_count} rows")
        return table_info

def main():
    """测试函数"""
    try:
        loader = SchemaLoader()
        meta_files = list(loader.meta_graphs_dir.glob("*.json"))
        if meta_files:
            test_table = meta_files[0].stem
            info = loader.get_table_info(test_table)
            print(f"\nTable: {test_table}")
            print(f"Columns: {info['column_names'][:5]}...")
            print(f"Rows: {info['row_count']}")
            print(f"Summary: {info['summary_text']}")
            print(f"Triplets: {info['meta_graph_triplets'][:5]}")
        else:
            print("No meta graph files found for testing.")
    except Exception as e:
        print(f"Test failed: {e}")

if __name__ == "__main__":
    main()
